"""의정부동 424 실무 공내역서 5공종 → 표준항목 마스터 JSON 추출.

출처: 의정부동 424 주상복합 신축공사 공내역서(실적 1건, 연면적 238,504㎡ —
건축 내역서 '임시동력+가설전기시설 연면적기준 M2 238504' 행에서 확인).
원칙: 원본 수치 무가공(수량 샘플 보존), 파생계수는 소비측에서 산출,
출처(provenance)를 JSON에 명시. 단가 칸은 공내역서라 비어 있음(전기 비고단가만 참고 보존).

실행: .venv/bin/python scripts/extract_boq_master.py
출력: app/services/cost/data/boq_master/{discipline}.json + _meta.json
"""
from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

import openpyxl
import xlrd

BASE = Path("/mnt/d/플랫폼제작/개발사업AI전주기자동화시스템/플랫폼_자료/적산_공내역서")
OUT = Path(__file__).resolve().parents[1] / "app" / "services" / "cost" / "data" / "boq_master"

PROJECT = {
    "name": "의정부동 424 주상복합 신축공사",
    "gfa_sqm": 238504.0,
    "gfa_basis": "건축 내역서 '임시동력+가설전기시설(연면적기준)' 수량",
    "sample_count": 1,
    "provenance": "실적 공내역서 1건(공개 단가 없음) — 원단위 계수는 n=1 참고치",
}

_SECTION_CODE = re.compile(r"^(\d{2,10})\s+(\S.*)$")
_SKIP_NAMES = ("합계", "소계", "총계", "[ 계 ]", "계 (", "순공사비", "공급가액", "부가가치세")


def _s(v: Any) -> str:
    return "" if v is None else str(v).strip()


def _num(v: Any) -> float | None:
    if v is None or isinstance(v, bool):
        return None
    try:
        return float(str(v).replace(",", ""))
    except ValueError:
        return None


def _is_skip(name: str) -> bool:
    return any(k in name for k in _SKIP_NAMES)


class Collector:
    """(품명,규격,단위) 중복 합산 수집기 — 결정론 정렬 출력."""

    def __init__(self) -> None:
        self.items: dict[tuple[str, str, str], dict[str, Any]] = {}
        self.sections: list[dict[str, Any]] = []
        self._seen_codes: set[str] = set()

    def add_section(self, code: str, name: str) -> None:
        if code and code not in self._seen_codes:
            self._seen_codes.add(code)
            self.sections.append({"code": code, "name": name, "level": max(1, len(code) // 2)})

    def add_item(self, section_code: str, section_name: str, name: str, spec: str,
                 unit: str, qty: float | None, ref_price: float | None = None) -> None:
        if not name or _is_skip(name) or not unit:
            return
        # 헤더행('품명'/'공종명', 단위='단위')·표지행('■…') 노이즈 제거
        compact = name.replace(" ", "")
        if unit == "단위" or compact in ("품명", "공종명", "공종") or name.startswith("■"):
            return
        key = (name, spec, unit)
        ent = self.items.get(key)
        if ent is None:
            ent = {
                "section_code": section_code, "section_name": section_name,
                "name": name, "spec": spec, "unit": unit,
                "qty_sample": 0.0, "row_count": 0,
            }
            self.items[key] = ent
        if qty is not None:
            ent["qty_sample"] = round(ent["qty_sample"] + qty, 4)
        ent["row_count"] += 1
        if ref_price is not None and "ref_mat_price" not in ent:
            ent["ref_mat_price"] = ref_price

    def dump(self, discipline: str, source_file: str) -> dict[str, Any]:
        items = sorted(self.items.values(), key=lambda x: (x["section_code"], x["name"], x["spec"]))
        for i, it in enumerate(items, 1):
            it["id"] = f"{discipline[:2]}-{i:04d}"
        return {
            "discipline": discipline,
            "source_file": source_file,
            "project": PROJECT,
            "sections": self.sections,
            "items": items,
        }


def extract_standard(path: Path, sheet: str, discipline: str,
                     ref_price_col: int | None = None) -> dict[str, Any]:
    """건축·기계·전기 공통형: 품명|규격|단위|수량 + 행내 다중 섹션 헤더."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    col = Collector()
    cur_code, cur_name = "", ""
    for row in ws.iter_rows(min_row=1, values_only=True):
        name, spec, unit = _s(row[0]), _s(row[1]) if len(row) > 1 else "", _s(row[2]) if len(row) > 2 else ""
        qty = _num(row[3]) if len(row) > 3 else None
        # 행 내 모든 셀에서 섹션 헤더 탐지(기계·전기: 한 행에 다단계 코드)
        found = False
        for cell in row[:6]:
            m = _SECTION_CODE.match(_s(cell))
            if m and not unit:
                cur_code, cur_name = m.group(1), m.group(2).strip()
                col.add_section(cur_code, cur_name)
                found = True
        if found:
            continue
        if unit and name:
            ref = _num(row[ref_price_col]) if ref_price_col and len(row) > ref_price_col else None
            col.add_item(cur_code, cur_name, name, spec, unit, qty, ref)
    wb.close()
    return col.dump(discipline, path.name)


def extract_landscape(path: Path) -> dict[str, Any]:
    """조경: 내역서 시트 — 공종|규격|수량|단위(수량·단위 역순)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["내역서"]
    col = Collector()
    cur_code, cur_name = "", ""
    sec_pat = re.compile(r"^(\d{2})\.\s*(\S.*)$")  # '01. 식 재 공' — 단위 '식'이어도 섹션
    for row in ws.iter_rows(min_row=1, values_only=True):
        name = _s(row[0])
        spec = _s(row[1]) if len(row) > 1 else ""
        qty = _num(row[2]) if len(row) > 2 else None
        unit = _s(row[3]) if len(row) > 3 else ""
        m = sec_pat.match(name)
        if m:
            cur_code, cur_name = f"L{m.group(1)}", m.group(2).strip()
            col.add_section(cur_code, cur_name)
            continue
        if unit and name:
            col.add_item(cur_code, cur_name, name, spec, unit, qty)
    wb.close()
    return col.dump("조경", path.name)


def extract_civil(path: Path) -> dict[str, Any]:
    """토목(.xls): 공종|규격|수량|단위 — Ⅰ./1./1) 텍스트 계층 섹션."""
    wb = xlrd.open_workbook(str(path), on_demand=True)
    sh = wb.sheet_by_name("내역서")
    col = Collector()
    cur_code, cur_name, sec_n = "", "", 0
    sec_pat = re.compile(r"^([ⅠⅡⅢⅣⅤⅥⅦⅧⅨⅩ]\.|\d{1,2}[.)]|\*\.)\s*\S")
    for r in range(sh.nrows):
        name = _s(sh.cell_value(r, 0))
        spec = _s(sh.cell_value(r, 1)) if sh.ncols > 1 else ""
        qty = _num(sh.cell_value(r, 2)) if sh.ncols > 2 else None
        unit = _s(sh.cell_value(r, 3)) if sh.ncols > 3 else ""
        if name and not unit and sec_pat.match(name) and not _is_skip(name):
            sec_n += 1
            cur_code, cur_name = f"C{sec_n:02d}", name
            col.add_section(cur_code, cur_name)
            continue
        if unit and name and qty:
            col.add_item(cur_code, cur_name, name, spec, unit, qty)
    return col.dump("토목", path.name)


def main(argv: list[str] | None = None) -> None:
    # N1: 프로젝트별 누적 구조 수용(additive). --project 미지정 시 기존 단일 플랫 구조 유지.
    import argparse  # noqa: PLC0415

    parser = argparse.ArgumentParser(description="실무 공내역서 5공종 → 표준항목 마스터 추출")
    parser.add_argument(
        "--project", default=None,
        help="프로젝트 식별자 — 지정 시 data/boq_master/<project>/ 하위에 누적 저장"
        "(N1 표본 통계용). 미지정 시 기존 평면 구조(default)로 저장(하위호환).")
    args = parser.parse_args(argv)

    out_dir = (OUT / args.project) if args.project else OUT
    out_dir.mkdir(parents=True, exist_ok=True)
    results = {
        "건축": extract_standard(
            BASE / "건축" / "의정부동 424 주상복합 신축공사_건축_공내역서.xlsx",
            "공종별내역서", "건축"),
        "기계소방": extract_standard(
            BASE / "기계,기계소방" / "의정부동 424 주상복합 신축공사_기계,소방_내역서.xlsm",
            "내역서", "기계소방"),
        "전기통신소방": extract_standard(
            BASE / "전기,전기소방,통신" / "의정부동 424 주상복합 신축공사_전기,통신,소방_공내역서.xlsm",
            "내역서", "전기통신소방", ref_price_col=13),
        "조경": extract_landscape(
            BASE / "조경" / "의정부동 424 주상복합 신축공사_조경_공내역서.xlsx"),
        "토목": extract_civil(
            BASE / "토목" / "의정부동 424 주상복합 신축공사_토목_공내역서.xls"),
    }
    meta = {"project": PROJECT, "disciplines": {}}
    fname = {"건축": "architecture", "기계소방": "mechanical", "전기통신소방": "electrical",
             "조경": "landscape", "토목": "civil"}
    for disc, data in results.items():
        out = out_dir / f"{fname[disc]}.json"
        out.write_text(json.dumps(data, ensure_ascii=False, indent=1), encoding="utf-8")
        meta["disciplines"][disc] = {
            "file": out.name, "sections": len(data["sections"]),
            "unique_items": len(data["items"]),
            "rows_aggregated": sum(i["row_count"] for i in data["items"]),
        }
        print(f"{disc}: 섹션 {len(data['sections'])} · 고유항목 {len(data['items'])} "
              f"· 원천행 {meta['disciplines'][disc]['rows_aggregated']}")
    (out_dir / "_meta.json").write_text(json.dumps(meta, ensure_ascii=False, indent=1), encoding="utf-8")
    print("저장 →", out_dir)


if __name__ == "__main__":
    main()
