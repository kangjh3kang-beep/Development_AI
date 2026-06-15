"""수지분석 v2 API 라우터 — 계산·비교·몬테카를로(실수지)·민감도·VCS·내보내기."""

from __future__ import annotations

import logging
import uuid
from datetime import datetime
from typing import Any

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession

from app.schemas.feasibility_v2 import (
    FeasibilityCalculateRequest,
    FeasibilityBaselineRequest,
    FeasibilityBaselineResponse,
    FeasibilityMultiRequest,
    FeasibilityResultResponse,
    FeasibilityMultiResponse,
    MonteCarloRequest,
    MonteCarloResponse,
    OptimizationRequest,
    RecommendationResponse,
    ModuleListResponse,
    SensitivityRequest,
    SensitivityResponse,
    VCSCommitRequest,
    VCSRollbackRequest,
)
from app.services.feasibility.feasibility_service_v2 import FeasibilityServiceV2
from app.services.feasibility.modules.base_module import ModuleInput
from app.services.feasibility.monte_carlo_engine import run_monte_carlo, MCVariable
from app.services.feasibility.ai_optimizer import optimize_slsqp
from app.services.feasibility.ai_recommendation import diagnose
from app.services.feasibility.version_control_db import FeasibilityVCSDB
from app.services.feasibility.sensitivity_engine import (
    SensitivityScenario,
    run_sensitivity_analysis,
)
from app.core.database import get_db
from app.core.billing_deps import enforce_llm_quota
from app.services.auth.auth_service import get_current_user
# 주의(혼선 제거): 아래 current_user의 실제 런타임 타입은 get_current_user가 반환하는
# apps.api.database.models.user.User(tenant_id 보유)다. 이 app.models.auth.User는
# stale(organization_id) 모델로 타입 힌트 표기용일 뿐 — current_user.tenant_id 접근은 런타임 정상.
# (from __future__ import annotations 활성 → 주석/힌트는 런타임 미평가, 동작 무영향.)
from app.models.auth import User

router = APIRouter(prefix="/api/v2/feasibility", tags=["feasibility-v2"])

logger = logging.getLogger(__name__)

_service = FeasibilityServiceV2()


def _parse_project_id(project_id: str) -> uuid.UUID:
    """project_id 문자열을 UUID로 변환. 'default' 등 비UUID 문자열은 결정적 UUID로 매핑."""
    try:
        return uuid.UUID(project_id)
    except ValueError:
        return uuid.uuid5(uuid.NAMESPACE_DNS, f"propai.feasibility.{project_id}")


# 시도 약식명 → 정식 행정구역명(VWorld 지오코딩 성공률↑). 17개 광역시도.
_SIDO_SHORT_TO_FULL = {
    "서울": "서울특별시",
    "부산": "부산광역시",
    "대구": "대구광역시",
    "인천": "인천광역시",
    "광주": "광주광역시",
    "대전": "대전광역시",
    "울산": "울산광역시",
    "세종": "세종특별자치시",
    "경기": "경기도",
    "강원": "강원특별자치도",
    "충북": "충청북도",
    "충남": "충청남도",
    "전북": "전북특별자치도",
    "전남": "전라남도",
    "경북": "경상북도",
    "경남": "경상남도",
    "제주": "제주특별자치도",
}


def _normalize_address(address: str) -> str:
    """약식 시도명("서울")을 정식 행정구역명("서울특별시")으로 보강해 지오코딩 성공률을 높인다.

    이미 정식명(접미사 포함)이면 그대로 둔다. 시도명 접두만 보강하며, 구·동·번지가
    누락된 경우는 보강할 수 없으므로 입력을 그대로 반환한다(무목업: 추정 생성 금지).
    """
    addr = (address or "").strip()
    if not addr:
        return addr
    for short, full in _SIDO_SHORT_TO_FULL.items():
        if addr.startswith(full):
            return addr  # 이미 정식명
        if addr.startswith(short):
            rest = addr[len(short):].lstrip()
            # "서울특별시"처럼 정식 접미사가 이미 붙은 경우는 위 분기에서 처리됨.
            # "서울 강남구..." / "서울강남구..." → "서울특별시 강남구..."
            return f"{full} {rest}" if rest else full
    return addr


def _request_to_input(req: FeasibilityCalculateRequest) -> ModuleInput:
    return ModuleInput(
        development_type=req.development_type,
        project_name=req.project_name,
        total_land_area_sqm=req.total_land_area_sqm,
        land_category=req.land_category,
        official_price_per_sqm=req.official_price_per_sqm,
        price_multiplier=req.price_multiplier,
        total_gfa_sqm=req.total_gfa_sqm,
        building_type=req.building_type,
        total_households=req.total_households,
        avg_sale_price_per_pyeong=req.avg_sale_price_per_pyeong,
        avg_area_pyeong=req.avg_area_pyeong,
        sale_ratio=req.sale_ratio,
        bridge_amount_won=req.bridge_amount_won,
        pf_amount_won=req.pf_amount_won,
        midpay_amount_won=req.midpay_amount_won,
        sido_name=req.sido_name,
        sigungu_name=req.sigungu_name,
        project_months=req.project_months,
        discount_rate=req.discount_rate,
        equity_won=req.equity_won,
        params=req.params,
    )


# ── 실수지 섭동(몬테카를로 base 모드 · /sensitivity 공용) ──────────
# 지원 변수명은 sensitivity_engine DEFAULT_SCENARIOS 관례를 따른다.
BASE_PERTURB_VARIABLES = (
    "sale_price",         # 평당 분양가(원/평) → avg_sale_price_per_pyeong
    "construction_cost",  # 총공사비(원) → params.construction_cost_override_won
    "land_cost",          # 총토지비(원) → 공시지가 비례 스케일
    "interest_rate",      # 대표금리(소수) → bridge/pf/midpay 3종 동률(pp) 적용
    "project_months",     # 사업기간(월)
)

# 실수지 모드 시뮬레이션 횟수 상한 — 표본마다 수지모듈 전체(공사비·세금·금융)를
# 재계산하므로 동기 API 응답시간 보호를 위해 제한(파이프라인 MC도 1,000회 관례).
MC_BASE_MAX_SIMULATIONS = 1_000


def _make_base_perturb_fn(base_req: FeasibilityCalculateRequest):
    """실수지 섭동 함수 팩토리 — 변수 dict를 base 입력에 반영해 v2 엔진으로 재계산.

    Returns:
        (perturb_fn, base_output, base_values)
        - perturb_fn(vals: dict[str, float]) -> ModuleOutput
        - base_values: 지원 변수 5종의 섭동 원점(실수지 산출 기준값)

    Raises:
        ValueError: base 입력 자체가 계산 불가(모듈 검증 실패)할 때.
    """
    base_inp = _request_to_input(base_req)
    base_out = _service.calculate(base_inp)  # 입력 검증 겸 기준 토지비·공사비 확보
    base_land = float(base_out.total_land_cost_won)
    base_constr = float(base_out.total_construction_cost_won)

    base_values: dict[str, float] = {
        "sale_price": float(base_req.avg_sale_price_per_pyeong),
        "construction_cost": base_constr,
        "land_cost": base_land,
        "interest_rate": float(base_inp.pf_rate),  # ModuleInput 표준 PF금리 기준
        "project_months": float(base_req.project_months),
    }

    def perturb_fn(vals: dict[str, float]):
        inp = _request_to_input(base_req)
        inp.params = dict(base_req.params or {})  # 표본 간 params 공유 오염 방지
        for name, value in vals.items():
            v = float(value)
            if name == "sale_price":
                inp.avg_sale_price_per_pyeong = max(0.0, v)  # 음수 분양가 불가 — 0 하한
            elif name == "construction_cost":
                # 기존 정밀공사비 주입 라인(construction_cost_override_won) 재사용.
                # 1원 하한: 0 이하면 override가 무시되어 표준단가 계산으로 암묵
                # 전환(모델 스위칭)되는 것을 방지.
                inp.params["construction_cost_override_won"] = max(1.0, v)
            elif name == "land_cost":
                # 토지비는 공시지가에 선형 → 공시지가 비례 스케일(params 보상비 불변).
                # 기준 토지비/공시지가가 0이면 섭동 불능 — 무변경(no-op) 정직 처리.
                if base_land > 0 and base_inp.official_price_per_sqm > 0:
                    inp.official_price_per_sqm = (
                        base_inp.official_price_per_sqm * max(0.0, v) / base_land
                    )
            elif name == "interest_rate":
                # 단일 대표금리 가정 — pp 변동분을 3종 대출금리에 동률 적용.
                shift = v - float(base_inp.pf_rate)
                inp.bridge_rate = max(0.0, base_inp.bridge_rate + shift)
                inp.pf_rate = max(0.0, base_inp.pf_rate + shift)
                inp.midpay_rate = max(0.0, base_inp.midpay_rate + shift)
            elif name == "project_months":
                inp.project_months = max(1, int(round(v)))
            else:
                raise ValueError(
                    f"실수지 섭동 미지원 변수 '{name}' — "
                    f"지원: {', '.join(BASE_PERTURB_VARIABLES)}"
                )
        return _service.calculate(inp)

    return perturb_fn, base_out, base_values


# ─────────────────────────────────────────────────────────────────────────────
# 신뢰 레이어(additive) — 총사업비 구성 근거 트레이스(evidence[]) + 법령 근거(legal_refs[]).
# auto_zoning.py의 _build_legal_refs/_build_evidence/_attach_trust_blocks 정본 패턴 준수:
# 기존 응답 필드는 1개도 변경하지 않고 두 블록만 가산하며, 구축 중 예외가 나면 빈
# 배열로 강등한다(graceful — 수지 결과 무손상). law.go.kr URL은 반드시
# legal_reference_registry.get_legal_refs 출력만 사용하고(여기서 URL 조립 금지),
# 레지스트리에 없는 근거는 링크 없이 텍스트만(할루시네이션 링크 금지).
# ─────────────────────────────────────────────────────────────────────────────


class FeasibilityResultTrustResponse(FeasibilityResultResponse):
    """/calculate 응답 + 신뢰 블록(additive — 기존 필드·타입 불변, 하위호환).

    evidence[]   : 총사업비 구성(토지비·공사비·금융비·세금합계) 산출 트레이스
                   {label, value, basis, legal_ref_key?}
    legal_refs[] : 실제 부과된 세목의 법령 근거(레지스트리 get_legal_refs 출력).
                   far_limit 등 설계한도 키는 수지 산출 근거가 아니므로 미포함.
    """

    evidence: list[dict[str, Any]] = []
    legal_refs: list[dict[str, Any]] = []


# 통합 세금엔진 항목코드 → 법령 근거 레지스트리 키. 레지스트리 보유분만 매핑하며,
# 그 외 세목(농어촌특별세·각종 부담금 등)은 링크 없이 합계 텍스트로만 표기한다.
_TAX_CODE_TO_REF_KEY: dict[str, str] = {
    "A01": "acquisition_tax",             # 취득세 — 지방세법 제11조
    "A02": "local_education_tax",         # 지방교육세 — 지방세법(루트)
    "A04": "stamp_tax",                   # 인지세 — 인지세법 제3조
    "D01": "capital_gains_tax",           # 양도소득세 — 소득세법 제104조
    "D05": "reconstruction_levy",         # 재건축부담금 — 재건축초과이익 환수법(루트)
    "D06": "comprehensive_property_tax",  # 종합부동산세 — 종합부동산세법(루트)
}


def _fmt_won(v) -> str:
    """원화 표기 — 1234567 → '1,234,567원'. 비수치는 '—'(정직 표기)."""
    try:
        return f"{int(round(float(v))):,}원"
    except (TypeError, ValueError):
        return "—"


def _applied_taxes(tax_detail) -> list[dict[str, Any]]:
    """tax_detail(통합 세금엔진 출력) 4단계에서 실제 부과(amount_won>0)된 세목 평탄화.

    각 항목: {code, name, amount_won, ref_key}. ref_key는 레지스트리 보유분만
    부착(그 외 None — 링크 없이 텍스트만). 구조가 다르면 빈 리스트(graceful).
    """
    out: list[dict[str, Any]] = []
    if not isinstance(tax_detail, dict):
        return out
    for stage in ("acquisition", "construction", "sale", "disposal"):
        st = tax_detail.get(stage)
        items = st.get("items") if isinstance(st, dict) else None
        for it in items or []:
            if not isinstance(it, dict):
                continue
            try:
                amount = float(it.get("amount_won") or 0)
            except (TypeError, ValueError):
                continue
            if amount <= 0:
                continue
            code = str(it.get("code") or "")
            out.append({
                "code": code,
                "name": str(it.get("name") or code or "세금항목"),
                "amount_won": int(round(amount)),
                "ref_key": _TAX_CODE_TO_REF_KEY.get(code),
            })
    return out


def _build_cost_trust_blocks(
    output, req: FeasibilityCalculateRequest
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """ModuleOutput → (evidence[], legal_refs[]) — 총사업비 구성 근거 트레이스.

    총사업비 = 토지비+공사비+금융비+기타경비+세금합계(aggregation_engine 합산식과
    동일 구성)를 구성요소별 한 줄 근거로 트레이스하고, 실제 부과된 세목 중
    레지스트리 보유분에만 법령 근거(legal_ref_key)를 연결한다. basis의 법령 표기는
    레지스트리 레코드(law_name·article)에서만 가져온다(단일출처).
    """
    taxes = _applied_taxes(getattr(output, "tax_detail", None))
    ref_keys: list[str] = []
    for t in taxes:
        rk = t.get("ref_key")
        if rk and rk not in ref_keys:
            ref_keys.append(rk)

    legal_refs: list[dict[str, Any]] = []
    if ref_keys:
        try:
            from app.services.legal.legal_reference_registry import get_legal_refs

            legal_refs = get_legal_refs(ref_keys)
        except Exception:  # noqa: BLE001 — 레지스트리 실패 시 링크 없이 텍스트만
            legal_refs = []
    ref_by_key = {r.get("key"): r for r in legal_refs}

    evidence: list[dict[str, Any]] = []
    params = req.params or {}

    # 총사업비 — 합산식(구성요소 전체)을 헤드라인으로 명시.
    evidence.append({
        "label": "총사업비",
        "value": _fmt_won(output.total_cost_won),
        "basis": "토지비 + 공사비 + 금융비 + 기타경비 + 세금 합계",
    })

    # 토지비 — 공시지가×면적×보정(+보상비). 취득세 등 제세는 세금 합계에 계상(이중계상 방지).
    if output.total_land_cost_won:
        basis = (
            f"공시지가 {_fmt_won(req.official_price_per_sqm)}/㎡ × "
            f"면적 {req.total_land_area_sqm:,.0f}㎡ × 보정 {req.price_multiplier:g}"
        )
        try:
            comp = float(params.get("compensation_won") or 0)
        except (TypeError, ValueError):
            comp = 0.0
        if comp > 0:
            basis += f" + 보상비 {_fmt_won(comp)}"
        evidence.append({
            "label": "토지비",
            "value": _fmt_won(output.total_land_cost_won),
            "basis": basis + " — 취득세 등 제세는 세금 합계에 계상",
        })

    # 공사비 — 정밀분석 주입값(override) 여부를 정직하게 구분 표기.
    if output.total_construction_cost_won:
        try:
            override_won = float(params.get("construction_cost_override_won") or 0)
        except (TypeError, ValueError):
            override_won = 0.0
        if override_won > 0:
            constr_basis = "공사비 정밀분석 결과 주입(construction_cost_override_won)"
        else:
            constr_basis = (
                f"연면적 {req.total_gfa_sqm:,.0f}㎡ × 표준단가({req.building_type})"
            )
        evidence.append({
            "label": "공사비",
            "value": _fmt_won(output.total_construction_cost_won),
            "basis": constr_basis,
        })

    # 금융비 — 브릿지·PF·중도금(finance_cost_engine 합산).
    if output.total_finance_cost_won:
        evidence.append({
            "label": "금융비",
            "value": _fmt_won(output.total_finance_cost_won),
            "basis": "브릿지·PF·중도금 대출이자·수수료 합산",
        })

    # 기타경비 — 입력 params(마케팅·관리·예비비) 합산.
    if output.total_other_cost_won:
        evidence.append({
            "label": "기타경비",
            "value": _fmt_won(output.total_other_cost_won),
            "basis": "마케팅·관리·예비비 합산(입력 params)",
        })

    # 세금 합계 — 통합 세금엔진 4단계 합산(부과 건수 정직 표기).
    if output.total_tax_cost_won:
        evidence.append({
            "label": "세금 합계",
            "value": _fmt_won(output.total_tax_cost_won),
            "basis": f"취득·공사·분양·양도 4단계 제세공과 합산({len(taxes)}건 부과)",
        })

    # 개별 세목 — 레지스트리 보유분만(법령 근거 연결). basis는 레지스트리 단일출처.
    for t in taxes:
        rk = t.get("ref_key")
        if not rk:
            continue
        ref = ref_by_key.get(rk) or {}
        law = str(ref.get("law_name") or "").strip()
        art = str(ref.get("article") or "").strip()
        basis = f"{law} {art}".strip()
        evidence.append({
            "label": f"세금 — {t['name']}",
            "value": _fmt_won(t["amount_won"]),
            "basis": basis or None,
            "legal_ref_key": rk,
        })

    return evidence, legal_refs


@router.post(
    "/calculate",
    response_model=FeasibilityResultTrustResponse,
    dependencies=[Depends(enforce_llm_quota)],
)
async def calculate_feasibility(req: FeasibilityCalculateRequest):
    """단일 수지분석 계산(+ 산출 근거 evidence[]·legal_refs[] 가산 — 기존 필드 불변)."""
    try:
        inp = _request_to_input(req)
        output = _service.calculate(inp)
        # 신뢰 블록(additive) — 구축 실패해도 수지 결과는 무손상(빈 배열 폴백).
        try:
            evidence, legal_refs = _build_cost_trust_blocks(output, req)
        except Exception as e:  # noqa: BLE001
            logger.warning("수지 근거 블록 부착 스킵: %s", str(e)[:120])
            evidence, legal_refs = [], []
        return FeasibilityResultTrustResponse(
            development_type=output.development_type,
            module_name=output.module_name,
            total_revenue_won=output.total_revenue_won,
            total_cost_won=output.total_cost_won,
            net_profit_won=output.net_profit_won,
            profit_rate_pct=output.profit_rate_pct,
            roi_pct=output.roi_pct,
            npv_won=output.npv_won,
            grade=output.grade,
            cost_breakdown_won=output.cost_detail,
            tax_detail=output.tax_detail,
            special_detail=output.special_detail,
            evidence=evidence,
            legal_refs=legal_refs,
        )
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))


@router.post(
    "/baseline",
    response_model=FeasibilityBaselineResponse,
    dependencies=[Depends(enforce_llm_quota)],
)
async def baseline_feasibility(req: FeasibilityBaselineRequest):
    """부지직후 시장표준 baseline 수지(설계·공사비 전 1차 산출).

    부지 데이터(주소/용도지역/면적)만으로 GFA 역산·시장표준 분양가·표준 공사비를
    시드해 기존 FeasibilityServiceV2 계산엔진으로 1차 수지를 산출한다.
    응답은 /calculate 와 동일 구조 + baseline 메타필드(is_baseline/confidence/
    sources/assumptions). 추정값은 정직하게 "시장표준/추정" 라벨, 실데이터 우선.
    """
    from app.services.feasibility.regional_pricing import (
        get_regional_sale_price_per_pyeong,
    )
    from app.services.feasibility.permit_validator import (
        get_permitted_types,
        DEVELOPMENT_TYPE_NAMES,
    )

    # ── 1) 부지 데이터 확보(실데이터 우선) ──
    # 프론트는 용도지역을 zone_type(이름) 또는 zone_code(이름/코드)로 보낼 수 있다.
    # 둘 다 수용해 단일 zone 변수로 통합(이전엔 zone_code가 항상 폐기되어 FAR 역산 미반영).
    zone = ((req.zone_type or "").strip() or (req.zone_code or "").strip())
    land_area = req.land_area_sqm or 0
    official_price = req.official_price_per_sqm or 0
    zone_limits: dict = {}
    sources: dict[str, Any] = {}

    # 약식 시도명("서울")을 정식명("서울특별시")으로 보강해 지오코딩 성공률↑
    norm_address = _normalize_address(req.address)

    need_autodetect = (not zone) or land_area <= 0 or official_price <= 0
    if need_autodetect:
        try:
            from app.services.zoning.auto_zoning_service import AutoZoningService

            zoning = await AutoZoningService().analyze_by_address(norm_address)
            if not zone and zoning.get("zone_type"):
                zone = str(zoning["zone_type"]).strip()
                sources["zone_type"] = "자동감지(공공데이터)"
            if land_area <= 0 and zoning.get("land_area_sqm"):
                land_area = float(zoning["land_area_sqm"])
                sources["land_area_sqm"] = "자동감지(공공데이터)"
            if official_price <= 0 and zoning.get("official_price_per_sqm"):
                official_price = float(zoning["official_price_per_sqm"])
                sources["official_price_per_sqm"] = "자동감지(공시지가)"
            zone_limits = zoning.get("zone_limits") or {}
        except Exception:  # noqa: BLE001 — 자동감지 실패 시 입력/표준으로 진행
            logger.warning("baseline: 용도지역 자동감지 실패 — 입력/표준값으로 진행")

    # 자동감지로 zone_limits를 못 얻었으면(자동감지 미수행/실패), 사용자가 보낸
    # 용도지역명(zone)으로 법정 상한 FAR/BCR을 정적 테이블에서 보강 → FAR 역산 반영.
    if not zone_limits and zone:
        try:
            from app.services.zoning.auto_zoning_service import (
                AutoZoningService,
                ZONE_LIMITS,
            )

            zone_key = AutoZoningService()._normalize_zone_name(zone)
            static_limits = ZONE_LIMITS.get(zone_key)
            if static_limits:
                zone_limits = {
                    "max_bcr_pct": static_limits["max_bcr"],
                    "max_far_pct": static_limits["max_far"],
                    "max_height_m": static_limits.get("max_height_m"),
                    "zone_key": zone_key,
                    "legal_basis": "국토의 계획 및 이용에 관한 법률 제78조",
                }
        except Exception:  # noqa: BLE001 — 정적 보강 실패 시 유형 표준 FAR로 진행
            logger.warning("baseline: 용도지역 정적 FAR/BCR 보강 실패")

    if req.zone_type or req.zone_code:
        sources.setdefault("zone_type", "사용자입력")
    if req.land_area_sqm:
        sources.setdefault("land_area_sqm", "사용자입력")
    if req.official_price_per_sqm:
        sources.setdefault("official_price_per_sqm", "사용자입력")

    if land_area <= 0:
        raise HTTPException(
            status_code=422,
            detail=(
                "부지면적을 자동감지하지 못했습니다. "
                "정확한 주소(시·구·동·번지) 또는 부지면적(land_area_sqm)을 입력하세요."
            ),
        )

    assumptions: dict[str, Any] = {}
    confidence_penalty = 0

    # ── 2) 개발유형 선택(용도지역 대표유형) ──
    dev_type = (req.development_type or "").strip()
    if not dev_type:
        permitted = get_permitted_types(zone) if zone else []
        if not permitted:
            # 용도지역 미상/개발불가 → 가장 일반적 유형으로 보수적 baseline
            dev_type = "M06"
            assumptions["development_type"] = "용도지역 미상/개발제한 — 일반분양(M06) 표준 가정"
            confidence_penalty += 1
        else:
            # 대표성·인허가 용이성 위해 일반분양(M06) 우선, 없으면 첫 허용유형
            dev_type = "M06" if "M06" in permitted else permitted[0]
            assumptions["development_type"] = (
                f"{zone} 인허가 가능유형 중 대표({DEVELOPMENT_TYPE_NAMES.get(dev_type, dev_type)}) 자동선택"
            )
    else:
        sources["development_type"] = "사용자입력"

    # ── 3) GFA 역산(조례/법정 FAR·BCR) ──
    ordinance_far = zone_limits.get("ordinance_far_pct") or 0
    legal_far = zone_limits.get("max_far_pct") or 0
    type_typical_far = float(_service._get_type_typical_far(dev_type))
    if ordinance_far or legal_far:
        zone_far = float(ordinance_far or legal_far)
        applied_far = min(zone_far, type_typical_far)
        far_source = "조례" if ordinance_far else "법정"
        sources["far_pct"] = f"용도지역 {far_source} 상한"
    else:
        applied_far = type_typical_far
        sources["far_pct"] = "개발유형 표준(추정)"
        confidence_penalty += 1
    total_gfa = land_area * applied_far / 100.0
    assumptions["gfa_inversion"] = (
        f"GFA={round(total_gfa)}㎡ = 부지 {round(land_area)}㎡ × 용적률 {round(applied_far)}%(적용)"
    )
    assumptions["applied_far_pct"] = round(applied_far, 1)

    # BCR로 층수 가정(far/bcr 비율)
    ordinance_bcr = zone_limits.get("ordinance_bcr_pct") or 0
    legal_bcr = zone_limits.get("max_bcr_pct") or 0
    applied_bcr = float(ordinance_bcr or legal_bcr or 60)
    est_floors = max(1, round(applied_far / applied_bcr)) if applied_bcr else 0
    assumptions["estimated_floors"] = est_floors
    assumptions["applied_bcr_pct"] = round(applied_bcr, 1)

    # ── 4) 세대수·평형 가정 ──
    avg_unit_area_sqm = _service._get_type_avg_unit_area(dev_type)
    total_hh = max(1, int(total_gfa / avg_unit_area_sqm))
    avg_area_pyeong = avg_unit_area_sqm / 3.305785

    # ── 5) 분양가 시드(시장표준 시세 테이블) ──
    sale_price_per_pyeong = get_regional_sale_price_per_pyeong(
        dev_type=dev_type, region=req.region, address=req.address
    )
    sources["avg_sale_price_per_pyeong"] = "지역 시세 테이블(시장표준)"
    assumptions["avg_sale_price_per_pyeong"] = (
        f"평당 {round(sale_price_per_pyeong / 10000)}만원(지역×유형 시장표준 시세)"
    )

    # ── 6) 토지비·공시지가 시드 ──
    if official_price <= 0:
        official_price = 1_500_000  # 표준 폴백(공시지가 미상)
        sources["official_price_per_sqm"] = "표준 폴백(공시지가 미상)"
        assumptions["official_price_per_sqm"] = "공시지가 미상 — 표준 150만원/㎡ 가정"
        confidence_penalty += 1
    price_multiplier = 1.1  # 공시지가→실거래 보수적 보정
    assumptions["land_price_multiplier"] = price_multiplier

    # ── 7) 자기자본 가정(미입력 시 토지비 기반) ──
    equity = req.equity_won or int(official_price * price_multiplier * land_area)
    if not req.equity_won:
        assumptions["equity_won"] = "자기자본 미입력 — 토지비 추정액으로 가정"
    else:
        sources["equity_won"] = "사용자입력"

    # ── 8) 공사비 표준단가 라벨(엔진은 SSOT 표준단가 자동적용) ──
    building_type = _service._get_building_type(dev_type)
    sources["construction_unit_cost"] = "표준 개산단가(DEFAULT_DIRECT_COST_PER_SQM, SSOT)"
    assumptions["building_type"] = building_type

    # ── 9) 기존 계산엔진 재사용 ──
    inp = ModuleInput(
        development_type=dev_type,
        project_name="baseline",
        total_land_area_sqm=land_area,
        official_price_per_sqm=official_price,
        price_multiplier=price_multiplier,
        total_gfa_sqm=total_gfa,
        building_type=building_type,
        total_households=total_hh,
        avg_sale_price_per_pyeong=sale_price_per_pyeong,
        avg_area_pyeong=avg_area_pyeong,
        sale_ratio=0.95,
        sido_name=req.region,
        project_months=_service._get_type_project_months(dev_type),
        discount_rate=0.08,
        equity_won=equity,
    )
    try:
        output = _service.calculate(inp)
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))

    # ── 10) 신뢰도 판정(추정 데이터 비중) ──
    confidence = "보통" if confidence_penalty <= 1 else "낮음"

    return FeasibilityBaselineResponse(
        development_type=output.development_type,
        module_name=output.module_name,
        total_revenue_won=output.total_revenue_won,
        total_cost_won=output.total_cost_won,
        net_profit_won=output.net_profit_won,
        profit_rate_pct=output.profit_rate_pct,
        roi_pct=output.roi_pct,
        npv_won=output.npv_won,
        grade=output.grade,
        cost_breakdown_won=output.cost_detail,
        tax_detail=output.tax_detail,
        special_detail=output.special_detail,
        is_baseline=True,
        confidence=confidence,
        sources=sources,
        assumptions=assumptions,
    )


@router.post("/compare", response_model=FeasibilityMultiResponse)
async def compare_feasibility(req: FeasibilityMultiRequest):
    """복수 개발유형 비교 분석."""
    inputs = [_request_to_input(p) for p in req.projects]
    result = _service.calculate_multi(inputs)
    responses = []
    for output in result["results"]:
        responses.append(FeasibilityResultResponse(
            development_type=output.development_type,
            module_name=output.module_name,
            total_revenue_won=output.total_revenue_won,
            total_cost_won=output.total_cost_won,
            net_profit_won=output.net_profit_won,
            profit_rate_pct=output.profit_rate_pct,
            roi_pct=output.roi_pct,
            npv_won=output.npv_won,
            grade=output.grade,
        ))
    return FeasibilityMultiResponse(results=responses, comparison=result["comparison"])


@router.get("/modules", response_model=ModuleListResponse)
async def list_modules():
    """사용 가능한 개발유형 모듈 목록."""
    return ModuleListResponse(modules=_service.list_available_modules())


@router.post("/monte-carlo", response_model=MonteCarloResponse)
async def run_monte_carlo_sim(req: MonteCarloRequest):
    """몬테카를로 시뮬레이션.

    - base 미제공(기존 계약): 변수 합(simple_npv)을 목적함수로 사용 — 하위호환.
    - base 제공: 표본을 실수지 입력에 반영해 FeasibilityServiceV2.calculate로
      net_profit_won 분포를 산출. 지원 변수: sale_price/construction_cost/
      land_cost/interest_rate/project_months (그 외 변수명은 422).
    """
    mc_vars = [
        MCVariable(
            name=v["name"],
            mean=v["mean"],
            std=v["std"],
            distribution=v.get("distribution", "normal"),
        )
        for v in req.variables
    ]

    if req.base is None:
        # 기존 동작 유지(하위호환) — 변수 단순 합산
        def simple_npv(vals):
            return sum(vals.values())

        result = run_monte_carlo(
            calculate_fn=simple_npv,
            variables=mc_vars,
            n_simulations=req.n_simulations,
            seed=req.seed,
        )
        return MonteCarloResponse(**result)

    # ── 실수지 모드 — 미지원 변수는 가짜 반영 대신 정직하게 거부 ──
    unknown = [v.name for v in mc_vars if v.name not in BASE_PERTURB_VARIABLES]
    if unknown:
        raise HTTPException(
            status_code=422,
            detail=(
                f"실수지 모드 미지원 변수: {', '.join(unknown)} — "
                f"지원: {', '.join(BASE_PERTURB_VARIABLES)}"
            ),
        )

    try:
        perturb_fn, _base_out, _base_values = _make_base_perturb_fn(req.base)
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))

    def real_profit_fn(vals: dict[str, float]) -> float:
        return float(perturb_fn(vals).net_profit_won)

    n_sim = min(req.n_simulations, MC_BASE_MAX_SIMULATIONS)
    note = None
    if n_sim < req.n_simulations:
        note = (
            f"실수지 모드: 요청 {req.n_simulations:,}회를 상한 "
            f"{MC_BASE_MAX_SIMULATIONS:,}회로 제한(표본마다 수지모듈 전체 재계산)"
        )

    result = run_monte_carlo(
        calculate_fn=real_profit_fn,
        variables=mc_vars,
        n_simulations=n_sim,
        seed=req.seed,
    )
    return MonteCarloResponse(
        **result,
        target_metric="net_profit_won",
        calc_source="feasibility_v2",
        note=note,
    )


@router.post("/sensitivity", response_model=SensitivityResponse)
async def run_sensitivity(req: SensitivityRequest):
    """민감도 분석(토네이도) — 실수지(FeasibilityServiceV2) 섭동 기반.

    base 수지입력을 원점으로 시나리오 변수(분양가/공사비/토지비/금리/공기)를
    변동시키며 수지를 재계산한다. scenarios 미지정 시 엔진 프리셋 5종 사용.
    """
    custom_scenarios = None
    if req.scenarios:
        unknown = [
            s.variable for s in req.scenarios
            if s.variable not in BASE_PERTURB_VARIABLES
        ]
        if unknown:
            raise HTTPException(
                status_code=422,
                detail=(
                    f"민감도 미지원 변수: {', '.join(unknown)} — "
                    f"지원: {', '.join(BASE_PERTURB_VARIABLES)}"
                ),
            )
        custom_scenarios = [
            SensitivityScenario(
                name=s.name, variable=s.variable, deltas_pct=s.deltas_pct
            )
            for s in req.scenarios
        ]

    try:
        perturb_fn, _base_out, base_values = _make_base_perturb_fn(req.base)
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))

    def sensitivity_fn(vals: dict[str, float]) -> dict[str, Any]:
        out = perturb_fn(vals)
        return {
            "development_type": out.development_type,
            "total_revenue_won": out.total_revenue_won,
            "total_cost_won": out.total_cost_won,
            "net_profit_won": out.net_profit_won,
            "profit_rate_pct": out.profit_rate_pct,
            "roi_pct": out.roi_pct,
            "npv_won": out.npv_won,
            "grade": out.grade,
        }

    result = run_sensitivity_analysis(
        base_values=base_values,
        calculate_fn=sensitivity_fn,
        scenarios=custom_scenarios,
    )
    return SensitivityResponse(
        base_result=result["base_result"],
        scenarios=result["scenarios"],
        tornado=result["tornado"],
        base_values=base_values,
        calc_source="feasibility_v2",
    )


@router.post("/optimize")
async def run_optimization(req: OptimizationRequest):
    """SLSQP 최적화."""
    variables = {k: tuple(v) for k, v in req.variables.items()}

    def objective(vals):
        return sum(vals.values())

    result = optimize_slsqp(
        objective_fn=objective,
        variables=variables,
        maximize=True,
        max_iter=req.max_iter,
    )
    return result


@router.post("/recommendations", response_model=RecommendationResponse)
async def get_recommendations(req: FeasibilityCalculateRequest):
    """AI 권고 생성."""
    inp = _request_to_input(req)
    output = _service.calculate(inp)

    total_cost = output.total_cost_won or 1
    recs = diagnose(
        profit_rate_pct=output.profit_rate_pct,
        roi_pct=output.roi_pct,
        finance_cost_ratio_pct=output.total_finance_cost_won / total_cost * 100,
        construction_cost_ratio_pct=output.total_construction_cost_won / total_cost * 100,
        tax_cost_ratio_pct=output.total_tax_cost_won / total_cost * 100,
        grade=output.grade,
    )
    return RecommendationResponse(
        recommendations=[
            {
                "rule_code": r.rule_code,
                "rule_name": r.rule_name,
                "severity": r.severity,
                "message": r.message,
                "suggestion": r.suggestion,
            }
            for r in recs
        ]
    )


@router.post("/repos/{project_id}/commit")
async def vcs_commit(
    project_id: str,
    req: VCSCommitRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """수지분석 커밋."""
    vcs = FeasibilityVCSDB(db, project_id=_parse_project_id(project_id), tenant_id=current_user.tenant_id)
    result = await vcs.commit(req.snapshot, req.message)
    await db.commit()                      # 외부 트랜잭션 확정 — 이후에만 원장 적재(고아 방지)
    # Phase 0 unit d: feasibility 커밋을 원장 단일 SSOT에 best-effort 일원화(실패 무중단).
    try:
        from app.services.ledger.ledger_adapters import record_feasibility_commit

        await record_feasibility_commit(
            commit=result,
            tenant_id=str(current_user.tenant_id) if current_user.tenant_id else None,
            project_id=str(_parse_project_id(project_id)),
            created_by=None,
        )
    except Exception as e:  # noqa: BLE001 — 원장 적재 실패가 커밋을 막지 않음
        logger.warning("원장 배선 append 실패(feasibility_vcs): %s", str(e)[:160])
    # Phase 1 성장루프: 수지 결과 자체를 'feasibility' 체인에 적재(read 대상) + 직전 prior read(best-effort).
    try:
        from app.services.ledger.ledger_adapters import record_feasibility_result
        from app.services.ledger.prior_context import load_prior

        _tid = str(current_user.tenant_id) if current_user.tenant_id else None
        _pid = str(_parse_project_id(project_id))
        prior = await load_prior(analysis_type="feasibility", tenant_id=_tid, project_id=_pid)
        await record_feasibility_result(result=req.snapshot, tenant_id=_tid, project_id=_pid)
        if prior:
            logger.info("feasibility 성장루프 — 직전 수지 prior 적용", prior_version=prior.get("version"))
    except Exception as e:  # noqa: BLE001
        logger.warning("feasibility 성장루프 배선 실패 — skipped: %s", str(e)[:160])
    return {"sha": result["sha"], "message": result["message"], "timestamp": result.get("timestamp", "")}


@router.post("/repos/{project_id}/rollback")
async def vcs_rollback(
    project_id: str,
    req: VCSRollbackRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """수지분석 롤백."""
    vcs = FeasibilityVCSDB(db, project_id=_parse_project_id(project_id), tenant_id=current_user.tenant_id)
    result = await vcs.rollback(req.target_sha)
    if not result:
        raise HTTPException(status_code=404, detail="커밋을 찾을 수 없습니다")
    await db.commit()                      # rollback도 신규 commit row 생성 — 확정 후 원장 적재
    # Phase 0 unit d: rollback 커밋도 원장 단일 SSOT에 best-effort 일원화(실패 무중단).
    try:
        from app.services.ledger.ledger_adapters import record_feasibility_commit

        await record_feasibility_commit(
            commit=result,
            tenant_id=str(current_user.tenant_id) if current_user.tenant_id else None,
            project_id=str(_parse_project_id(project_id)),
            created_by=None,
        )
    except Exception as e:  # noqa: BLE001 — 원장 적재 실패가 롤백을 막지 않음
        logger.warning("원장 배선 append 실패(feasibility_vcs/rollback): %s", str(e)[:160])
    return {"sha": result["sha"], "message": result["message"]}


@router.get("/repos/{project_id}/log")
async def vcs_log(
    project_id: str,
    max_count: int = 50,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """커밋 이력."""
    vcs = FeasibilityVCSDB(db, project_id=_parse_project_id(project_id), tenant_id=current_user.tenant_id)
    log_entries = await vcs.log(max_count)
    return {
        "commits": [
            {"sha": c["sha"], "message": c["message"], "parent_sha": c["parent_sha"], "timestamp": c["timestamp"]}
            for c in log_entries
        ]
    }


@router.get("/repos/{project_id}/diff/{sha_a}/{sha_b}")
async def vcs_diff(
    project_id: str,
    sha_a: str,
    sha_b: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """두 커밋 간 diff."""
    vcs = FeasibilityVCSDB(db, project_id=_parse_project_id(project_id), tenant_id=current_user.tenant_id)
    return await vcs.diff(sha_a, sha_b)


@router.post("/export-excel", response_class=Response)
async def export_feasibility_excel(req: FeasibilityCalculateRequest):
    """수지분석 결과를 Excel 파일로 내보낸다."""
    from app.services.export.excel_export_service import ExcelExportService

    try:
        inp = _request_to_input(req)
        output = _service.calculate(inp)

        result_dict = {
            "development_type": output.development_type,
            "module_name": output.module_name,
            "total_revenue_won": output.total_revenue_won,
            "total_cost_won": output.total_cost_won,
            "net_profit_won": output.net_profit_won,
            "profit_rate_pct": output.profit_rate_pct,
            "roi_pct": output.roi_pct,
            "npv_won": output.npv_won,
            "grade": output.grade,
            "cost_breakdown_won": output.cost_detail,
            "tax_detail": output.tax_detail,
        }

        export_svc = ExcelExportService()
        file_bytes, content_type = export_svc.feasibility_to_xlsx(result_dict)

        ext = "xlsx" if "spreadsheet" in content_type else "csv"
        return Response(
            content=file_bytes,
            media_type=content_type,
            headers={
                "Content-Disposition": f'attachment; filename="feasibility_{output.development_type}.{ext}"'
            },
        )
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))


# ------------------------------------------------------------------
# Auto-Recommend Top 3 + Finalize
# ------------------------------------------------------------------


class AutoRecommendRequest(BaseModel):
    address: str
    land_area_sqm: float | None = None
    region: str = "서울"
    equity_won: int = 10_000_000_000
    use_llm: bool = True  # AI 내러티브(수지 해석) 포함 여부(사용자 선택)


@router.post("/auto-recommend", dependencies=[Depends(enforce_llm_quota)])
async def auto_recommend_top3(req: AutoRecommendRequest):
    """부지 주소로부터 최적 사업모델 Top 3 자동 추천."""
    service = FeasibilityServiceV2()
    return await service.auto_recommend_top3(
        address=req.address,
        land_area_sqm=req.land_area_sqm,
        region=req.region,
        equity_won=req.equity_won,
        use_llm=req.use_llm,
    )


class FinalizeRequest(BaseModel):
    project_id: str
    development_type: str
    module_input: dict  # The refined ModuleInput from user


@router.post("/finalize")
async def finalize_business_model(req: FinalizeRequest):
    """선택된 사업모델을 최종 확정."""
    service = FeasibilityServiceV2()
    # Calculate final result
    inp = ModuleInput(**req.module_input)
    result = service.calculate(inp)
    return {
        "project_id": req.project_id,
        "status": "finalized",
        "development_type": req.development_type,
        "final_result": {
            "development_type": result.development_type,
            "module_name": result.module_name,
            "total_revenue_won": result.total_revenue_won,
            "total_cost_won": result.total_cost_won,
            "net_profit_won": result.net_profit_won,
            "profit_rate_pct": result.profit_rate_pct,
            "roi_pct": result.roi_pct,
            "npv_won": result.npv_won,
            "grade": result.grade,
            "cost_breakdown_won": result.cost_detail,
            "tax_detail": result.tax_detail,
        },
        "finalized_at": datetime.now().isoformat(),
    }


# ── 다기간 DCF 월별 현금흐름(베팅 B) ──────────────────────────────
class CashflowRequest(BaseModel):
    land_cost_won: float
    construction_cost_won: float
    construction_months: int = 24
    total_revenue_won: float
    sale_start_month: int = 6          # 시공 개시 기준 분양 시작(월)
    sale_duration_months: int = 6
    bridge_loan_rate: float = 0.08
    pf_loan_rate: float = 0.065
    equity_ratio: float = 0.3
    design_months: int = 3
    design_cost_ratio: float = 0.03
    discount_rate_annual: float = 0.06  # NPV 할인율(연)
    # R1(additive): integrated_tax_engine.calculate_all_taxes 키워드 입력(부분집합).
    # 지정 시 38종 세금을 시점 매핑 주입해 summary에 after_tax_irr_annual_pct·total_tax_won 가산.
    # 미지정(None, 기본)이면 기존 세전 현금흐름과 완전 동일.
    tax_inputs: dict | None = None


def _build_cashflow(req: CashflowRequest) -> dict:
    from app.services.feasibility.cashflow_generator import (
        CashflowGenerator,
        build_tax_schedule_from_integrated,
    )

    # ── R1 세후 IRR(additive): tax_inputs 지정 시에만 통합 세금엔진(38종) 주입 ──
    tax_schedule = None
    if req.tax_inputs:
        import inspect

        from app.services.tax.integrated_tax_engine import calculate_all_taxes

        allowed = set(inspect.signature(calculate_all_taxes).parameters)
        tax_kwargs = {k: v for k, v in req.tax_inputs.items() if k in allowed}
        tax_schedule = build_tax_schedule_from_integrated(
            calculate_all_taxes(**tax_kwargs)
        )

    cf = CashflowGenerator().generate_monthly_cashflow(
        land_cost=req.land_cost_won,
        construction_cost=req.construction_cost_won,
        construction_months=max(1, req.construction_months),
        total_revenue=req.total_revenue_won,
        sale_start_month=max(0, req.sale_start_month),
        sale_duration_months=max(1, req.sale_duration_months),
        bridge_loan_rate=req.bridge_loan_rate,
        pf_loan_rate=req.pf_loan_rate,
        equity_ratio=req.equity_ratio,
        design_months=max(0, req.design_months),
        design_cost_ratio=req.design_cost_ratio,
        tax_schedule=tax_schedule,
    )
    # 월 할인율로 NPV 재계산(엔진 IRR과 별개로 사용자 지정 할인율 반영)
    rmonthly = (1 + req.discount_rate_annual) ** (1 / 12) - 1
    npv = 0.0
    for r in cf["rows"]:
        net = (r.get("inflow", 0) or 0) - (r.get("outflow", 0) or 0)
        m = r.get("month", 0) or 0
        npv += net / ((1 + rmonthly) ** m)
    cf["summary"]["npv_won"] = round(npv)
    cf["summary"]["discount_rate_annual_pct"] = round(req.discount_rate_annual * 100, 2)
    return cf


@router.post("/cashflow")
async def cashflow(req: CashflowRequest):
    """다기간(월별) DCF 현금흐름 + IRR·NPV·peak 자금소요. (은행제출용 정밀 사업성)"""
    try:
        return _build_cashflow(req)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(status_code=422, detail=f"현금흐름 산정 실패: {str(e)[:160]}")


# ── 개발금융(PF·브릿지·이자·LTV·DSCR) ────────────────────────────
class DevelopmentFinanceRequest(BaseModel):
    """수지(P2)에서 흘러온 총사업비·자기자본·토지/공사비를 입력받아 개발금융 산출."""

    total_project_cost_won: float
    equity_ratio: float | None = None       # 자기자본 비율(미입력 시 equity_won 사용, 둘 다 없으면 0.3)
    equity_won: float | None = None         # 자기자본 절대액(있으면 비율보다 우선)
    land_cost_won: float | None = None      # 미입력 시 총사업비의 30% 가정(브릿지 산정용)
    construction_cost_won: float | None = None
    annual_noi_won: float | None = None     # 연 순영업이익(DSCR 산정용, 없으면 분양형으로 간주)
    credit_grade: str = "A"
    presale_ratio: float = 0.0
    bridge_months: int = 12
    pf_months: int = 30


def _build_development_finance(req: DevelopmentFinanceRequest) -> dict:
    """finance_cost_engine 재사용 — PF/브릿지/이자, LTV/DSCR(표준 비율식)."""
    from app.services.feasibility.finance_cost_engine import (
        calculate_bridge_loan,
        calculate_pf_loan,
    )

    total_cost = max(0.0, float(req.total_project_cost_won))
    if total_cost <= 0:
        raise ValueError("total_project_cost_won must be > 0")

    # ── 자기자본 결정(절대액 우선 → 비율 → 표준 0.3) ──
    if req.equity_won is not None and req.equity_won > 0:
        equity_won = float(req.equity_won)
        equity_ratio = equity_won / total_cost
    else:
        equity_ratio = req.equity_ratio if req.equity_ratio is not None else 0.3
        equity_ratio = min(max(equity_ratio, 0.0), 1.0)
        equity_won = total_cost * equity_ratio

    # ── 자금 구조(cashflow_generator 관례: 브릿지=토지비 중 타인자본, PF=나머지) ──
    land_cost = float(req.land_cost_won) if req.land_cost_won else total_cost * 0.3
    bridge_amount = int(max(0.0, land_cost * (1 - equity_ratio)))
    pf_amount = int(max(0.0, total_cost - equity_won - bridge_amount))

    # ── finance_cost_engine 재사용(이자·수수료) ──
    bridge = calculate_bridge_loan(
        amount_won=bridge_amount, months=max(1, req.bridge_months)
    )
    pf = calculate_pf_loan(
        amount_won=pf_amount,
        months=max(1, req.pf_months),
        credit_grade=req.credit_grade,
        presale_ratio=req.presale_ratio,
    )

    total_debt = bridge_amount + pf_amount
    total_finance_cost = bridge["total_bridge_cost_won"] + pf["total_pf_cost_won"]

    # ── LTV(총부채/총사업비), DSCR(연 NOI/연 부채상환액) — 표준 비율식 ──
    ltv = round(total_debt / total_cost, 4) if total_cost > 0 else 0.0

    # 연간 이자(부채상환액 근사): PF 잔액×PF금리 + 브릿지 잔액×브릿지금리
    annual_debt_service = pf_amount * pf["rate"] + bridge_amount * bridge["rate"]
    dscr: float | None = None
    if req.annual_noi_won is not None and annual_debt_service > 0:
        dscr = round(float(req.annual_noi_won) / annual_debt_service, 2)

    return {
        "total_project_cost_won": int(total_cost),
        "equity_won": int(equity_won),
        "equity_ratio": round(equity_ratio, 4),
        "pf_loan": {
            "amount_won": pf_amount,
            "rate": pf["rate"],
            "interest_won": pf["interest_won"],
            "guarantee_fee_won": pf["guarantee_fee_won"],
            "months": pf["months"],
            "total_cost_won": pf["total_pf_cost_won"],
        },
        "bridge_loan": {
            "amount_won": bridge_amount,
            "rate": bridge["rate"],
            "interest_won": bridge["interest_won"],
            "arrangement_fee_won": bridge["arrangement_fee_won"],
            "months": bridge["months"],
            "total_cost_won": bridge["total_bridge_cost_won"],
        },
        "total_debt_won": total_debt,
        "ltv": ltv,
        "dscr": dscr,
        "annual_debt_service_won": int(annual_debt_service),
        "total_financing_cost_won": total_finance_cost,
    }


@router.post("/development-finance")
async def development_finance(req: DevelopmentFinanceRequest):
    """개발금융(PF·브릿지·이자·LTV·DSCR·자기자본비율) 산출.

    수지분석(총사업비)·공사비·토지비를 입력으로 finance_cost_engine을 재사용해
    PF대출·금리·총이자·LTV·DSCR을 반환. 새 금융계산 로직 미작성(엔진 재사용).
    """
    try:
        return _build_development_finance(req)
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except Exception as e:  # noqa: BLE001
        raise HTTPException(
            status_code=422, detail=f"개발금융 산정 실패: {str(e)[:160]}"
        )


@router.post("/cashflow/excel")
async def cashflow_excel(req: CashflowRequest):
    """월별 현금흐름을 Excel(xlsx)로 다운로드."""
    try:
        cf = _build_cashflow(req)
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active or wb.create_sheet()
        ws.title = "월별 현금흐름(DCF)"

        # 요약
        s = cf["summary"]
        ws.append(["■ 사업성 요약 (DCF)"])
        ws["A1"].font = Font(bold=True, size=13)
        summary_rows = [
            ("총 분양수입(원)", s.get("total_inflow")),
            ("총 사업비(원)", s.get("total_outflow")),
            ("순이익(원)", s.get("net_profit")),
            ("수익률(%)", s.get("profit_rate_pct")),
            ("IRR(연,%)", s.get("irr_annual_pct")),
            ("NPV(원)", s.get("npv_won")),
            ("할인율(연,%)", s.get("discount_rate_annual_pct")),
            ("최대 자금소요(peak, 원)", s.get("peak_negative_cashflow")),
            ("자기자본(원)", s.get("equity_amount")),
            ("브릿지론(원)", s.get("bridge_loan_amount")),
            ("PF론(원)", s.get("pf_loan_amount")),
        ]
        for k, v in summary_rows:
            ws.append([k, v])
        ws.append([])

        # 월별 표 헤더
        hdr_row = ws.max_row + 1
        headers = ["월", "단계", "유입(원)", "유출(원)", "순현금(원)", "누적현금(원)"]
        ws.append(headers)
        head_fill = PatternFill("solid", fgColor="1F2937")
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=hdr_row, column=c)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = head_fill
            cell.alignment = Alignment(horizontal="center")

        for r in cf["rows"]:
            net = (r.get("inflow", 0) or 0) - (r.get("outflow", 0) or 0)
            ws.append([
                r.get("month"), r.get("phase", ""),
                round(r.get("inflow", 0) or 0), round(r.get("outflow", 0) or 0),
                round(net), round(r.get("cumulative", 0) or 0),
            ])

        # 숫자 포맷·열너비
        for col in ("C", "D", "E", "F"):
            for cell in ws[col]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0"
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 16
        for col in ("C", "D", "E", "F"):
            ws.column_dimensions[col].width = 18

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=propai_cashflow_dcf.xlsx"},
        )
    except Exception as e:  # noqa: BLE001
        raise HTTPException(status_code=422, detail=f"엑셀 생성 실패: {str(e)[:160]}")
