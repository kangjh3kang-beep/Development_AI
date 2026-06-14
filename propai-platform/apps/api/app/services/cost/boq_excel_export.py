"""공내역서 엑셀 내보내기 (B2) — 파라메트릭 초안(generate_draft) → xlsx bytes.

실무 공내역서 포맷 미러:
- 시트 = 공종별 1시트.
- 헤더 2행: 품명 | 규격 | 단위 | 수량 | 재료비(단가·금액) | 노무비(단가·금액)
            | 경비(단가·금액) | 합계(단가·금액) | 비고.
- 단가·금액 칸 전부 공란 = 공내역서 표준(입찰자 기입).
- 섹션 행 굵게(음영), 하단에 출처(provenance)·주의 문구.

정직성: 단가를 절대 생성하지 않는다. 전기 참고 재료단가(ref_mat_price)는
비고란에 '참고'로만 표기(원본 실적값 — 단가칸 미기입). 결정론(LLM 0).
"""

from __future__ import annotations

import re
from io import BytesIO
from typing import Any

_HEADER_TOP = ["품명", "규격", "단위", "수량", "재료비", "", "노무비", "", "경비", "", "합계", "", "비고"]
_HEADER_SUB = ["", "", "", "", "단가", "금액", "단가", "금액", "단가", "금액", "단가", "금액", ""]
_N_COLS = 13
_COL_WIDTHS = [34, 18, 7, 11, 9, 11, 9, 11, 9, 11, 9, 12, 26]

_DRIVER_LABEL = {
    "gfa": "연면적 원단위",
    "households": "세대당 원단위",
    "landscape_area": "조경면적 원단위",
    "fixed": "고정(횟수성)",
}

_INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")


def _sheet_title(name: str) -> str:
    """엑셀 시트명 제약(금지문자·31자) 반영."""
    cleaned = _INVALID_SHEET_CHARS.sub(" ", str(name)).strip() or "공종"
    return cleaned[:31]


def _remark(item: dict[str, Any]) -> str:
    """비고: 드라이버 라벨(+폴백 표기) + 전기 참고 재료단가(참고용 — 단가칸 미기입)."""
    basis = item.get("qty_basis") or {}
    parts: list[str] = []
    label = _DRIVER_LABEL.get(basis.get("driver"), basis.get("driver") or "")
    if label:
        parts.append(label)
    if basis.get("fallback_from"):
        parts.append(f"{basis['fallback_from']} 미제공→gfa 폴백")
    ref_mat = item.get("ref_mat_price")
    if isinstance(ref_mat, (int, float)):
        parts.append(f"참고 재료단가 {ref_mat:,.0f}원(실적 1건)")
    return " · ".join(parts)


def _round_won(value: float) -> int:
    return int(round(value))


def _cost_cells(item: dict[str, Any]) -> list[Any]:
    """금액 모드(N3) 단가/금액 8칸 산출 — [재료단가, 재료금액, 노무단가, 노무금액,
    경비단가, 경비금액, 합계단가, 합계금액]. 미결합 단가는 None(빈칸 유지·정직).
    """
    qty = item.get("qty")
    qf = float(qty) if isinstance(qty, (int, float)) and not isinstance(qty, bool) else None
    mu, lu, eu = item.get("mat_unit"), item.get("labor_unit"), item.get("exp_unit")

    def _amt(u: Any) -> int | None:
        return _round_won(qf * u) if (qf is not None and isinstance(u, (int, float))
                                      and not isinstance(u, bool)) else None

    units = [u for u in (mu, lu, eu) if isinstance(u, (int, float)) and not isinstance(u, bool)]
    tot_unit = sum(units) if units else None
    tot_amt = item.get("amount") if isinstance(item.get("amount"), (int, float)) else None
    return [mu, _amt(mu), lu, _amt(lu), eu, _amt(eu), tot_unit, tot_amt]


def build_xlsx(draft: dict[str, Any], *, priced: bool = False) -> bytes:
    """generate_draft(또는 join_prices) 응답 → 공내역서 xlsx bytes (openpyxl).

    시트 수 = draft.disciplines 키 수(0이면 안내 시트 1장 — 빈 통합문서 금지).

    priced=False(기본): 단가·금액 칸 공란 = 공내역서 표준(입찰자 기입). 하위호환 유지.
    priced=True(N3 금액 모드): join_prices 가 결합한 단가/금액을 채우고, 공종별 '소계' 행과
      별도 '총계' 시트(공종별 합계 + 총합)를 추가한다. 미결합 항목은 빈칸 유지(가짜 단가 금지).
    """
    from openpyxl import Workbook  # noqa: PLC0415 — 사용 시점 임포트(기동 비용 절감)
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: PLC0415
    from openpyxl.utils import get_column_letter  # noqa: PLC0415

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right")
    head_fill = PatternFill(fill_type="solid", start_color="FFD9D9D9")
    sect_fill = PatternFill(fill_type="solid", start_color="FFF2F2F2")
    thin = Side(style="thin", color="FF999999")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    badges = draft.get("badges") or {}
    provenance = draft.get("provenance") or {}
    summary = draft.get("summary") or {}
    disciplines: dict[str, Any] = draft.get("disciplines") or {}

    wb = Workbook()
    wb.remove(wb.active)

    money_fmt = "#,##0"
    # priced 모드 총계 집계: 공종별 [재료금액, 노무금액, 경비금액, 합계금액].
    grand: list[tuple[str, list[int]]] = []

    for disc, block in disciplines.items():
        ws = wb.create_sheet(title=_sheet_title(disc))
        disc_sub = [0, 0, 0, 0]  # 재료/노무/경비/합계 금액 소계

        # ── 헤더 2행(병합) ──
        ws.append(_HEADER_TOP)
        ws.append(_HEADER_SUB)
        for col_pair_start in (5, 7, 9, 11):  # 재료비/노무비/경비/합계 — 가로 병합
            ws.merge_cells(start_row=1, start_column=col_pair_start,
                           end_row=1, end_column=col_pair_start + 1)
        for col in (1, 2, 3, 4, 13):  # 품명/규격/단위/수량/비고 — 세로 병합
            ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
        for row in (1, 2):
            for col in range(1, _N_COLS + 1):
                c = ws.cell(row=row, column=col)
                c.font = bold
                c.alignment = center
                c.fill = head_fill
                c.border = box

        # ── 데이터: 섹션 행(굵게) + 항목 행(단가·금액 공란) ──
        current_section: tuple[Any, Any] | None = None
        for item in block.get("items") or []:
            sec_key = (item.get("section_code"), item.get("section_name"))
            if sec_key != current_section:
                current_section = sec_key
                code, name = sec_key
                ws.append([f"[{code}] {name}"] + [""] * (_N_COLS - 1))
                r = ws.max_row
                for col in range(1, _N_COLS + 1):
                    c = ws.cell(row=r, column=col)
                    c.font = bold
                    c.fill = sect_fill
                    c.border = box
            cost = _cost_cells(item) if priced else [None] * 8
            if priced:
                for i, amt_idx in enumerate((1, 3, 5, 7)):  # 재료/노무/경비/합계 금액
                    v = cost[amt_idx]
                    if isinstance(v, (int, float)):
                        disc_sub[i] += int(v)
            ws.append([
                item.get("name"), item.get("spec"), item.get("unit"), item.get("qty"),
                *cost,
                _remark(item),
            ])
            r = ws.max_row
            for col in range(1, _N_COLS + 1):
                ws.cell(row=r, column=col).border = box
            ws.cell(row=r, column=4).alignment = right
            ws.cell(row=r, column=4).number_format = "#,##0.###"
            if priced:  # 단가·금액 칸 우측정렬·천단위 서식
                for col in range(5, 13):
                    ws.cell(row=r, column=col).alignment = right
                    ws.cell(row=r, column=col).number_format = money_fmt

        if priced:  # ── 공종 소계 행 ──
            ws.append([f"소계 ({disc})", "", "", "",
                       None, disc_sub[0], None, disc_sub[1], None, disc_sub[2],
                       None, disc_sub[3], ""])
            r = ws.max_row
            for col in range(1, _N_COLS + 1):
                c = ws.cell(row=r, column=col)
                c.font = bold
                c.fill = sect_fill
                c.border = box
                if col in (6, 8, 10, 12):
                    c.alignment = right
                    c.number_format = money_fmt
            grand.append((str(disc), list(disc_sub)))

        # ── 하단 출처·주의 문구 ──
        ws.append([""] * _N_COLS)
        mode_note = (
            "※ 단가·금액 = 결합 단가(DB/fallback·도면참고단가). 미결합 항목은 공란 유지"
            "(가짜 단가 금지) — 부분 단가·전문 적산 검토 필수."
            if priced else
            "※ 단가·금액란 공란 = 공내역서 양식(입찰자 기입). 본 문서는 단가를 생성하지 않음."
        )
        notes = [
            f"※ {badges.get('note', '실적 1건 기반 원단위 초안 — 전문 적산 검토 필수')}"
            f" (신뢰도: {badges.get('confidence', '낮음(n=1)')})",
            mode_note,
            f"※ 출처: {provenance.get('name', '미상')} 실적 공내역서 "
            f"{provenance.get('sample_count', '?')}건 — {provenance.get('provenance', '')}",
        ]
        warnings = (summary.get("warnings") or []) if isinstance(summary, dict) else []
        notes.extend(f"※ 경고: {w}" for w in warnings)
        for note in notes:
            ws.append([note] + [""] * (_N_COLS - 1))
            ws.cell(row=ws.max_row, column=1).font = Font(italic=True, size=9)

        for idx, width in enumerate(_COL_WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        ws.freeze_panes = "A3"

    if priced and grand:  # ── 총계 시트(공종별 합계 + 총합) ──
        ws = wb.create_sheet(title="총계")
        ws.append(["공종", "재료비", "노무비", "경비", "합계"])
        for col in range(1, 6):
            c = ws.cell(row=1, column=col)
            c.font = bold
            c.fill = head_fill
            c.border = box
            c.alignment = center
        tot = [0, 0, 0, 0]
        for name, sub in grand:
            ws.append([name, sub[0], sub[1], sub[2], sub[3]])
            for i in range(4):
                tot[i] += sub[i]
            r = ws.max_row
            for col in range(2, 6):
                ws.cell(row=r, column=col).number_format = money_fmt
                ws.cell(row=r, column=col).alignment = right
        ws.append(["총계", tot[0], tot[1], tot[2], tot[3]])
        r = ws.max_row
        for col in range(1, 6):
            c = ws.cell(row=r, column=col)
            c.font = bold
            c.fill = sect_fill
            if col >= 2:
                c.number_format = money_fmt
                c.alignment = right
        ws.append([""] * 5)
        ws.append(["※ 금액은 단가가 결합된 항목만의 부분합 — 미결합 항목 제외(가짜 단가 금지). "
                   "전문 적산사 검토 필수."])
        ws.cell(row=ws.max_row, column=1).font = Font(italic=True, size=9)
        for idx, width in enumerate((20, 16, 16, 16, 18), start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

    if not wb.sheetnames:  # 빈 draft 도 유효 xlsx 반환(정직 안내)
        ws = wb.create_sheet(title="안내")
        ws.append(["공종 데이터 없음 — generate_draft(disciplines=...) 결과가 비어 있습니다."])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
