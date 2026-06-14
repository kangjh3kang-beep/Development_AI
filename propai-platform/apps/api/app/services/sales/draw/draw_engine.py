"""동·호 추첨 엔진 — 즉석추첨(누르면 남은 동호 중 무작위 1개 공개) + seed 해시체인 감사.

흐름:
  1) 추첨그룹 생성(create_group) + 대상자 등록(add_candidates / import_excel / from_customers)
  2) 그룹 동·호판(unit pool) 지정(set_pool) — 그룹이 추첨할 대상 세대 집합
  3) 대상자 순번대로 추첨(draw_for_candidate): 남은 가용 세대 중 seed 기반 무작위 1개 배정
     - 배정 시 세대 상태 HOLD 전이 + 이벤트 원장 DRAW_ASSIGN(seed·group·candidate) 기록(감사)
  공정성: seed=secrets, 선택 = random.Random(seed).choice(sorted(remaining)) → seed·remaining·결과를
  원장에 남겨 누구나 재현·검증(부정 재추첨 방지). content_hash 체인으로 변조탐지.

지정모드와 공존: 지정모드는 lifecycle_actions(직접 배정), 추첨모드는 본 엔진(무작위). 모드는 프론트 토글.
"""
from __future__ import annotations

import hashlib
import io
import random
import secrets
from typing import Any

from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.services.sales.units.event_ledger import append_event

_DDL_GROUPS = (
    "CREATE TABLE IF NOT EXISTS sales_draw_groups ("
    "  id uuid PRIMARY KEY DEFAULT gen_random_uuid(),"
    "  site_id uuid NOT NULL,"
    "  name varchar(120) NOT NULL,"
    "  status varchar(12) NOT NULL DEFAULT 'OPEN',"   # OPEN/CLOSED
    "  unit_pool jsonb,"                               # 그룹 동·호판 unit_id 배열(미지정=현장 전체 가용)
    "  created_at timestamptz NOT NULL DEFAULT now()"
    ")"
)
_DDL_CAND = (
    "CREATE TABLE IF NOT EXISTS sales_draw_candidates ("
    "  id uuid PRIMARY KEY DEFAULT gen_random_uuid(),"
    "  group_id uuid NOT NULL,"
    "  site_id uuid NOT NULL,"
    "  seq integer NOT NULL,"                          # 추첨 순번
    "  name varchar(120) NOT NULL,"
    "  phone varchar(40),"
    "  customer_id uuid,"
    "  assigned_unit_id uuid,"                          # 추첨 배정된 세대
    "  draw_seed varchar(32),"
    "  drawn_at timestamptz,"
    "  UNIQUE (group_id, seq)"
    ")"
)
_READY = False


async def _ensure(db: AsyncSession) -> None:
    global _READY
    if _READY:
        return
    await db.execute(text(_DDL_GROUPS))
    await db.execute(text(_DDL_CAND))
    await db.execute(text("CREATE INDEX IF NOT EXISTS ix_draw_cand_group ON sales_draw_candidates(group_id, seq)"))
    await db.commit()
    _READY = True


async def create_group(db: AsyncSession, site_id, name: str) -> dict[str, Any]:
    await _ensure(db)
    if not (name or "").strip():
        raise ValueError("추첨그룹 이름을 입력하세요")
    row = (await db.execute(text(
        "INSERT INTO sales_draw_groups (site_id, name) VALUES (:s,:n) RETURNING id"),
        {"s": str(site_id), "n": name.strip()})).first()
    await db.commit()
    return {"id": str(row[0]), "name": name.strip(), "status": "OPEN"}


async def list_groups(db: AsyncSession, site_id) -> list[dict[str, Any]]:
    """현장 추첨그룹 목록(대상자수·완료수 포함)."""
    await _ensure(db)
    rows = (await db.execute(text(
        "SELECT g.id, g.name, g.status, g.created_at, "
        "  (SELECT count(*) FROM sales_draw_candidates c WHERE c.group_id=g.id) AS cand, "
        "  (SELECT count(*) FROM sales_draw_candidates c WHERE c.group_id=g.id AND c.assigned_unit_id IS NOT NULL) AS drawn "
        "FROM sales_draw_groups g WHERE g.site_id=:s ORDER BY g.created_at DESC"),
        {"s": str(site_id)})).all()
    return [{"id": str(r[0]), "name": r[1], "status": r[2], "created_at": str(r[3]),
             "candidates": int(r[4]), "drawn": int(r[5])} for r in rows]


async def set_pool(db: AsyncSession, site_id, group_id, unit_ids: list[str]) -> dict[str, Any]:
    """그룹 동·호판(추첨 대상 세대 집합) 지정."""
    await _ensure(db)
    import json
    await db.execute(text(
        "UPDATE sales_draw_groups SET unit_pool=CAST(:p AS jsonb) WHERE id=:g AND site_id=:s"),
        {"p": json.dumps([str(u) for u in unit_ids]), "g": str(group_id), "s": str(site_id)})
    await db.commit()
    return {"ok": True, "pool_size": len(unit_ids)}


async def _next_seq(db: AsyncSession, group_id) -> int:
    r = (await db.execute(text(
        "SELECT COALESCE(MAX(seq),0) FROM sales_draw_candidates WHERE group_id=:g"), {"g": str(group_id)})).first()
    return int(r[0] or 0) + 1


async def add_candidates(db: AsyncSession, site_id, group_id, rows: list[dict]) -> dict[str, Any]:
    """대상자 일괄 등록 — rows: [{name, phone?, customer_id?}]. 순번은 기존 뒤에 이어 부여."""
    await _ensure(db)
    seq = await _next_seq(db, group_id)
    n = 0
    for r in rows:
        name = str(r.get("name") or "").strip()
        if not name:
            continue
        await db.execute(text(
            "INSERT INTO sales_draw_candidates (group_id, site_id, seq, name, phone, customer_id) "
            "VALUES (:g,:s,:seq,:n,:p,:c)"),
            {"g": str(group_id), "s": str(site_id), "seq": seq, "n": name,
             "p": (r.get("phone") or None), "c": (str(r["customer_id"]) if r.get("customer_id") else None)})
        seq += 1
        n += 1
    await db.commit()
    return {"ok": True, "added": n}


def parse_excel(content: bytes) -> list[dict]:
    """고객명부 Excel(.xlsx) 파싱 — 1행 헤더(이름/성명/name, 연락처/전화/phone) 자동인식, 그 외 첫2열=이름,전화."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c or "").strip().lower() for c in rows[0]]
    def _idx(keys, default):
        for i, h in enumerate(header):
            if any(k in h for k in keys):
                return i
        return default
    ni = _idx(["이름", "성명", "name"], 0)
    pi = _idx(["연락처", "전화", "휴대", "phone", "tel"], 1)
    out = []
    for r in rows[1:]:
        if not r or all(c is None for c in r):
            continue
        name = str(r[ni]).strip() if ni < len(r) and r[ni] is not None else ""
        phone = str(r[pi]).strip() if pi < len(r) and r[pi] is not None else None
        if name:
            out.append({"name": name, "phone": phone})
    return out


async def from_customers(db: AsyncSession, site_id, group_id, customer_ids: list[str] | None = None) -> dict[str, Any]:
    """계약자/고객 명부에서 대상자 선별 등록. customer_ids 미지정 시 현장 전체 고객."""
    await _ensure(db)
    q = "SELECT id, name, phone FROM sales_customers WHERE site_id=:s AND deleted_at IS NULL"
    params: dict[str, Any] = {"s": str(site_id)}
    if customer_ids:
        q += " AND id = ANY(:ids)"
        params["ids"] = [str(c) for c in customer_ids]
    try:
        rows = (await db.execute(text(q), params)).all()
    except Exception:  # noqa: BLE001 — 컬럼차이 시 최소 컬럼
        await db.rollback()
        rows = (await db.execute(text(
            "SELECT id, name, NULL FROM sales_customers WHERE site_id=:s"), {"s": str(site_id)})).all()
    cands = [{"name": r[1] or "고객", "phone": r[2], "customer_id": str(r[0])} for r in rows]
    return await add_candidates(db, site_id, group_id, cands)


async def _remaining_units(db: AsyncSession, site_id, group_id) -> list[str]:
    """그룹 동·호판에서 아직 배정되지 않은(AVAILABLE) 세대 목록."""
    import json
    g = (await db.execute(text(
        "SELECT unit_pool FROM sales_draw_groups WHERE id=:g AND site_id=:s"),
        {"g": str(group_id), "s": str(site_id)})).first()
    pool = None
    if g and g[0]:
        pool = g[0] if isinstance(g[0], list) else json.loads(g[0])
    if pool:
        rows = (await db.execute(text(
            "SELECT id FROM sales_unit_inventory WHERE site_id=:s AND id = ANY(:ids) "
            "AND status='AVAILABLE' AND deleted_at IS NULL"),
            {"s": str(site_id), "ids": [str(u) for u in pool]})).all()
    else:  # 풀 미지정 → 현장 전체 가용
        rows = (await db.execute(text(
            "SELECT id FROM sales_unit_inventory WHERE site_id=:s AND status='AVAILABLE' AND deleted_at IS NULL"),
            {"s": str(site_id)})).all()
    # 이미 같은 그룹서 배정된 세대 제외(이중배정 방지).
    assigned = {str(r[0]) for r in (await db.execute(text(
        "SELECT assigned_unit_id FROM sales_draw_candidates WHERE group_id=:g AND assigned_unit_id IS NOT NULL"),
        {"g": str(group_id)})).all()}
    return [str(r[0]) for r in rows if str(r[0]) not in assigned]


async def draw_for_candidate(db: AsyncSession, site_id, group_id, candidate_id, by=None) -> dict[str, Any]:
    """즉석추첨 — 대상자가 누르면 남은 동호 중 seed 기반 무작위 1개를 배정·공개(감사 기록)."""
    await _ensure(db)
    cand = (await db.execute(text(
        "SELECT seq, name, assigned_unit_id FROM sales_draw_candidates WHERE id=:c AND group_id=:g"),
        {"c": str(candidate_id), "g": str(group_id)})).first()
    if not cand:
        raise ValueError("추첨 대상자를 찾을 수 없습니다")
    if cand[2]:
        raise ValueError("이미 추첨이 완료된 대상자입니다")
    # ★HOLD 선점은 'RETURNING id' 로 원자 확정한다(이중배정 방지). 동시 추첨/지정 경합으로 방금 선점된
    #   세대(0행)면 그 세대를 빼고 남은 가용 세대를 재조회해 재추첨한다. 모든 후보가 소진되면 명확한 에러.
    # 공정 추첨: 매 시도마다 secrets seed → 결정적 선택(재현·검증 가능). sorted 로 입력 순서 영향 제거.
    chosen: str | None = None
    seed = ""
    pool_sorted: list[str] = []
    pool_hash = ""
    excluded: set[str] = set()  # 이번 호출에서 선점 실패해 제외한 세대(재조회 후에도 중복 시도 방지).
    while True:
        remaining = [u for u in await _remaining_units(db, site_id, group_id) if u not in excluded]
        if not remaining:
            raise ValueError("남은 가용 세대가 없습니다(추첨 종료)")
        seed = secrets.token_hex(8)
        pool_sorted = sorted(remaining)
        candidate_unit = random.Random(seed).choice(pool_sorted)
        pool_hash = hashlib.sha256(",".join(pool_sorted).encode()).hexdigest()[:16]
        # HOLD 선점(원자 조건부 UPDATE) — 0행이면 이미 다른 곳에서 선점됨 → 재추첨.
        won = (await db.execute(text(
            "UPDATE sales_unit_inventory SET status='HOLD' WHERE id=:u AND status='AVAILABLE' RETURNING id"),
            {"u": candidate_unit})).first()
        if won:
            chosen = candidate_unit
            break
        excluded.add(candidate_unit)  # 방금 선점된 세대 — 제외하고 재추첨.

    # 세대 정보(공개용)는 HOLD 확정 후 조회. DRAW_ASSIGN 이벤트도 확정 성공분만 기록한다.
    u = (await db.execute(text(
        "SELECT dong, ho FROM sales_unit_inventory WHERE id=:u"), {"u": chosen})).first()
    await db.execute(text(
        "UPDATE sales_draw_candidates SET assigned_unit_id=:u, draw_seed=:seed, drawn_at=now() WHERE id=:c"),
        {"u": chosen, "seed": seed, "c": str(candidate_id)})

    # 감사: 이벤트 원장에 추첨 배정 기록(seed·group·candidate·pool_hash → 사후 검증).
    # HOLD UPDATE + candidate UPDATE + 이벤트 원장을 한 트랜잭션으로 묶어 한 번에 커밋(원자성).
    ev = await append_event(db, site_id, chosen, "DRAW_ASSIGN", from_status="AVAILABLE", to_status="HOLD",
                            message=f"추첨 배정: {cand[1]}(순번 {cand[0]})", by=by,
                            meta={"group_id": str(group_id), "candidate_id": str(candidate_id),
                                  "seed": seed, "pool_hash": pool_hash, "pool_size": len(pool_sorted)},
                            do_commit=False)
    await db.commit()
    return {
        "ok": True, "candidate": {"seq": int(cand[0]), "name": cand[1]},
        "assigned_unit": {"id": chosen, "dong": u[0] if u else None, "ho": u[1] if u else None},
        "seed": seed, "pool_hash": pool_hash, "pool_size": len(pool_sorted), "remaining_after": len(remaining) - 1,
        "event": ev,
    }


async def group_status(db: AsyncSession, site_id, group_id) -> dict[str, Any]:
    """그룹 현황 — 대상자(순번·배정세대)·진행률·남은 세대."""
    await _ensure(db)
    g = (await db.execute(text(
        "SELECT name, status FROM sales_draw_groups WHERE id=:g AND site_id=:s"),
        {"g": str(group_id), "s": str(site_id)})).first()
    if not g:
        raise ValueError("추첨그룹을 찾을 수 없습니다")
    cands = (await db.execute(text(
        "SELECT c.seq, c.name, c.phone, c.assigned_unit_id, c.draw_seed, c.drawn_at, u.dong, u.ho, c.id "
        "FROM sales_draw_candidates c LEFT JOIN sales_unit_inventory u ON u.id=c.assigned_unit_id "
        "WHERE c.group_id=:g ORDER BY c.seq ASC"), {"g": str(group_id)})).all()
    roster = [{
        "id": str(c[8]),  # candidate_id — 프론트 추첨 버튼용
        "seq": int(c[0]), "name": c[1], "phone": c[2],
        "assigned_unit_id": str(c[3]) if c[3] else None,
        "assigned_label": (f"{c[6]}동 {c[7]}호" if c[3] and c[6] else None),
        "seed": c[4], "drawn_at": str(c[5]) if c[5] else None, "done": bool(c[3]),
    } for c in cands]
    remaining = await _remaining_units(db, site_id, group_id)
    drawn = sum(1 for r in roster if r["done"])
    return {"group_id": str(group_id), "name": g[0], "status": g[1],
            "candidates": len(roster), "drawn": drawn, "remaining_units": len(remaining), "roster": roster}
