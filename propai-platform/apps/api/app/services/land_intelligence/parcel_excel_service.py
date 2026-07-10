"""다필지 토지조서 엑셀 — 플랫폼 최적 양식 생성 + 업로드 파싱(주소/필지 추출).

목적: 사용자가 다필지를 일일이 검색하지 않고, 플랫폼 최적 양식 엑셀에 작성해 업로드하면
필지(주소·지번·법정동코드·PNU·면적·지목·소유구분)를 추출해 다필지 주소등록에 주입한다.

PNU 결정 우선순위(가짜 금지·정직표기):
  ① PNU열(19자리) 그대로  ② 법정동코드(bcode 10자리)+지번 조합  ③ 주소 VWorld 지오코딩(좌표·PNU)
무자료/해석불가 행은 status로 정직 표기(스킵하지 않고 사유 노출 — 사용자가 보정 가능).
의존성: openpyxl·pandas(이미 설치). 외부호출은 ③ 지오코딩만(상한·동시성 제한).
"""
from __future__ import annotations

import asyncio
import io
import re
from collections import Counter
from typing import Any

import structlog

logger = structlog.get_logger(__name__)

# 양식 컬럼(순서·헤더 = 플랫폼 표준). 필수=소재지(주소). 나머지는 있으면 정확도↑.
# ★예시행의 면적/지목은 '비움' — 면적은 비워두면 공부상(VWorld)에서 자동조회되며, 예시값을
#   남겨두면(예 14959·8500) 그대로 적재돼 합계가 부풀려지던 사고가 있었다(예시값 미삭제 사고).
TEMPLATE_COLUMNS = [
    ("연번", "예시1(삭제)"),
    ("소재지(주소)", "경기도 의정부시 의정부동 224"),
    ("지번", "224"),
    ("법정동코드(bcode·10자리)", "4115010100"),
    ("PNU(필지고유번호·19자리)", ""),
    ("지목", ""),
    ("면적(㎡)", ""),
    ("소유구분", "사유"),
    ("토지사용동의(O/X)", "X"),
    ("지구단위계획동의(O/X)", "X"),
    ("시행자지정동의(O/X)", "X"),
    ("비고", "← 예시행(삭제 후 작성). 면적·지목은 비우면 자동조회"),
]
_MAX_ROWS = 500  # 업로드 행 상한(과도 방지)
_GEOCODE_CONCURRENCY = 8  # VWorld 재시도(백오프) 보호하에 상향 — 대량 엑셀 처리 가속

# 헤더 자동감지 — 정규화(공백/특수문자 제거·소문자) 후 후보집합 매칭.
_H_ADDR = {"소재지", "소재지주소", "주소", "지번주소", "도로명주소", "address", "소재",
           "토지소재지", "부동산소재지", "물건소재지", "토지소재", "위치", "소재지번", "물건지"}
_H_JIBUN = {"지번", "번지", "지번번지", "jibun", "lot", "지번지", "지번(번지)", "본번부번"}
_H_BCODE = {"법정동코드", "bcode", "법정동코드10자리", "법정동", "동코드", "admcd", "법정동코드bcode10자리"}
_H_PNU = {"pnu", "필지고유번호", "pnu코드", "고유번호", "필지번호", "필지고유번호19자리"}
_H_JIMOK = {"지목", "지목명", "landcategory", "jimok"}
_H_AREA = {"면적", "면적㎡", "면적m2", "토지면적", "area", "areasqm", "대지면적", "공부면적", "지적면적", "실면적", "면적m²"}
_H_OWNER = {"소유구분", "소유", "ownertype", "소유자구분"}
_H_LABEL = {"비고", "라벨", "명칭", "메모", "note", "label"}

# 전각 숫자·하이픈 → 반각(엑셀에 전각으로 입력된 지번 대응). 예: ２２４－１ → 224-1
_FW_DIGITS = str.maketrans("０１２３４５６７８９－‐―ー", "0123456789----")


def _clean_num_str(s: Any) -> str:
    """코드성 문자열(지번·법정동코드·PNU) 정리: 전각→반각, 엑셀이 숫자로 읽어 붙은 '.0' 꼬리 제거.

    엑셀에서 224를 숫자로 저장하면 dtype=str로 읽을 때 '224.0'이 되어 지오코딩이 깨진다.
    면적(소수 가능)에는 적용하지 않는다(별도 _to_float 사용).
    """
    t = str(s or "").strip().translate(_FW_DIGITS)
    t = re.sub(r"\.0+$", "", t)  # '224.0' → '224' (정수형 꼬리만)
    return t
# 동의서 3종(O/X·Y/N·동의/미동의 체크) — 정비/도시개발·지구단위·시행자지정 동의율 산정용.
_H_CONSENT_LAND = {"토지사용동의", "토지사용동의서", "사용동의", "토지동의", "landconsent"}
_H_CONSENT_DISTRICT = {"지구단위계획동의", "지구단위동의", "지구단위계획동의서", "지구단위동의서", "districtconsent"}
_H_CONSENT_OPERATOR = {"시행자지정동의", "시행자지정동의서", "시행자동의", "사업시행자동의", "operatorconsent"}


def _to_consent(v: Any) -> bool | None:
    """동의 체크 셀 → True/False/None(미기재). O·Y·동의·1·V·체크=동의."""
    s = str(v or "").strip().lower()
    if not s:
        return None
    if s in {"o", "ㅇ", "y", "yes", "동의", "1", "v", "✓", "true", "t", "체크", "있음"}:
        return True
    if s in {"x", "ㅌ", "n", "no", "미동의", "0", "false", "f", "없음", "-"}:
        return False
    return None


def _norm(h: Any) -> str:
    return re.sub(r"[\s\-_()·.,/]+", "", str(h or "")).lower()


def _detect_columns(headers: list[Any]) -> dict[str, str | None]:
    """헤더 목록 → 역할별 실제 컬럼명 매핑(첫 일치 우선)."""
    sets = {
        "address": _H_ADDR, "jibun": _H_JIBUN, "bcode": _H_BCODE, "pnu": _H_PNU,
        "jimok": _H_JIMOK, "area": _H_AREA, "owner": _H_OWNER, "label": _H_LABEL,
        "consent_land": _H_CONSENT_LAND, "consent_district": _H_CONSENT_DISTRICT,
        "consent_operator": _H_CONSENT_OPERATOR,
    }
    found: dict[str, str | None] = {k: None for k in sets}
    for h in headers:
        n = _norm(h)
        for role, candidates in sets.items():
            if found[role] is None and n in candidates:
                found[role] = h
                break
    # 2차: 정확매칭 실패 역할은 부분문자열로 추정(헤더가 변형돼도 최대한 매핑).
    #   ★PNU·법정동을 지번보다 먼저 검사해 '필지고유번호'가 '지번'으로 오인되지 않게 한다.
    sub_rules = [
        ("pnu", ("pnu", "고유번호")),
        ("bcode", ("법정동", "동코드")),
        ("address", ("소재", "주소", "위치")),
        ("jibun", ("지번", "번지")),
        ("area", ("면적",)),
        ("owner", ("소유",)),
        ("jimok", ("지목",)),
        ("consent_land", ("토지사용",)),
        ("consent_district", ("지구단위",)),
        ("consent_operator", ("시행자",)),
    ]
    used = {v for v in found.values() if v}
    for h in headers:
        if h in used:
            continue
        n = _norm(h)
        for role, subs in sub_rules:
            if found[role] is None and any(s in n for s in subs):
                found[role] = h
                used.add(h)
                break
    return found


def _detect_header_row(df0: Any) -> int:
    """헤더 키워드(소재지/지번/면적/PNU/법정동/소유/지목)가 가장 많이 등장하는 행을 헤더로 본다.

    실무 토지조서는 1행에 '○○구역 토지조서' 같은 제목·설명행을 두고 머리글이 2~3행에
    오는 경우가 흔하다. pd.read_excel은 기본 1행을 헤더로 쓰므로 그대로면 컬럼이 전부
    오인식된다. 앞 8행을 훑어 헤더다운 행의 0-based 인덱스를 반환(없으면 0=기존 동작).
    """
    sub_keys = ("소재", "지번", "번지", "면적", "법정동", "pnu", "고유번호", "소유", "지목")
    best_i, best_score = 0, -1
    try:
        scan = min(len(df0), 8)
        for i in range(scan):
            cells = [_norm(v) for v in df0.iloc[i].tolist()]
            score = sum(1 for c in cells if c and any(k in c for k in sub_keys))
            if score > best_score:
                best_score, best_i = score, i
            if score >= 3:  # 충분히 헤더다움 → 즉시 채택
                return i
    except Exception:  # noqa: BLE001
        return 0
    return best_i if best_score > 0 else 0


def _looks_transposed(df0: Any) -> bool:
    """1열(A열)에 표준 필드라벨(소재지·지번·PNU·면적·지목 등)이 여러 개 나타나면 전치(세로형) 의심.

    ★전치 양식은 필드명이 '행'이 아니라 '열'로 나열돼 _detect_header_row(행 단위 스캐너)가
    필드라벨 행 하나를 우연히 헤더로 오인(그 행의 라벨셀 1개만 매칭)하기 쉽다 — 이 경우
    규칙기반이 address 등을 "찾은 것처럼" 보여 S1 LLM 트리거가 새지 않는 사고를 막기 위한
    별도 신호(1열에 라벨이 3개 이상 반복되면 명백히 전치 의심 → 신뢰도 낮음으로 간주).
    """
    sub_keys = ("소재", "지번", "번지", "면적", "법정동", "pnu", "고유번호", "소유", "지목")
    try:
        col0 = [_norm(v) for v in df0.iloc[:12, 0].tolist()]
    except Exception:  # noqa: BLE001
        return False
    hits = sum(1 for c in col0 if c and any(k in c for k in sub_keys))
    return hits >= 3


def _pnu_from_bcode(bcode: str, jibun: str) -> str | None:
    """bcode(10)+지번 → PNU(19). 구조: bcode(10)+대지구분(1:산이면2)+본번(4)+부번(4)."""
    b = re.sub(r"\D", "", str(bcode or ""))
    if len(b) < 10:
        return None
    b = b[:10]
    m = re.search(r"(산)?\s*(\d+)(?:-(\d+))?", str(jibun or ""))
    if not m:
        return None
    san = "2" if m.group(1) else "1"
    return f"{b}{san}{m.group(2).zfill(4)}{(m.group(3) or '0').zfill(4)}"


def _to_float(v: Any) -> float | None:
    try:
        f = float(re.sub(r"[,\s㎡m²]", "", str(v)))
        return f if f > 0 else None
    except (ValueError, TypeError):
        return None


def _expand_merged_cells(raw: bytes, df: Any, header_row: int = 0) -> Any:
    """엑셀 병합 셀의 좌상단 값을 병합범위 전체 셀에 채운다(forward-fill).

    토지조서 엑셀은 한 필지(지번)에 소유자가 여럿이면 여러 행을 두고 지번·소재지 칸을
    세로로 '병합'하는 경우가 흔하다. pandas(read_excel)는 병합 시 좌상단 셀에만 값을
    넣고 나머지 행은 NaN(빈값)으로 둔다 → 병합된 '지번'이 소실돼 그 행들이 소재지(동)만
    남고, 번지가 없어 '동 단위 수렴'으로 보완필요 처리되던 근본버그.
    openpyxl로 병합범위를 읽어 좌상단 값을 같은 범위의 모든 데이터 셀(빈칸)에 채운다.

    header_row(0-based)는 머리글 행 위치. 데이터 첫 행은 엑셀 (header_row+2)행 = df index 0.
    """
    try:
        from openpyxl import load_workbook

        # read_only=False 여야 merged_cells.ranges가 채워진다. data_only=True=캐시값(수식 대비).
        wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=False)
        ws = wb.worksheets[0]
        nrows, ncols = len(df), df.shape[1]
        base = header_row + 2  # 엑셀(1-based) 데이터 첫 행 → df index 0
        for rng in list(ws.merged_cells.ranges):
            top = ws.cell(row=rng.min_row, column=rng.min_col).value
            if top is None or str(top).strip() == "":
                continue
            for r in range(rng.min_row, rng.max_row + 1):
                df_row = r - base
                if df_row < 0 or df_row >= nrows:
                    continue  # 헤더/제목 병합·범위초과는 무시
                for c in range(rng.min_col, rng.max_col + 1):
                    df_col = c - 1  # 엑셀 1열(A)=df 0열
                    if df_col < 0 or df_col >= ncols:
                        continue
                    cur = df.iat[df_row, df_col]
                    # 빈칸(NaN·""·"nan")만 채우고 기존 값은 보존.
                    if cur is None or str(cur).strip() == "" or str(cur).strip().lower() == "nan":
                        df.iat[df_row, df_col] = str(top)
    except Exception as e:  # noqa: BLE001
        logger.warning("excel_merged_expand_failed", error=str(e)[:120])
    return df


_LLM_ROLES = {
    "address", "jibun", "bcode", "pnu", "jimok", "area", "owner", "label",
    "consent_land", "consent_district", "consent_operator",
}
# 표 구조 시그니처(시트목록+현재시트+헤더)별 LLM 분류 결과 캐시 — 동일 양식 반복 업로드 시
# LLM 재호출/비용 방지(_llm_detect_columns 시절의 헤더전용 캐시를 구조질의로 확장).
_STRUCT_CACHE: dict[tuple, dict[str, Any]] = {}
_LLM_REVERIFY_CONCURRENCY = 5  # 행별 재질의 동시 LLM 호출 상한(대량 실패행 폭주 방지)
_MAX_REVERIFY_ROWS = 60  # 패스당 재질의 대상 상한(양식이 근본적으로 안 맞으면 LLM 낭비 방지)
# 지번 형식 게이트 — '산' 지번 포함, 본번(-부번) 형태만 정상(그 외는 needs_review 후보).
_JIBUN_RE = re.compile(r"^산?\d+(-\d+)?$")
_UNRESOLVED_STATUSES = ("failed", "ambiguous", "need_geocode")
_ISSUE_LABELS = {
    "jibun_format": "지번 형식 불명확(숫자·산 지번 형태가 아님)",
    "pnu_format": "PNU 형식 오류(19자리 숫자가 아님)",
    "unresolved_failed": "주소 미해소(failed)",
    "unresolved_ambiguous": "동명이의 등으로 모호(ambiguous)",
    "unresolved_need_geocode": "주소 미해소(지오코딩 결과 없음)",
}


def _row_issues(p: dict[str, Any]) -> list[str]:
    """행별 결정론 검증 게이트 — 지번 형식·PNU 19자리·미해소 상태(failed/ambiguous/need_geocode)."""
    issues: list[str] = []
    jb = p.get("jibun")
    if jb and not _JIBUN_RE.match(jb):
        issues.append("jibun_format")
    pnu = p.get("pnu")
    if pnu and (len(pnu) != 19 or not pnu.isdigit()):
        issues.append("pnu_format")
    st = p.get("status")
    if st in _UNRESOLVED_STATUSES:
        issues.append(f"unresolved_{st}")
    return issues


def _sheet_previews_xlsx(raw: bytes, max_rows: int = 15) -> dict[str, list[list[str]]]:
    """xlsx 전체 시트명별 첫 N행 미리보기(값만 문자열화) — LLM 시트선택 판단용. 실패 시 빈dict."""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    except Exception:  # noqa: BLE001
        return {}
    out: dict[str, list[list[str]]] = {}
    try:
        for name in wb.sheetnames:
            ws = wb[name]
            preview: list[list[str]] = []
            for i, r in enumerate(ws.iter_rows(values_only=True)):
                if i >= max_rows:
                    break
                preview.append(["" if v is None else str(v)[:60] for v in r])
            out[name] = preview
    except Exception:  # noqa: BLE001
        return out
    return out


async def _llm_analyze_structure(
    sheet_previews: dict[str, list[list[str]]],
    current_sheet: str,
    headers: list[str],
    sample_rows: list[dict],
) -> tuple[dict[str, Any], bool]:
    """규칙기반 신뢰도가 낮을 때 LLM에게 표 구조 전체를 ★한 번의 JSON 질의로★ 물어본다
    (시트선택·헤더행/데이터범위·전치판정·복합셀분해·컬럼역할 — 중복 프롬프트 금지, 기존
    _llm_detect_columns를 이 확장 함수가 흡수).

    반환: (LLM 원시 제안 dict, LLM 실제 호출 여부). ★환각 차단은 여기서 하지 않는다 — 시트가
    재선택되면 유효 헤더 집합 자체가 바뀌므로, 실존 검증(값이 실제 헤더/시트/행범위에 있는지)은
    최종 시트·헤더가 확정된 뒤 호출측(parse())이 수행한다. 캐시: 시트+헤더 시그니처로 재호출 방지.
    LLM 토큰은 _record_llm_billing 계측(service=parcel_excel_structure_detect).
    """
    if not headers and not sheet_previews:
        return {}, False
    cache_key = (tuple(sorted(sheet_previews.keys())), current_sheet, tuple(headers))
    if cache_key in _STRUCT_CACHE:
        return dict(_STRUCT_CACHE[cache_key]), False
    try:
        from langchain_core.messages import HumanMessage, SystemMessage

        from app.services.ai.llm_provider import get_llm
    except Exception:  # noqa: BLE001
        return {}, False
    try:
        llm = get_llm(timeout=45, max_tokens=800)
    except Exception:  # noqa: BLE001
        logger.info("엑셀 LLM 구조분석 생략 — 사용가능 LLM 키 없음(규칙기반 폴백)")
        return {}, False
    import json as _json
    sample = [{str(k): str(v)[:40] for k, v in r.items()} for r in sample_rows[:3]]
    sys = (
        "너는 한국 부동산 '토지조서' 엑셀의 표 구조를 분석하는 에이전트다. 비표준 양식(다중시트·"
        "세로형·복합셀 등)의 실제 데이터 표 구조를 판단해 JSON만 출력한다."
    )
    human = (
        f"현재 시트: {current_sheet}\n"
        f"시트별 미리보기(각 최대15행, 행은 0-based): {sheet_previews}\n"
        f"현재 헤더: {headers}\n샘플행(최대3): {sample}\n\n"
        "다음 JSON 키를 판단되는 것만 포함해 출력(모르면 생략, 값이 확실하지 않으면 생략):\n"
        "- sheet_name: 실제 토지조서 데이터가 있는 시트명(현재 시트가 맞으면 생략)\n"
        "- header_row: 머리글 행의 0-based 인덱스(미리보기 기준 행 번호)\n"
        "- data_start_row / data_end_row: 실제 데이터 행 범위(0-based, 미리보기 기준 행 번호)\n"
        "- is_transposed: 필드명이 세로(첫 열)로 나열된 '전치' 양식이면 true\n"
        "- compound_cell: 주소+지번+지목+면적이 한 컬럼에 합쳐진 경우 "
        '{"column":"실제컬럼명","regex":"명명그룹(?P<addr>..)(?P<jibun>..)(?P<jimok>..)(?P<area>..)를 '
        '포함한 파이썬 정규식(필요한 그룹만)"}\n'
        "- columns: 표준역할(address/jibun/bcode/pnu/jimok/area/owner/label/"
        "consent_land/consent_district/consent_operator)별 실제 컬럼명\n"
        'JSON 객체만 출력. 예: {"header_row":2,"columns":{"address":"소재지"}}'
    )
    try:
        resp = await llm.ainvoke([SystemMessage(content=sys), HumanMessage(content=human)])
        text = resp.content if hasattr(resp, "content") else str(resp)
        m = re.search(r"\{.*\}", str(text), re.S)
        if not m:
            return {}, True
        data = _json.loads(m.group(0))
        if not isinstance(data, dict):
            return {}, True
    except Exception as e:  # noqa: BLE001
        logger.warning("엑셀 LLM 구조분석 실패: %s", str(e)[:160])
        return {}, False

    # ★LLM 토큰 사용량을 표준 계측 경로(_record_llm_billing→llm_usage_log)에 best-effort 기록.
    try:
        usage = getattr(resp, "usage_metadata", None) or {}
        in_tok = int(usage.get("input_tokens", 0) or 0)
        out_tok = int(usage.get("output_tokens", 0) or 0)
        if in_tok or out_tok:
            from app.services.ai.base_interpreter import _record_llm_billing
            model = getattr(llm, "model", None) or getattr(llm, "model_name", "") or "unknown"
            await _record_llm_billing(str(model), in_tok, out_tok, service="parcel_excel_structure_detect")
    except Exception:  # noqa: BLE001
        pass

    if len(_STRUCT_CACHE) > 256:  # 장수명 워커 무한증가 방지(상한 초과 시 비움)
        _STRUCT_CACHE.clear()
    _STRUCT_CACHE[cache_key] = dict(data)
    return data, True


async def _llm_reverify_row(raw_cells: dict[str, str], issues: list[str]) -> tuple[dict[str, str], bool]:
    """S3 선별 재질의 — 결정론 게이트에 실패한 행 1건의 원본 셀 텍스트만 제시해 역할별 값을 묻는다.

    환각 차단: LLM이 답한 값이 그 행의 원본 셀 텍스트 안에 '실제로 등장하는 부분문자열'일 때만
    채택(새 값 생성 금지 — 새로 만들어낸 값은 폐기). 반환: (역할→후보값, LLM 실제 호출 여부).
    """
    if not raw_cells:
        return {}, False
    try:
        from langchain_core.messages import HumanMessage, SystemMessage

        from app.services.ai.llm_provider import get_llm
    except Exception:  # noqa: BLE001
        return {}, False
    try:
        llm = get_llm(timeout=30, max_tokens=300)
    except Exception:  # noqa: BLE001
        return {}, False
    import json as _json
    sys = (
        "너는 토지조서 엑셀 한 행에서 문제된 값을 원문 그대로 찾아주는 보조 에이전트다. "
        "제시된 원본 셀 텍스트 안에 실제로 존재하는 부분문자열만 답한다(새로 만들어내지 않는다). "
        "JSON만 출력한다."
    )
    human = (
        f"원본 셀(컬럼:값): {raw_cells}\n검증 실패 사유: {issues}\n\n"
        "역할(address/jibun/jimok/area 중 해당하는 것만): 원본 셀 값 안에 실제로 등장하는 "
        '정확한 부분문자열로만 답하라(추정·생성 금지). 예: {"jibun":"224-1"}'
    )
    try:
        resp = await llm.ainvoke([SystemMessage(content=sys), HumanMessage(content=human)])
        text = resp.content if hasattr(resp, "content") else str(resp)
        m = re.search(r"\{.*\}", str(text), re.S)
        if not m:
            return {}, True
        data = _json.loads(m.group(0))
    except Exception as e:  # noqa: BLE001
        logger.warning("엑셀 LLM 행 재질의 실패: %s", str(e)[:160])
        return {}, False

    try:
        usage = getattr(resp, "usage_metadata", None) or {}
        in_tok = int(usage.get("input_tokens", 0) or 0)
        out_tok = int(usage.get("output_tokens", 0) or 0)
        if in_tok or out_tok:
            from app.services.ai.base_interpreter import _record_llm_billing
            model = getattr(llm, "model", None) or getattr(llm, "model_name", "") or "unknown"
            await _record_llm_billing(str(model), in_tok, out_tok, service="parcel_excel_row_reverify")
    except Exception:  # noqa: BLE001
        pass

    # 환각 차단: 원본 셀 전체 텍스트에 실제로 등장하는 부분문자열일 때만 채택.
    haystack = " ".join(str(v) for v in raw_cells.values())
    out: dict[str, str] = {}
    for role, val in (data or {}).items():
        if role in {"address", "jibun", "jimok", "area"} and isinstance(val, str) and val.strip() and val in haystack:
            out[role] = val.strip()
    return out, True


def build_template_xlsx() -> bytes:
    """토지조서 다필지 업로드용 표준 엑셀 양식(예시행 + 안내 시트) 생성."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "토지조서"
    head_fill = PatternFill("solid", fgColor="1F4E79")
    head_font = Font(bold=True, color="FFFFFF", size=11)
    for col, (name, _ex) in enumerate(TEMPLATE_COLUMNS, start=1):
        c = ws.cell(row=1, column=col, value=name)
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[c.column_letter].width = max(12, len(name) + 4)
    # 예시 2행(작성 가이드) — ★면적·지목은 비워둠(자동조회 대상). 예시값 미삭제 시 면적 오적재 방지.
    ws.append([ex for _n, ex in TEMPLATE_COLUMNS])
    ws.append(["예시2(삭제)", "서울특별시 강남구 역삼동 737", "737", "1168010100", "", "", "", "사유", "O", "X", "X", "← 예시행(삭제 후 작성)"])
    for r in (2, 3):
        for col in range(1, len(TEMPLATE_COLUMNS) + 1):
            ws.cell(row=r, column=col).font = Font(italic=True, color="888888")

    # 안내 시트.
    g = wb.create_sheet("작성안내")
    notes = [
        "■ PropAI 다필지 토지조서 업로드 양식",
        "",
        "1) [소재지(주소)] 는 필수입니다. 나머지는 비워도 됩니다(있으면 정확도↑).",
        "2) PNU(19자리)를 알면 그 열에 입력하세요 — 가장 정확합니다.",
        "3) PNU가 없으면 [법정동코드(10자리)] + [지번] 으로 자동 구성됩니다.",
        "4) 둘 다 없으면 [소재지(주소)] 를 좌표·필지로 자동 조회합니다(다소 느릴 수 있음).",
        "5) '산' 지번은 지번 칸에 '산12-3' 처럼 적으세요.",
        "6) 면적은 비워두면 공부상(VWorld) 면적을 자동 조회합니다. 직접 입력 시 공부상과 대조해 "
        "크게 다르면(1.5배↑) 공부상으로 보정하고 경고합니다. ㎡ 숫자만(콤마 가능).",
        "7) [토지사용동의]·[지구단위계획동의]·[시행자지정동의] 는 O(동의) / X(미동의)로 표기하세요.",
        "   → 정비·도시개발사업의 동의율(소유자 동의 비율) 산정과 시행자 지정요건 판정에 활용됩니다.",
        f"8) 한 번에 최대 {_MAX_ROWS}필지까지 업로드됩니다.",
        "",
        "※ 예시행(2~3행)은 삭제하고 실제 필지를 입력하세요.",
    ]
    for i, t in enumerate(notes, start=1):
        cell = g.cell(row=i, column=1, value=t)
        if i == 1:
            cell.font = Font(bold=True, size=13)
    g.column_dimensions["A"].width = 70

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class ParcelExcelService:
    """업로드 엑셀 → 필지 목록 추출."""

    async def parse(self, raw: bytes, filename: str, use_llm: bool = True) -> dict[str, Any]:
        """업로드 엑셀/CSV → 필지 목록(+검증 리포트) 추출.

        use_llm(기본 True=기존 동작 보존): False면 규칙기반만 동작(LLM 0호출·무과금·실패는
        정직 사유표기). True면 규칙기반 신뢰도가 낮을 때만 구조 LLM(S1)·행별 재질의(S3)를 사용.
        """
        import pandas as pd

        name = (filename or "").lower()
        is_csv = name.endswith(".csv")
        try:
            if is_csv:
                # ★CSV 인코딩 폴백 — 한국 엑셀이 내보낸 CSV는 CP949/EUC-KR가 흔하다(UTF-8만
                #   시도하면 한글이 깨지거나 읽기 실패).
                df0 = None
                for enc in ("utf-8-sig", "cp949", "euc-kr", "latin1"):
                    try:
                        df0 = pd.read_csv(io.BytesIO(raw), dtype=str, keep_default_na=False,
                                          header=None, encoding=enc)
                        break
                    except (UnicodeDecodeError, ValueError):
                        continue
                if df0 is None:
                    return {"error": "CSV 인코딩을 인식하지 못했습니다(UTF-8/CP949). 엑셀(.xlsx)로 저장해 올려보세요.", "parcels": []}
                current_sheet = "CSV"
            else:
                # ★머리글 행 자동탐지 — 제목·설명행이 앞에 있는 양식이면 헤더가 1행이 아니다.
                df0 = pd.read_excel(io.BytesIO(raw), dtype=str, engine="openpyxl", header=None)
                current_sheet = ""
                try:
                    from openpyxl import load_workbook
                    _wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
                    current_sheet = _wb.sheetnames[0] if _wb.sheetnames else ""
                except Exception:  # noqa: BLE001
                    pass
        except Exception as e:  # noqa: BLE001
            return {"error": f"엑셀/CSV를 읽지 못했습니다: {str(e)[:120]} — 엑셀(.xlsx)로 저장해 다시 시도해 주세요.", "parcels": []}

        def _rebuild(df0_: Any, hdr_: int) -> tuple[Any, list[str]]:
            """raw grid(header=None)+헤더행 인덱스 → (데이터프레임, 헤더목록). 재선택/전치 후 공용 재사용."""
            d = df0_.iloc[hdr_ + 1:].reset_index(drop=True)
            d.columns = [str(v) if v is not None and str(v).strip() else f"col{i}"
                         for i, v in enumerate(df0_.iloc[hdr_].tolist())]
            return d, [str(h) for h in d.columns]

        hdr = _detect_header_row(df0)
        df, headers = _rebuild(df0, hdr)
        if not is_csv:
            # ★병합 셀 forward-fill — 한 지번을 여러 행에 '병합'한 토지조서에서 병합
            #   연속행의 지번이 소실(NaN)돼 보완필요로 빠지던 근본버그를 차단(지번 기준 복원).
            df = _expand_merged_cells(raw, df, header_row=hdr)
            headers = [str(h) for h in df.columns]
        df = df.fillna("")
        cols = _detect_columns(headers)
        engine = "rule"
        llm_used = False
        structure_notes: list[str] = []

        def _valid_ratio(frame: Any, colmap: dict) -> float:
            keys = [colmap.get(r) for r in ("address", "pnu", "bcode") if colmap.get(r)]
            if not keys or len(frame) == 0:
                return 0.0
            good = sum(1 for _, r in frame.iterrows() if any(str(r.get(k, "")).strip() for k in keys))
            return good / len(frame)

        # ── S1: 표 구조 인식(LLM, use_llm 게이트 하) ──
        #   규칙기반이 필수컬럼(주소/PNU/법정동코드)을 못 찾거나, 컬럼은 찾았지만 유효행<50%
        #   (다중시트·전치·오탐 헤더행 등 구조 문제 의심)일 때만 LLM에게 표 구조 전체를 질의한다.
        need_structure_llm = (
            (not cols["address"] and not cols["pnu"] and not cols["bcode"])
            or (len(df) > 0 and _valid_ratio(df, cols) < 0.5)
            or _looks_transposed(df0)
        )
        if need_structure_llm and use_llm:
            sheet_previews = (
                _sheet_previews_xlsx(raw) if not is_csv
                else {current_sheet: [[str(v) for v in r] for r in df0.head(15).values.tolist()]}
            )
            struct, called = await _llm_analyze_structure(
                sheet_previews=sheet_previews, current_sheet=current_sheet,
                headers=headers, sample_rows=df.to_dict("records")[:3],
            )
            llm_used = llm_used or called
            if struct:
                # 1) 시트 재선택(xlsx만 — 실존 시트명일 때만 채택. 환각 차단).
                chosen_sheet = struct.get("sheet_name")
                if (not is_csv and isinstance(chosen_sheet, str) and chosen_sheet in sheet_previews
                        and chosen_sheet != current_sheet):
                    try:
                        df0 = pd.read_excel(io.BytesIO(raw), dtype=str, engine="openpyxl",
                                             header=None, sheet_name=chosen_sheet)
                        current_sheet = chosen_sheet
                        structure_notes.append(f"LLM이 '{chosen_sheet}' 시트를 토지조서 데이터로 재선택")
                    except Exception:  # noqa: BLE001
                        pass

                # 2) 전치(세로형) 판정 — 결정론 전치(행↔열 swap) 후 재파싱.
                if struct.get("is_transposed") is True:
                    df0 = df0.T.reset_index(drop=True)
                    structure_notes.append("LLM이 세로형(전치) 양식으로 판단 — 전치 후 재파싱")

                # 3) 헤더행 재적용(실존 범위 검증 — 미달/범위초과는 규칙기반 유지).
                hdr_llm = struct.get("header_row")
                hdr = hdr_llm if isinstance(hdr_llm, int) and 0 <= hdr_llm < len(df0) else _detect_header_row(df0)
                df, headers = _rebuild(df0, hdr)
                # ★전치·시트재선택 후에는 원본 병합범위 좌표계가 어긋나 merge-expand는 건너뜀(과도적용 방지).
                df = df.fillna("")
                cols = _detect_columns(headers)

                # 4) 데이터 범위(data_start/end_row) — 실존 범위 내일 때만 슬라이스(엑셀 1-based 행 매핑).
                ds, de = struct.get("data_start_row"), struct.get("data_end_row")
                if isinstance(ds, int) and isinstance(de, int) and 0 <= ds <= de:
                    base = hdr + 2  # 엑셀 1-based 데이터 첫 행
                    keep = [i for i in range(len(df)) if ds <= (base + i) <= de]
                    if keep:
                        df = df.iloc[keep].reset_index(drop=True)
                        structure_notes.append(f"LLM 판정 데이터범위({ds}~{de}행)로 슬라이스")

                # 5) 복합셀 분해 — 채택 전 샘플 매치율 ≥60% 검증(미달 폐기).
                compound = struct.get("compound_cell")
                if isinstance(compound, dict):
                    c_col, c_rx = compound.get("column"), compound.get("regex")
                    if isinstance(c_col, str) and c_col in headers and isinstance(c_rx, str):
                        try:
                            rx = re.compile(c_rx)
                        except re.error:
                            rx = None
                        _GMAP = {"addr": "address", "jibun": "jibun", "jimok": "jimok", "area": "area"}
                        groups = [g for g in _GMAP if rx and g in (rx.groupindex or {})]
                        if rx and groups:
                            vals = [str(v) for v in df[c_col].tolist()]
                            nonblank = [v for v in vals if v.strip()]
                            matched = sum(1 for v in nonblank if rx.search(v))
                            rate = matched / len(nonblank) if nonblank else 0.0
                            if rate >= 0.6:
                                for g in groups:
                                    role = _GMAP[g]
                                    if cols.get(role):
                                        continue  # 이미 매핑된 역할은 덮지 않음
                                    vcol = f"__cc_{g}"
                                    df[vcol] = [
                                        (m.group(g) or "") if (m := rx.search(v)) else "" for v in vals
                                    ]
                                    cols[role] = vcol
                                structure_notes.append(f"복합셀 '{c_col}' 분해 적용(매치율 {rate:.0%})")
                            else:
                                structure_notes.append(f"복합셀 '{c_col}' 분해 폐기(매치율 {rate:.0%}<60%)")

                # 6) 컬럼 역할(실존 헤더 검증 — 환각 차단: 실제 헤더에 존재할 때만 채택).
                llm_cols = struct.get("columns")
                if isinstance(llm_cols, dict):
                    for role, col in llm_cols.items():
                        if role in _LLM_ROLES and isinstance(col, str) and col in headers and not cols.get(role):
                            cols[role] = col
                if structure_notes or (isinstance(llm_cols, dict) and llm_cols):
                    engine = "rule+llm"

        if not cols["address"] and not cols["pnu"] and not cols["bcode"]:
            return {
                "error": "필수 컬럼을 찾지 못했습니다 — 최소 [소재지(주소)] 또는 [PNU] 또는 [법정동코드]가 필요합니다. 표준 양식을 내려받아 작성하거나, 헤더에 '소재지/지번/면적' 등을 명시해 주세요.",
                "detected_columns": cols, "headers": headers, "parcels": [],
            }

        rows = df.to_dict("records")[:_MAX_ROWS]
        parcels: list[dict[str, Any]] = []
        raw_rows: list[dict] = []  # parcels와 같은 인덱스로 원본 셀 보관(S3 재질의용 — 응답에는 미포함)
        need_geocode: list[int] = []  # 주소만 있는 행 인덱스(지오코딩 대상)
        _cum_area = 0.0
        declared_total: float | None = None
        excluded_n = 0  # 합계/소계 등 표 내 집계행으로 감지되어 제외된 행 수

        def _g(row: dict, role: str) -> str:
            col = cols.get(role)
            return str(row.get(col, "")).strip() if col else ""

        _SUMMARY_KEYS = {"합계", "소계", "총계", "총합", "계", "합산", "total", "sum"}

        def _is_summary_token(s: str) -> bool:
            return bool(s) and (s in _SUMMARY_KEYS or s.startswith(("합계", "소계", "총계")))

        for row in rows:
            # 전각숫자→반각, 숫자형 '224.0'→'224' 정규화(코드성 칸). 면적은 _to_float가 별도 처리.
            address = _g(row, "address").translate(_FW_DIGITS)
            jibun = _clean_num_str(_g(row, "jibun"))
            bcode = re.sub(r"\D", "", _clean_num_str(_g(row, "bcode")))
            pnu_raw = re.sub(r"\D", "", _clean_num_str(_g(row, "pnu")))
            area_raw = _g(row, "area")
            area = _to_float(area_raw)
            jimok = _g(row, "jimok") or None
            owner = _g(row, "owner") or None
            label = _g(row, "label") or None
            # 동의서 3종 체크(O/X·동의/미동의) — 정비·도시개발 동의율 산정에 활용.
            consent_land = _to_consent(_g(row, "consent_land"))
            consent_district = _to_consent(_g(row, "consent_district"))
            consent_operator = _to_consent(_g(row, "consent_operator"))
            if not (address or pnu_raw or bcode):
                continue  # 완전 빈 행 스킵

            # S3 집계게이트: 합계·소계 등 표 내 집계행 감지 — ①키워드(주소/비고칸)
            #   ②면적=상위행 누적합(±1%) 중 하나면 필지가 아닌 집계행으로 보고 제외.
            _ax = address.replace(" ", "")
            _lb = (label or "").replace(" ", "")
            is_summary = not jibun and not pnu_raw and not bcode and (_is_summary_token(_ax) or _is_summary_token(_lb))
            if not is_summary and area and _cum_area > 0 and not jibun and not pnu_raw:
                if abs(area - _cum_area) / _cum_area <= 0.01:
                    is_summary = True
            if is_summary:
                if area:
                    declared_total = area  # 합계행 선언값 — 추출 면적합과 ±1% 대조(S4)
                excluded_n += 1
                continue

            pnu = pnu_raw if len(pnu_raw) == 19 else (_pnu_from_bcode(bcode, jibun) if bcode else None)
            status = "ok" if pnu else ("need_geocode" if address else "failed")
            p: dict[str, Any] = {
                "address": address or None, "jibun": jibun or None,
                "bcode": (bcode[:10] if len(bcode) >= 10 else None),
                "pnu": pnu, "area_sqm": area, "jimok": jimok,
                "owner_type": owner, "label": label,
                # 동의서 3종(True=동의/False=미동의/None=미기재).
                "consent_land": consent_land,
                "consent_district": consent_district,
                "consent_operator": consent_operator,
                # ★자동 보강 필드(주소만 입력해도 PNU 확보 후 채워짐). 무자료=None(가짜 금지).
                "zone_type": None, "official_price_per_sqm": None,
                # 소유자·권리관계는 공공API로 확보 불가 → 등기부등본 열람/발급 필요(사용자 안내).
                "registry_needed": not owner,
                "status": status,
            }
            if area_raw.strip() and area is None:
                # S3 게이트: 입력 면적이 있었지만 형식불량(음수/0/숫자아님) — 정직 기록(공부상 자동보강 대상).
                p["_corrections"] = [{
                    "field": "area_sqm", "before": area_raw, "after": None,
                    "reason": f"입력 면적 '{area_raw}' 형식 불량(음수/0/숫자아님) — 공부상 자동조회로 대체",
                }]
            if status == "need_geocode":
                need_geocode.append(len(parcels))
            parcels.append(p)
            raw_rows.append(row)
            if area:
                _cum_area += area

        # ③ 주소만 있는 행 — VWorld 지오코딩으로 좌표·PNU 확보(법정동코드 몰라도 됨).
        if need_geocode:
            await self._geocode_fill(parcels, need_geocode)

        # ④ ★데이터 보완·보강 — PNU 확보된 필지의 빈 칸(면적·지목·용도지역·공시지가)을
        #    토지특성(NED)으로 자동 채운다. 사용자는 주소만 적으면 나머지는 시스템이 조회.
        await self._enrich_fill(parcels)

        # ★bad-data 자가치유: 엑셀의 어긋난 법정동코드로 만든 PNU 라 용도지역을 못 불러온 행을
        #   주소로 재해소(올바른 PNU 재확보). 중복PNU 탐지 전에 수행해 치유된 PNU 가 dedup 에 반영되게 함.
        await self._heal_stale_pnu(parcels)

        # C3: 중복 PNU 탐지 — 서로 다른 행이 같은 PNU로 해석되면 거의 항상 오매칭(동명이의·
        #     번지누락). 2건 이상 동일 PNU는 ambiguous로 강등(가짜 확신 차단). 1필지만 같은 PNU면 정상.
        def _input_key(p: dict) -> str:
            # 입력 주소+지번을 공백 제거해 비교(같은 필지를 가리키는 행인지 판정).
            a = re.sub(r"\s+", "", (p.get("address") or ""))
            j = re.sub(r"\s+", "", (p.get("jibun") or ""))
            return f"{a}|{j}"

        pnu_counts = Counter(p["pnu"] for p in parcels if p.get("pnu"))
        dup_pnus = {pnu for pnu, c in pnu_counts.items() if c >= 2}
        dup_warning = None
        if dup_pnus:
            dup_n = 0  # 오매칭으로 강등한 행
            co_n = 0   # 같은 필지의 공유지분 등으로 정상 유지한 행
            for pnu in dup_pnus:
                group = [p for p in parcels if p.get("pnu") == pnu]
                keys = {_input_key(p) for p in group}
                if len(keys) == 1:
                    # 동일 입력(주소+지번)이 같은 PNU → 같은 필지의 공유지분·다소유자 행
                    #   (지번 병합 토지조서 등)으로 '의도된 중복' → 강등하지 않고 표시만.
                    for p in group:
                        p["co_owner"] = True
                    group[0]["co_owner_rep"] = True  # 대표 행(면적 1회 반영 기준)
                    # ★면적 중복합산 방지: 그룹의 면적이 모두 같으면(대표 필지면적이 병합/자동보강
                    #   으로 행마다 복제된 것) 대표 행만 남기고 나머지는 None — 부지면적이 N배로
                    #   과대 합산되는 것을 막는다. 행마다 면적이 다르면 소유자별 '지분면적'으로 보고
                    #   그대로 보존한다(합이 곧 필지면적이라 정상).
                    nonnull = [round(p["area_sqm"], 1) for p in group if p.get("area_sqm")]
                    if nonnull and len(set(nonnull)) == 1:
                        for p in group[1:]:
                            p["area_sqm"] = None
                            p["reason"] = (p.get("reason") or
                                           "동일 필지의 공유지분(다소유자) 행 — 면적은 대표 행에 "
                                           "1회만 반영(부지면적 중복합산 방지).")
                    co_n += len(group)
                    continue
                # 서로 다른 입력이 같은 PNU로 수렴 → 번지누락/동명이의 오매칭 → 강등.
                for p in group:
                    p["status"] = "ambiguous"
                    p["reason"] = (p.get("reason") or
                                   "여러 필지가 동일 PNU로 해석됨 — 번지 누락/동명이의 가능. "
                                   "시·군·구+번지 포함 정확한 주소로 보완하세요.")
                    p["pnu"] = None  # 오매칭 PNU 제거(잘못된 면적/용도 보강 차단)
                    p["area_sqm"] = None  # 오매칭 면적 제거(동일평수 오표시 차단)
                    dup_n += 1
            parts = []
            if dup_n:
                parts.append(f"동일 PNU로 중복 해석된 {dup_n}필지를 '모호'로 표기(오매칭 방지)")
            if co_n:
                parts.append(f"같은 필지의 공유지분 등 {co_n}행은 정상 등록")
            dup_warning = (" · ".join(parts) + ".") if parts else None

        # ── S3: 반복검증 루프 — 결정론 게이트에 실패한 행만 선별 재질의(최대 2회) ──
        reverify = await self._reverify_loop(parcels, raw_rows, use_llm)
        llm_used = llm_used or reverify["llm_called"]

        # ── S4: 행별 최종분류(verified/corrected/needs_review) + 검증 리포트(additive) ──
        verification_report = self._build_verification_report(
            parcels, declared_total=declared_total, excluded_n=excluded_n,
            dup_warning=dup_warning, llm_used=llm_used, passes=reverify["passes"],
            structure_notes=structure_notes,
        )

        ok = sum(1 for p in parcels if p["status"] == "ok")
        enriched = sum(1 for p in parcels if p.get("zone_type") or p.get("official_price_per_sqm"))
        return {
            "parcels": parcels,
            "total_rows": len(rows),
            "parsed_count": len(parcels),
            "resolved_count": ok,
            "enriched_count": enriched,
            "detected_columns": cols,
            "column_engine": engine,  # rule | rule+llm (LLM 에이전트가 구조/컬럼 분류에 관여했는지)
            "duplicate_pnu_warning": dup_warning,
            "examples": parcels[:3],
            # 소유자·권리관계(근저당·지상권 등)는 공공API 미제공 → 등기부등본 열람/발급으로 확보.
            "registry_guidance": {
                "needed_count": sum(1 for p in parcels if p.get("registry_needed")),
                "route": "/land-schedule",
                "message": ("소유자·권리관계(근저당·지상권 등)는 공공데이터로 확인할 수 없습니다 — "
                            "토지조서 화면의 '등기부등본 열람/발급'으로 확보하세요."),
            },
            # ★additive — 기존 키(parcels/status 등)는 불변. 확정분(verified+corrected)만
            #   기본 주입 대상(각 parcel의 injectable 플래그로 표시).
            "verification_report": verification_report,
            "note": (f"{len(parcels)}필지 인식 · PNU 확정 {ok}건 · 면적/용도/공시지가 자동보강 {enriched}건. "
                     + ("🤖 비표준 양식이라 LLM 에이전트가 구조를 자동 분석했습니다. " if engine == "rule+llm" else "")
                     + "주소·지번만 입력해도 좌표·PNU·면적·용도지역·공시지가는 자동 수집됩니다. "
                     "소유자·권리관계는 등기부등본 열람/발급으로 확인하세요. "
                     "status=failed 행은 주소를 보완해 주세요(가짜값 없음)."),
        }

    async def _reverify_loop(
        self, parcels: list[dict[str, Any]], raw_rows: list[dict], use_llm: bool,
    ) -> dict[str, Any]:
        """S3: 결정론 게이트(_row_issues)에 실패한 행만 선별 재질의(최대 2회).

        패스마다 남은 실패행만 대상으로 재평가(고쳐진 행은 다음 패스에서 자연히 제외).
        주소/지번이 교정돼 미해소 상태였던 행은 재지오코딩까지 재시도(기존 인프라 재사용).
        use_llm=False면 즉시 0회 반환(LLM 0호출 보장).
        """
        if not use_llm:
            return {"passes": 0, "llm_called": False}
        called = False
        pass_n = 0
        sem = asyncio.Semaphore(_LLM_REVERIFY_CONCURRENCY)
        for _ in range(2):
            pending = [i for i, p in enumerate(parcels) if _row_issues(p)][:_MAX_REVERIFY_ROWS]
            if not pending:
                break
            pass_n += 1
            retry_geocode: list[int] = []

            async def _one(i: int) -> None:
                nonlocal called
                p = parcels[i]
                raw = raw_rows[i] if i < len(raw_rows) else {}
                issues = _row_issues(p)
                async with sem:
                    cand, hit = await _llm_reverify_row({str(k): str(v) for k, v in raw.items()}, issues)
                called = called or hit
                _FIELD = {"address": "address", "jibun": "jibun", "jimok": "jimok", "area": "area_sqm"}
                changed_addr = False
                for role, val in cand.items():
                    field = _FIELD.get(role)
                    if not field:
                        continue
                    before = p.get(field)
                    new_val = _to_float(val) if role == "area" else val
                    if new_val and new_val != before:
                        p[field] = new_val
                        p.setdefault("_corrections", []).append({
                            "field": field, "before": before, "after": new_val,
                            "reason": "LLM 재질의(원문 부분문자열 검증 통과)",
                        })
                        if field == "jibun":
                            p["jibun"] = _clean_num_str(str(new_val))
                        if field in ("address", "jibun"):
                            changed_addr = True
                if changed_addr and p.get("status") in ("failed", "ambiguous") and not p.get("pnu"):
                    p["status"] = "need_geocode"
                    retry_geocode.append(i)

            await asyncio.gather(*[_one(i) for i in pending], return_exceptions=True)
            if retry_geocode:
                await self._geocode_fill(parcels, retry_geocode)
                await self._enrich_fill(parcels)
        return {"passes": pass_n, "llm_called": called}

    def _build_verification_report(
        self, parcels: list[dict[str, Any]], *, declared_total: float | None, excluded_n: int,
        dup_warning: str | None, llm_used: bool, passes: int, structure_notes: list[str],
    ) -> dict[str, Any]:
        """S4: 행별 최종분류(verified/corrected/needs_review) + 요약 리포트."""
        warnings: list[str] = list(structure_notes)
        if dup_warning:
            warnings.append(dup_warning)
        if declared_total:
            extracted_sum = sum(p["area_sqm"] for p in parcels if p.get("area_sqm"))
            if extracted_sum > 0:
                diff = abs(extracted_sum - declared_total) / declared_total
                if diff > 0.01:
                    warnings.append(
                        f"추출 면적합 {extracted_sum:,.1f}㎡가 표의 합계행 값 {declared_total:,.1f}㎡와 "
                        f"{diff:.1%} 차이 — 데이터 확인이 필요합니다."
                    )

        corrections: list[dict] = []
        for p in parcels:
            c = p.pop("_corrections", None)
            if c:
                corrections.extend(c)
            issues = _row_issues(p)
            has_correction = bool(c) or bool(p.get("area_warning")) or bool(p.get("co_owner"))
            if issues:
                vstatus = "needs_review"
            elif has_correction:
                vstatus = "corrected"
            else:
                vstatus = "verified"
            reasons = [_ISSUE_LABELS.get(i, i) for i in issues]
            if p.get("reason"):
                reasons.append(p["reason"])
            if p.get("area_warning"):
                reasons.append(p["area_warning"])
            p["verification_status"] = vstatus
            p["verification_reasons"] = reasons
            # 확정분(verified+corrected)만 기본 주입 대상 — needs_review는 사용자 확인 후 반영.
            p["injectable"] = vstatus in ("verified", "corrected")

        counts = Counter(p["verification_status"] for p in parcels)
        return {
            "counts": {
                "verified": counts.get("verified", 0),
                "corrected": counts.get("corrected", 0),
                "needs_review": counts.get("needs_review", 0),
                "excluded": excluded_n,
            },
            "corrections": corrections,
            "warnings": warnings,
            "llm_used": llm_used,
            "passes": passes,
        }

    async def _geocode_fill(self, parcels: list[dict[str, Any]], idxs: list[int]) -> None:
        try:
            from app.services.external_api.vworld_service import VWorldService
        except Exception:  # noqa: BLE001
            return
        vworld = VWorldService()
        sem = asyncio.Semaphore(_GEOCODE_CONCURRENCY)

        async def one(i: int) -> None:
            p = parcels[i]
            # C1: 지번을 결합해 모호성 감소(번지 없는 동명만 입력 시 동명이의 오매칭 방지).
            addr = p["address"] or ""
            jibun = (p.get("jibun") or "").strip()
            query = f"{addr} {jibun}".strip() if (jibun and jibun not in addr) else addr
            async with sem:
                try:
                    geo = await vworld.geocode_address(query)
                    if not geo and query != addr:  # 결합 실패 시 원주소 폴백
                        geo = await vworld.geocode_address(addr)
                except Exception:  # noqa: BLE001
                    geo = None
            if not geo:
                return
            p["lat"] = geo.get("lat")
            p["lon"] = geo.get("lon")
            gp = str(geo.get("pnu") or "")
            if len(gp) != 19:
                return
            # C2: 지역 검증 — 입력 bcode가 있으면 시군구(앞5자리)까지 대조. 불일치=동명이의 오매칭
            #     가능 → 자동확정 금지, ambiguous 정직표기(가짜 확신 금지).
            in_bcode = (p.get("bcode") or "")
            if in_bcode and len(in_bcode) >= 5 and not gp.startswith(in_bcode[:5]):
                p["status"] = "ambiguous"
                p["reason"] = (f"지오코딩 결과 지역({gp[:5]})이 입력 법정동({in_bcode[:5]})과 불일치 — "
                               "동명이의 가능. 시·군·구 포함 정확한 주소로 보완하세요.")
                return
            # C2 보강(★사용자요청 — 법정동+번지 자동해석): bcode·시군구 없는 '순수 동명'이라도 번지가
            #   있으면 검색 API(search_address) 후보가 '단일 법정동(동일 bcode 10자리)'으로 수렴하는지
            #   확인한다. 단일 수렴=동명이의 아님 → 그 필지로 자동확정 + 전체 시군구 주소 자동보완.
            #   복수 법정동(진짜 동명이의)이면 ambiguous 정직표기(가짜 확신 금지).
            _has_sigungu = re.search(
                r"(특별시|광역시|특별자치시|특별자치도|[가-힣]+도)\s|[가-힣]+시\s|[가-힣]+군\s|[가-힣]+구\s",
                (addr + " "),
            )
            if not in_bcode and not _has_sigungu:
                resolved = False
                if re.search(r"\d", f"{addr} {jibun}"):  # 번지(숫자) 있어야 의미
                    try:
                        async with sem:
                            cands = await vworld.search_address(query, size=8)
                    except Exception:  # noqa: BLE001
                        cands = []
                    cand_pnus = [str(c.get("pnu") or "") for c in cands if len(str(c.get("pnu") or "")) == 19]
                    bcodes = {cp[:10] for cp in cand_pnus}
                    if len(bcodes) == 1 and cand_pnus:
                        # 단일 법정동 수렴 → 최적 후보(첫 후보=검색 best match)로 확정.
                        chosen = next((c for c in cands if str(c.get("pnu") or "") in cand_pnus), cands[0])
                        gp = str(chosen.get("pnu") or "") or gp
                        p["address"] = chosen.get("address") or addr  # 전체 시군구 주소 자동보완
                        if chosen.get("lat"):
                            p["lat"] = chosen.get("lat")
                        if chosen.get("lon"):
                            p["lon"] = chosen.get("lon")
                        resolved = True
                if not resolved:
                    p["status"] = "ambiguous"
                    p["reason"] = ("시·군·구 없는 동명 — 검색 결과가 여러 법정동(동명이의)으로 갈려 자동확정하지 "
                                   "않음. 시·군·구를 포함하거나 정확한 지번을 선택해 주세요.")
                    return
            # C2-2: 번지(본번) 숫자가 전혀 없는 '동·읍·면 단위' 주소는 지오코딩이 동 대표지점
            #   1필지로 수렴 → 여러 행이 같은 필지(동일 면적·지목)로 잘못 매핑된다(예: '서울 동작구
            #   상도동' 53㎡·지목 전이 반복). 자동확정 금지(번지 필요). 도로명·지번은 숫자를 포함.
            if not re.search(r"\d", f"{addr} {jibun}"):
                p["status"] = "ambiguous"
                p["reason"] = ("번지(본번) 없이 동·읍·면 단위만 입력 — 동 대표지점 1필지로 수렴해 여러 "
                               "필지가 동일하게 매핑됩니다. 번지를 포함해 주세요(예: 상도동 210-453).")
                return
            p["pnu"] = gp
            p["bcode"] = p["bcode"] or gp[:10]
            p["status"] = "ok"

        await asyncio.gather(*[one(i) for i in idxs], return_exceptions=True)

    async def _enrich_fill(self, parcels: list[dict[str, Any]]) -> None:
        """PNU 확보된 필지의 빈 칸(면적·지목·용도지역·공시지가)을 NED 토지특성으로 보강.

        사용자가 주소만 입력해도 PNU만 있으면 면적/용도/공시지가를 자동 조회한다(가짜값 금지).
        이미 입력된 값(엑셀에 적어둔 면적/지목)은 보존하고 빈 칸만 채운다.
        """
        targets = [
            i for i, p in enumerate(parcels)
            if p.get("pnu") and not (p.get("area_sqm") and p.get("zone_type") and p.get("official_price_per_sqm"))
        ]
        if not targets:
            return
        try:
            from app.services.external_api.vworld_service import VWorldService
        except Exception:  # noqa: BLE001
            return
        vworld = VWorldService()
        sem = asyncio.Semaphore(_GEOCODE_CONCURRENCY)

        async def one(i: int) -> None:
            p = parcels[i]
            async with sem:
                try:
                    lc = await vworld.get_land_characteristics(p["pnu"])
                except Exception:  # noqa: BLE001
                    lc = None
            if not isinstance(lc, dict):
                return
            # ★면적 교차검증(신뢰루프): 입력 면적(엑셀)이 있으면 공부상(VWorld 토지특성)과 대조.
            #   빈칸이면 공부상으로 보강. 큰 괴리(>50%)면 입력 오기로 보고 공부상 우선 채택 +
            #   정직 경고(단일 오기가 합계를 부풀리는 것 방지 — 예 211-204 엑셀 8,500 vs 공부상 236).
            lc_area = lc.get("area_sqm")
            if lc_area:
                lc_area = round(float(lc_area), 1)
                excel_area = p.get("area_sqm")
                if not excel_area:
                    p["area_sqm"] = lc_area
                elif lc_area > 0:
                    ratio = excel_area / lc_area
                    if ratio > 1.5 or ratio < 0.67:
                        p["area_input_sqm"] = excel_area  # 입력값 보존(참고용)
                        p["area_sqm"] = lc_area            # 공부상(권원) 면적 채택
                        p["area_warning"] = (
                            f"입력 면적 {excel_area:g}㎡가 공부상 면적 {lc_area:g}㎡와 "
                            f"{ratio:.1f}배 차이 — 공부상 면적으로 보정했습니다(입력값 점검 필요)."
                        )
            if not p.get("jimok") and lc.get("land_category"):
                p["jimok"] = lc["land_category"]
            if lc.get("zone_type"):
                p["zone_type"] = lc["zone_type"]
            if lc.get("official_price_per_sqm"):
                p["official_price_per_sqm"] = int(lc["official_price_per_sqm"])

        await asyncio.gather(*[one(i) for i in targets], return_exceptions=True)

    async def _heal_stale_pnu(self, parcels: list[dict[str, Any]]) -> bool:
        """★bad-data 자가치유: PNU는 확보됐지만 용도지역을 못 불러온 행을 주소로 재해소한다.

        엑셀에 적힌 법정동코드(bcode)가 실제 주소와 어긋나면 그 코드로 만든 PNU가 '존재하지 않는
        필지'를 가리켜 NED 토지특성 조회가 빈값이 된다(용도지역·건폐율·용적률이 영구히 안 불러와짐).
        이런 행은 need_geocode(=PNU 없는 행)에 안 들어가 재지오코딩도 안 돼 그대로 멈춰 있었다.

        해법(공공데이터=권위 원천 우선): zone_type을 못 채운 PNU 행 중 주소/지번이 있는 행은
        잘못된 PNU를 폐기하고 주소로 재지오코딩 → 올바른 PNU 재확보 → 한 번 더 보강한다.
        엑셀의 비권위 코드가 zone 로드를 깨뜨리던 문제를 근본 해소. 무한루프 방지로 호출측에서 1회만.

        Returns: 재해소를 시도한 행이 있었으면 True(호출측이 후처리 재집계 여부 판단).
        """
        stale = [
            i for i, p in enumerate(parcels)
            if p.get("pnu")
            and not p.get("zone_type")
            and (p.get("address") or p.get("jibun"))
            and not p.get("_stale_pnu_retried")  # 1회 재시도 가드
        ]
        if not stale:
            return False
        for i in stale:
            p = parcels[i]
            p["_stale_pnu_retried"] = True
            p["pnu"] = None  # 잘못된 PNU 폐기 → _geocode_fill 이 주소로 재해소
            # 어긋난 bcode 도 비운다(재지오코딩 결과와의 불일치 경고 오발동 방지·주소 권위 우선).
            p["bcode"] = ""
        await self._geocode_fill(parcels, stale)
        await self._enrich_fill(parcels)
        # 재해소 실패(주소로도 PNU 못 잡음) 행은 정직하게 failed 로 — bad PNU 로 set 됐던 'ok' 가
        # pnu=None 인 채 남아 'status=ok 인데 데이터 없음'(silent bad data)이 되는 것을 차단.
        for i in stale:
            p = parcels[i]
            if not p.get("pnu") and p.get("status") not in ("ambiguous", "failed"):
                p["status"] = "failed"
                p["reason"] = (p.get("reason") or
                               "입력된 법정동코드가 주소와 어긋나 용도지역을 확인하지 못했습니다 — "
                               "시·군·구+번지 포함 정확한 주소로 보완하세요(잘못된 코드는 무시).")
        return True

    async def enrich_parcel_list(
        self, items: list[dict[str, Any]], with_building: bool = True,
    ) -> list[dict[str, Any]]:
        """주소/지번/PNU 목록 → 필지별 토지정보(면적·지목·용도지역·공시지가) 일괄 보강.

        개별등록·엑셀 공통 경로. PNU 없으면 지번결합 지오코딩(_geocode_fill) 후 NED 토지특성
        보강(_enrich_fill). with_building=True면 건축물대장 표제부로 집합건물(공동주택·빌라)
        여부·건물명·용도·세대수를 best-effort 부착(호실/대지지분 안내용, 1콜/필지·동시성 제한).
        무목업: 실패는 status="failed"+reason 정직표기(가짜값 금지).
        """
        parcels: list[dict[str, Any]] = []
        for it in items:
            pnu = str(it.get("pnu") or "").strip()
            # ★엑셀 입력 면적(비권위)을 area_sqm 초기값으로 실어 _enrich_fill 교차검증을 활성화한다.
            #   PNU가 있으면 _enrich_fill이 공부상과 대조→큰 괴리는 공부상 채택+area_warning 생성,
            #   PNU가 없으면(_enrich_fill 미적용) 이 입력 면적이 그대로 area_sqm로 남는다(참고 보존).
            try:
                area_in = float(it.get("area_input_sqm")) if it.get("area_input_sqm") is not None else None
            except (TypeError, ValueError):
                area_in = None
            parcels.append({
                "rid": it.get("__rid"),  # 호출측 행 식별자 — 응답에서 그대로 echo(주소 충돌 매칭 회피)
                "address": str(it.get("address") or "").strip(),
                "jibun": str(it.get("jibun") or "").strip(),
                "bcode": str(it.get("bcode") or "").strip(),
                "pnu": pnu if len(pnu) == 19 else None,
                "area_sqm": area_in if (area_in and area_in > 0) else None,
                "zone_type": None, "jimok": None,
                "official_price_per_sqm": None,
                "status": "ok" if len(pnu) == 19 else "pending",
            })

        need_geocode = [i for i, p in enumerate(parcels) if not p["pnu"]]
        if need_geocode:
            await self._geocode_fill(parcels, need_geocode)
        await self._enrich_fill(parcels)

        # ★bad-data 자가치유: 잘못된 bcode 로 만든 PNU 라 용도지역을 못 불러온 행을 주소로 재해소.
        await self._heal_stale_pnu(parcels)

        # PNU 미확보(지오코딩 실패) → 정직 실패표기. ambiguous는 _geocode_fill이 이미 표기.
        for p in parcels:
            if p["status"] == "pending":
                if p["pnu"]:
                    p["status"] = "ok"
                else:
                    p["status"] = "failed"
                    p.setdefault("reason", "PNU 미확보 — 시·군·구 포함 정확한 주소로 보완하세요.")

        if with_building:
            await self._building_fill(parcels)
        return parcels

    async def _building_fill(self, parcels: list[dict[str, Any]]) -> None:
        """PNU 확보 필지에 건축물대장 표제부로 집합건물(공동주택·빌라) 여부·건물명·세대수 부착.

        is_aggregate는 용도(공동주택/다세대/연립/아파트) 또는 세대/호수>1로 best-effort 판정.
        정확한 호별 대지지분은 화면에서 /zoning/land-share(전유부 전수)로 확정(여기선 가벼운 플래그만).
        """
        targets = [i for i, p in enumerate(parcels) if p.get("pnu")]
        if not targets:
            return
        try:
            from app.services.external_api.building_registry_service import BuildingRegistryService
        except Exception:  # noqa: BLE001
            return
        breg = BuildingRegistryService()
        sem = asyncio.Semaphore(_GEOCODE_CONCURRENCY)
        _AGG = ("공동주택", "다세대", "연립", "아파트", "도시형생활주택", "오피스텔")

        async def one(i: int) -> None:
            p = parcels[i]
            async with sem:
                try:
                    t = await breg.get_title_by_pnu(p["pnu"])
                except Exception:  # noqa: BLE001
                    t = None
            if not isinstance(t, dict):
                return
            purpose = str(t.get("main_purpose", "") or "")
            units = max(int(t.get("household_count") or 0), int(t.get("ho_count") or 0))
            is_agg = any(k in purpose for k in _AGG) or units > 1
            p["building"] = {
                "is_aggregate": bool(is_agg),
                "building_name": t.get("building_name", ""),
                "main_purpose": purpose,
                "unit_count": units or None,
                # ★건축물현황(동수·건축연한·노후도) 산출용 — 표제부 실값 그대로 전달(가짜 금지).
                "use_approval_date": t.get("use_approval_date", "") or "",  # 사용승인일(YYYYMMDD) → 건축연한
                "ground_floors": t.get("ground_floors"),       # 0층 보존(or None 금지 — 0이 소실됨)
                "underground_floors": t.get("underground_floors"),
                "total_area_sqm": t.get("total_area_sqm") or None,  # 연면적(0=무자료)
                "structure": t.get("structure", "") or "",
                "dong_count": t.get("dong_count") or None,        # 표제부 동수
                "is_demolished": bool(t.get("is_demolished")),     # 멸실(추정)
            }

        await asyncio.gather(*[one(i) for i in targets], return_exceptions=True)
