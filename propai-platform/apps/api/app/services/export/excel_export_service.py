"""v61 Excel 내보내기 서비스 — 수지분석 + 원가계산서 XLSX 생성.

openpyxl 미설치 시 CSV 폴백으로 동작한다.
"""

from __future__ import annotations

import csv
import io
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError:
    openpyxl = None  # type: ignore[assignment]


class ExcelExportService:
    """XLSX/CSV 파일 바이트를 생성한다."""

    def feasibility_to_xlsx(self, result: dict[str, Any]) -> tuple[bytes, str]:
        """수지분석 결과를 Excel 바이트로 변환한다.

        Returns:
            (file_bytes, content_type)
        """
        if openpyxl is not None:
            return self._feasibility_xlsx(result), \
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return self._feasibility_csv(result), "text/csv"

    def cost_sheet_to_xlsx(self, rows: list[list[Any]]) -> tuple[bytes, str]:
        """원가계산서 행렬을 Excel 바이트로 변환한다.

        Args:
            rows: OriginCostCalculator.to_excel_data() 반환값.

        Returns:
            (file_bytes, content_type)
        """
        if openpyxl is not None:
            return self._cost_xlsx(rows), \
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return self._cost_csv(rows), "text/csv"

    # ── openpyxl 구현 ──

    def _feasibility_xlsx(self, result: dict[str, Any]) -> bytes:
        # ★수식 임베드: 값만 넣지 않고 "파생 셀"(순이익·수익률·비용비율·합계)에 엑셀 수식(=)을 써서,
        #   다운로드 후 사용자가 총수입/총비용/비용항목을 바꾸면 엑셀이 스스로 재계산하게 한다.
        #   수식 분모는 엔진(aggregation_engine)과 동일: 순이익=총수입−총비용, 수익률=순이익÷총수입×100.
        #   ROI·NPV·등급은 엔진 고유 산식(분모 가변·DCF·임계값)이라 값으로 두어 오답을 막는다.
        wb = openpyxl.Workbook()
        won_fmt = '#,##0" 원"'   # 통화 표시 형식
        pct_fmt = '0.00"%"'      # 백분율 표시 형식

        # 요약 시트
        ws = wb.active
        ws.title = "수지분석 요약"
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        def _num(v: Any) -> float:
            """수치만 안전 추출(비수치·bool·None → 0.0). 셀에 문자열 대신 진짜 숫자를 넣기 위함."""
            if isinstance(v, bool) or v is None:
                return 0.0
            try:
                return float(v)
            except (TypeError, ValueError):
                return 0.0

        ws.merge_cells("A1:C1")
        ws["A1"] = "PropAI v61 — 수지분석 보고서"
        ws["A1"].font = Font(bold=True, size=14)

        revenue = _num(result.get("total_revenue_won", 0))
        cost = _num(result.get("total_cost_won", 0))

        # 헤더(3행)
        for col_idx, val in enumerate(["항 목", "값", "비 고"], start=1):
            c = ws.cell(row=3, column=col_idx, value=val)
            c.border = thin_border
            c.font = header_font_white
            c.fill = header_fill

        # 4행부터 데이터. B열은 값 또는 수식, 총수입=B6·총비용=B7·순이익=B8 참조.
        # (label, B값또는수식, 숫자포맷, 비고)
        data_rows = [
            ("개발유형", result.get("development_type", "-"), None, ""),
            ("모듈", result.get("module_name", "-"), None, ""),
            ("총수입", revenue, won_fmt, "엔진 산출값(수정 가능)"),
            ("총비용", cost, won_fmt, "엔진 산출값(수정 가능)"),
            ("순이익", "=B6-B7", won_fmt, "총수입 − 총비용 (자동계산)"),
            ("수익률", "=IF(B6=0,0,B8/B6*100)", pct_fmt, "순이익 ÷ 총수입 (자동계산)"),
            ("ROI", _num(result.get("roi_pct", 0)), pct_fmt, "순이익 ÷ 총사업비 (엔진)"),
            ("NPV", _num(result.get("npv_won", 0)), won_fmt, "무차입 DCF (엔진)"),
            ("등급", result.get("grade", "-"), None, ""),
        ]
        for i, (label, bval, fmt, memo) in enumerate(data_rows, start=4):
            a = ws.cell(row=i, column=1, value=label)
            b = ws.cell(row=i, column=2, value=bval)
            c = ws.cell(row=i, column=3, value=memo)
            for cell in (a, b, c):
                cell.border = thin_border
            if fmt:
                b.number_format = fmt

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 24
        ws.column_dimensions["C"].width = 26

        # 비용 구성 시트 — 합계=SUM(), 비율=항목÷합계 수식(자동재계산)
        cost_breakdown = result.get("cost_breakdown_won", {})
        if cost_breakdown:
            ws2 = wb.create_sheet("비용 구성")
            ws2.append(["비용 항목", "금액 (원)", "비율"])
            first = 2  # 첫 데이터 행
            # 무음 행손실 방지: 모든 항목 보존(수치는 숫자·수식비율, 비수치는 문자·빈비율).
            items = list(cost_breakdown.items())
            for key, val in items:
                is_num = isinstance(val, (int, float)) and not isinstance(val, bool)
                ws2.append([key, _num(val) if is_num else str(val), None])
            n = len(items)
            last = first + n - 1
            total_row = first + n  # 합계 행
            for idx, (key, val) in enumerate(items):
                r = first + idx
                is_num = isinstance(val, (int, float)) and not isinstance(val, bool)
                if is_num:
                    ws2.cell(row=r, column=2).number_format = won_fmt
                    ratio = ws2.cell(row=r, column=3,
                                     value=f"=IF($B${total_row}=0,0,B{r}/$B${total_row})")
                    ratio.number_format = "0.0%"
            # 합계 행(SUM은 문자 셀을 자동 무시)
            ws2.cell(row=total_row, column=1, value="합계").font = Font(bold=True)
            sc = ws2.cell(row=total_row, column=2, value=f"=SUM(B{first}:B{last})")
            sc.number_format = won_fmt
            sc.font = Font(bold=True)
            ws2.column_dimensions["A"].width = 25
            ws2.column_dimensions["B"].width = 22

        # 세금 상세 시트 — 숫자 + 합계 SUM(자동)
        tax_detail = result.get("tax_detail", {})
        if tax_detail:
            ws3 = wb.create_sheet("세금 상세")
            ws3.append(["세금 항목", "금액 (원)"])
            tax_items = list(tax_detail.items())
            for key, val in tax_items:
                is_num = isinstance(val, (int, float)) and not isinstance(val, bool)
                ws3.append([key, _num(val) if is_num else str(val)])
            for idx, (key, val) in enumerate(tax_items):
                if isinstance(val, (int, float)) and not isinstance(val, bool):
                    ws3.cell(row=2 + idx, column=2).number_format = won_fmt
            if tax_items:
                tr = 2 + len(tax_items)
                ws3.cell(row=tr, column=1, value="합계").font = Font(bold=True)
                tsc = ws3.cell(row=tr, column=2, value=f"=SUM(B2:B{tr - 1})")
                tsc.number_format = won_fmt
                tsc.font = Font(bold=True)
            ws3.column_dimensions["A"].width = 25
            ws3.column_dimensions["B"].width = 22

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _cost_xlsx(self, rows: list[list[Any]]) -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "원가계산서"
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, size=11, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        ws.merge_cells("A1:C1")
        ws["A1"] = "PropAI v61 — 원가계산서 (2026 법정요율 적용)"
        ws["A1"].font = Font(bold=True, size=14)

        for row_idx, row_data in enumerate(rows, start=3):
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                if row_idx == 3:
                    cell.font = header_font
                    cell.fill = header_fill
                if col_idx == 2:
                    cell.alignment = Alignment(horizontal="right")

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 30

        # 합계 행 강조
        total_row = len(rows) + 2
        for col in range(1, 4):
            cell = ws.cell(row=total_row, column=col)
            cell.font = Font(bold=True)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── CSV 폴백 ──

    def _feasibility_csv(self, result: dict[str, Any]) -> bytes:
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(["항목", "값"])
        writer.writerow(["개발유형", result.get("development_type", "")])
        writer.writerow(["총수입", result.get("total_revenue_won", 0)])
        writer.writerow(["총비용", result.get("total_cost_won", 0)])
        writer.writerow(["순이익", result.get("net_profit_won", 0)])
        writer.writerow(["수익률(%)", result.get("profit_rate_pct", 0)])
        writer.writerow(["ROI(%)", result.get("roi_pct", 0)])
        writer.writerow(["NPV", result.get("npv_won", 0)])
        writer.writerow(["등급", result.get("grade", "")])

        cost_bd = result.get("cost_breakdown_won", {})
        if cost_bd:
            writer.writerow([])
            writer.writerow(["비용 항목", "금액"])
            for k, v in cost_bd.items():
                writer.writerow([k, v])

        return buf.getvalue().encode("utf-8-sig")

    def _cost_csv(self, rows: list[list[Any]]) -> bytes:
        buf = io.StringIO()
        writer = csv.writer(buf)
        for row in rows:
            writer.writerow(row)
        return buf.getvalue().encode("utf-8-sig")
