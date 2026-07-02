"""apick(에이픽) 부동산등기부등본 어댑터 — 스크래핑 중개사(주소 직접·MD5 키).

흐름: ① /rest/iros/1(열람) 주소→ic_id  ② /rest/iros_download/1(다운로드) ic_id→PDF/Excel.
Excel(xlsx)을 openpyxl로 텍스트 추출해 기존 등기 분석 파이프라인(registry_text)에 투입하고,
PDF는 base64로 받아 보관용 URL 생성에 사용한다.

장점: 고객 인터넷등기소 ID/PW 불필요(apick 자체 계정 스크래핑) → 계정 노출 0, 월 고정비 0.
한계: 구조화 JSON 미제공 → Excel/PDF 텍스트 추출 후 LLM 분석(CODEF는 구조화 JSON 직접).
"""

from __future__ import annotations

import asyncio
import base64
import io
import os
from typing import Any

import structlog

logger = structlog.get_logger()

_HOST = (os.getenv("APICK_HOST") or "https://apick.app").rstrip("/")

# CODEF realty_type 의미 → apick type 매핑(0토지+건물,1집합건물,2토지,3건물)
_TYPE_MAP = {"0": "집합건물", "1": "집합건물", "2": "토지", "3": "건물"}


def apick_key() -> str:
    return (os.getenv("APICK_CL_AUTH_KEY") or os.getenv("APICK_API_KEY") or "").strip()


def apick_ready() -> bool:
    return bool(apick_key())


def _xlsx_to_text(content: bytes) -> str:
    """apick Excel(xlsx) 등기부를 줄 단위 텍스트로 평탄화(LLM 분석 입력용)."""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        lines: list[str] = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
                if cells:
                    lines.append(" ".join(cells))
        return "\n".join(lines)
    except Exception as e:  # noqa: BLE001
        logger.warning("apick Excel 파싱 실패", err=str(e)[:120])
        return ""


async def fetch_registry(
    *, address: str | None = None, unique_num: str | None = None,
    realty_type: str | None = None,
) -> dict[str, Any]:
    """apick 등기부 열람+다운로드 → 표준 결과(registry_text·pdf_base64 포함)."""
    item: dict[str, Any] = {"address": address}
    key = apick_key()
    if not key:
        return {**item, "status": "not_configured", "message": "apick 인증키(APICK_CL_AUTH_KEY) 미설정"}
    if not (address or unique_num):
        return {**item, "status": "error", "message": "주소 또는 부동산 고유번호가 필요합니다."}

    import httpx

    headers = {"CL_AUTH_KEY": key}
    rtype = _TYPE_MAP.get((realty_type or "").strip(), "집합건물")
    form: dict[str, str] = {"type": rtype}
    if unique_num:
        form["unique_num"] = unique_num
    elif address:
        form["address"] = address

    try:
        async with httpx.AsyncClient(timeout=60.0) as client:
            # ① 열람 요청 → ic_id (성공:{"data":{"ic_id":N,"success":1}} / 에러:data.error 또는 result.error)
            #   ★열람 단계는 과금 전(ic_id 발급 전)이라, 일시 타임아웃/연결오류는 1회 재시도해 신뢰성↑.
            r1 = None
            for attempt in range(2):
                try:
                    r1 = await client.post(f"{_HOST}/rest/iros/1", headers=headers, data=form)
                    break
                except (httpx.TimeoutException, httpx.ConnectError, httpx.ReadError) as te:
                    if attempt == 0:
                        await asyncio.sleep(1.5)
                        continue
                    # 2회 모두 타임아웃/연결오류 → 정직 분류(빈 메시지 예외도 타입명으로 진단 가능).
                    logger.warning("apick 열람 연결/타임아웃 실패", err=f"{type(te).__name__}: {str(te)[:120]}")
                    return {**item, "status": "provider_timeout",
                            "message": "apick 등기 발급 서버 응답 지연·연결 실패(대법원 인터넷등기소/발급 연동 지연). "
                                       "잠시 후 ‘자동채움’을 다시 시도하거나 등기부 내용을 직접 입력하세요."}
            if r1 is None:  # 도달 불가(위에서 return) — 타입 안전 가드.
                return {**item, "status": "provider_timeout",
                        "message": "apick 등기 발급 서버 응답을 받지 못했습니다. 잠시 후 다시 시도하세요."}
            try:
                j1 = r1.json()
            except Exception:  # noqa: BLE001
                return {**item, "status": "provider_error",
                        "message": f"apick 응답 파싱 실패(HTTP {r1.status_code})"}
            data = j1.get("data") or {}
            ic_id = data.get("ic_id")
            if not ic_id or int(data.get("success") or 0) != 1:
                # 401 등에도 raise하지 않고 apick 실제 사유 노출(예: "아이디 로그인")
                apick_err = data.get("error") or (j1.get("result") or {}).get("error")
                # 대법원 인터넷등기소 점검/일시중단은 제공사(에이픽/텔코) 공통 영향 → 별도 분류
                err_txt = str(apick_err or "")
                is_maint = any(k in err_txt for k in ("점검", "유지보수", "일시", "중단", "maintenance", "unavailable"))
                if is_maint:
                    return {**item, "status": "provider_maintenance",
                            "message": f"대법원 인터넷등기소 점검·일시중단으로 등기 발급이 일시 불가합니다 — {err_txt} "
                                       "(에이픽/텔코 등 발급 연동은 인터넷등기소에 의존). 점검 종료 후 재시도하세요.",
                            "raw": j1}
                return {**item, "status": "provider_error",
                        "message": apick_err
                        or "apick 등기 열람 실패(주소/고유번호 확인 또는 계정 결제 필요)",
                        "raw": j1}

            # ② 다운로드 — 발급 PDF/Excel 생성에 20~30초 소요 → 처리중(헤더 result=2)이면 폴링
            async def _download(fmt: str) -> bytes | None:
                for _ in range(12):  # 최대 ~36초
                    rr = await client.post(f"{_HOST}/rest/iros_download/1", headers=headers,
                                           data={"ic_id": str(ic_id), "format": fmt})
                    res = rr.headers.get("result")
                    ctype = (rr.headers.get("content-type") or "").lower()
                    if rr.status_code == 200 and rr.content and "json" not in ctype and res != "2":
                        return rr.content  # 바이너리(PDF/xlsx)
                    if res == "2" or rr.status_code == 202:  # 처리중 → 대기 후 재시도
                        await asyncio.sleep(3)
                        continue
                    return None
                return None

            xlsx = await _download("excel")
            pdf = await _download("pdf")

        registry_text = _xlsx_to_text(xlsx) if xlsx else ""
        pdf_b64 = base64.b64encode(pdf).decode() if pdf else None
        if not registry_text and not pdf_b64:
            return {**item, "status": "provider_error",
                    "message": "apick 다운로드 결과가 비어 있습니다.", "ic_id": ic_id}

        return {
            **item, "status": "ok", "origin": "apick", "ic_id": ic_id,
            "registry_text": registry_text or None,   # LLM 분석 입력(없으면 PDF만)
            "pdf_base64": pdf_b64,
            "has_pdf": bool(pdf_b64),
            "owner": None,  # 구조화 미제공 → LLM이 registry_text에서 추출
        }
    except (httpx.TimeoutException, httpx.ConnectError, httpx.ReadError) as e:
        # ★빈 메시지 예외(httpx 타임아웃류는 str(e)가 비기 쉬움) → 타입명으로 진단 가능하게 기록.
        logger.warning("apick 등기 발급 타임아웃/연결 실패", err=f"{type(e).__name__}: {str(e)[:120]}")
        return {**item, "status": "provider_timeout",
                "message": "apick 등기 발급 서버 응답 지연(타임아웃) — 발급 PDF 생성 지연 또는 대법원 인터넷등기소 "
                           "지연일 수 있습니다. 잠시 후 다시 시도하거나 등기부 내용을 직접 입력하세요."}
    except Exception as e:  # noqa: BLE001
        logger.warning("apick 등기 조회 실패", err=f"{type(e).__name__}: {str(e)[:150]}")
        return {**item, "status": "error", "message": f"{type(e).__name__}: {str(e)[:180]}"}
