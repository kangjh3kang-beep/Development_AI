"""수지분석 엑셀 '수식 임베드' 회귀가드.

다운로드한 엑셀이 값 스냅샷이 아니라 살아있는 수식(순이익·수익률·비용비율·합계)을 담아,
사용자가 총수입/총비용/비용항목을 바꾸면 엑셀이 자동 재계산하도록 보장한다.
수식 분모는 엔진(aggregation_engine)과 동일해야 한다: 순이익=총수입−총비용, 수익률=순이익/총수입×100.
"""
from __future__ import annotations

import io

import openpyxl

from app.services.export.excel_export_service import ExcelExportService

_RESULT = {
    "development_type": "M08_officetel",
    "module_name": "오피스텔",
    "total_revenue_won": 50_000_000_000,
    "total_cost_won": 40_000_000_000,
    "net_profit_won": 10_000_000_000,
    "profit_rate_pct": 20.0,
    "roi_pct": 25.0,
    "npv_won": 8_000_000_000,
    "grade": "B",
    "cost_breakdown_won": {"토지비": 15_000_000_000, "공사비": 20_000_000_000, "기타": 5_000_000_000},
    "tax_detail": {"취득세": 2_000_000_000, "개발부담금": 1_000_000_000},
}


def _load():
    data, ct = ExcelExportService().feasibility_to_xlsx(_RESULT)
    assert ct.endswith("spreadsheetml.sheet")
    return openpyxl.load_workbook(io.BytesIO(data))


def test_summary_derived_cells_are_live_formulas():
    ws = _load()["수지분석 요약"]
    # 총수입/총비용은 진짜 숫자(문자열 "원" 아님) — 사용자가 바꾸면 하류 수식이 재계산됨
    assert isinstance(ws["B6"].value, (int, float))
    assert isinstance(ws["B7"].value, (int, float))
    # 순이익·수익률은 수식(=)
    assert ws["B8"].value == "=B6-B7"
    assert ws["B9"].value == "=IF(B6=0,0,B8/B6*100)"


def test_formula_denominators_match_engine():
    """수식이 산출할 값이 엔진값과 일치하는지(분모 정합) 직접 계산 검증."""
    ws = _load()["수지분석 요약"]
    revenue, cost = ws["B6"].value, ws["B7"].value
    # 순이익 = 총수입 − 총비용 (=B6-B7)
    assert revenue - cost == _RESULT["net_profit_won"]
    # 수익률 = 순이익 / 총수입 × 100 (=B8/B6*100)
    assert round((revenue - cost) / revenue * 100, 2) == _RESULT["profit_rate_pct"]


def test_cost_breakdown_sum_and_ratio_formulas():
    ws2 = _load()["비용 구성"]
    # 3개 항목(2~4행) + 합계(5행)
    assert ws2["B5"].value == "=SUM(B2:B4)"
    assert ws2["C2"].value == "=IF($B$5=0,0,B2/$B$5)"
    # 항목 금액은 숫자
    assert isinstance(ws2["B2"].value, (int, float))


def test_tax_sum_formula():
    ws3 = _load()["세금 상세"]
    assert ws3["B4"].value == "=SUM(B2:B3)"


def test_no_silent_row_loss_on_nonnumeric_cost_item():
    """비수치 비용 항목도 보존(무음 행손실 금지) — 합계 SUM은 문자 셀 자동 무시."""
    result = dict(_RESULT)
    result["cost_breakdown_won"] = {"토지비": 15_000_000_000, "비고": "협의 중", "공사비": 20_000_000_000}
    data, _ = ExcelExportService().feasibility_to_xlsx(result)
    ws2 = openpyxl.load_workbook(io.BytesIO(data))["비용 구성"]
    labels = [ws2.cell(row=r, column=1).value for r in range(2, 5)]
    assert "비고" in labels  # 비수치 항목 보존됨
    assert ws2["B5"].value == "=SUM(B2:B4)"  # 합계는 전체 범위(문자 무시)
