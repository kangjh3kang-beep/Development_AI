"""B2 파라메트릭 적산 엔진(boq_parametric_engine) + 엑셀 내보내기 테스트.

검증 축:
- gfa 2배 → gfa 드라이버 항목 수량 2배 / fixed(식·qty<=2) 항목 불변
- households 미제공 폴백 warning · 제공 시 세대 항목 households 비례
- 조경면적 미제공 폴백 warning · 제공 시 landscape_area 드라이버
- 유효숫자 반올림(>=100 정수 / >=1 1자리 / <1 3자리)
- 응답 계약(disciplines/summary/provenance/badges) · 정직 배지(n=1)
- 엑셀: bytes 시그니처(PK) · 시트 수 = 공종 수 · 헤더/공란 단가
"""

import os
import sys
from io import BytesIO

import pytest

# apps/api 루트를 Python path에 추가 (conftest 와 동일 규약 — 단독 실행 대비)
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from app.services.cost.boq_excel_export import build_xlsx  # noqa: E402
from app.services.cost.boq_parametric_engine import (  # noqa: E402
    REF_GFA_SQM,
    REF_HOUSEHOLDS,
    REF_LANDSCAPE_AREA_SQM,
    generate_draft,
    round_qty,
)

ALL_DISCIPLINES = ["건축", "기계소방", "전기통신소방", "조경", "토목"]


def _items(draft, discipline):
    return draft["disciplines"][discipline]["items"]


def _by_driver(draft, discipline, driver):
    return [
        it for it in _items(draft, discipline)
        if it["qty_basis"]["driver"] == driver
    ]


# ──────────────────────────────────────────────
# 유효숫자 반올림
# ──────────────────────────────────────────────


class TestRoundQty:
    def test_100이상_정수(self):
        assert round_qty(123.456) == 123.0
        assert round_qty(238504.0) == 238504.0

    def test_1이상_1자리(self):
        assert round_qty(12.34) == 12.3
        assert round_qty(99.94) == 99.9

    def test_1미만_3자리(self):
        assert round_qty(0.12345) == 0.123
        assert round_qty(0.0006) == 0.001
        assert round_qty(0.0) == 0.0


# ──────────────────────────────────────────────
# gfa 드라이버 — 연면적 2배 → 수량 2배
# ──────────────────────────────────────────────


class TestGfaScaling:
    def test_gfa_2배_수량_2배(self):
        base = generate_draft({"gfa_sqm": REF_GFA_SQM}, disciplines=["토목"])
        double = generate_draft({"gfa_sqm": REF_GFA_SQM * 2}, disciplines=["토목"])
        base_gfa = {it["id"]: it for it in _by_driver(base, "토목", "gfa")}
        assert base_gfa, "토목 gfa 드라이버 항목이 존재해야 함"
        checked = 0
        for it in _by_driver(double, "토목", "gfa"):
            sample = it["qty_basis"]["sample_qty"]
            assert it["qty"] == round_qty(sample * 2.0)
            checked += 1
        assert checked == len(base_gfa)

    def test_ref와_동일_gfa면_표본수량_유지(self):
        draft = generate_draft({"gfa_sqm": REF_GFA_SQM}, disciplines=["건축"])
        for it in _by_driver(draft, "건축", "gfa"):
            assert it["qty"] == round_qty(it["qty_basis"]["sample_qty"])

    def test_qty_basis_계약(self):
        draft = generate_draft({"gfa_sqm": REF_GFA_SQM * 2}, disciplines=["토목"])
        it = _by_driver(draft, "토목", "gfa")[0]
        basis = it["qty_basis"]
        assert basis["ref_value"] == REF_GFA_SQM
        assert basis["project_value"] == REF_GFA_SQM * 2
        assert basis["sample_qty"] >= 0


# ──────────────────────────────────────────────
# fixed 드라이버 — '식' & qty_sample<=2 수량 유지
# ──────────────────────────────────────────────


class TestFixedDriver:
    def test_gfa_2배에도_fixed_불변(self):
        double = generate_draft({"gfa_sqm": REF_GFA_SQM * 2}, disciplines=["토목"])
        fixed = _by_driver(double, "토목", "fixed")
        assert fixed, "토목에 '식'·qty<=2 횟수성 항목이 존재해야 함"
        for it in fixed:
            assert it["unit"] == "식"
            assert it["qty"] == it["qty_basis"]["sample_qty"]
            assert it["qty_basis"]["ref_value"] is None
            assert it["qty_basis"]["project_value"] is None

    def test_식_대수량은_fixed_아님(self):
        """unit '식'이라도 qty_sample>2 면 fixed 제외(횟수성 아님) — gfa 스케일."""
        draft = generate_draft({"gfa_sqm": REF_GFA_SQM}, disciplines=["건축"])
        for it in _items(draft, "건축"):
            if it["unit"] == "식" and it["qty_basis"]["sample_qty"] > 2:
                assert it["qty_basis"]["driver"] != "fixed"


# ──────────────────────────────────────────────
# households 드라이버 — 세대 비례·미제공 폴백
# ──────────────────────────────────────────────


class TestHouseholdsDriver:
    def test_세대수_제공시_households_비례(self):
        draft = generate_draft(
            {"gfa_sqm": REF_GFA_SQM, "households": REF_HOUSEHOLDS * 2},
            disciplines=["기계소방"],
        )
        hh = _by_driver(draft, "기계소방", "households")
        assert hh, "기계소방 '단위세대' 항목이 households 드라이버여야 함"
        for it in hh:
            assert it["qty"] == round_qty(it["qty_basis"]["sample_qty"] * 2.0)
            assert it["qty_basis"]["ref_value"] == REF_HOUSEHOLDS
        assert not any("households" in w for w in draft["summary"]["warnings"])

    def test_미제공시_gfa_폴백_및_warning(self):
        draft = generate_draft({"gfa_sqm": REF_GFA_SQM}, disciplines=["기계소방"])
        assert not _by_driver(draft, "기계소방", "households")
        fallback = [
            it for it in _items(draft, "기계소방")
            if it["qty_basis"].get("fallback_from") == "households"
        ]
        assert fallback, "세대 항목이 gfa 폴백으로 표기되어야 함"
        for it in fallback:
            assert it["qty_basis"]["driver"] == "gfa"
            assert it["qty_basis"]["ref_value"] == REF_GFA_SQM
        assert any("households 미제공" in w for w in draft["summary"]["warnings"])


# ──────────────────────────────────────────────
# 조경(landscape_area) 드라이버 — 미제공 폴백
# ──────────────────────────────────────────────


class TestLandscapeDriver:
    def test_조경면적_미제공_폴백_warning(self):
        draft = generate_draft({"gfa_sqm": REF_GFA_SQM}, disciplines=["조경"])
        assert any(
            "조경면적 미제공 — 연면적 비례는 부정확" in w
            for w in draft["summary"]["warnings"]
        )
        non_fixed = [
            it for it in _items(draft, "조경")
            if it["qty_basis"]["driver"] != "fixed"
        ]
        assert non_fixed
        for it in non_fixed:
            assert it["qty_basis"]["driver"] == "gfa"
            assert it["qty_basis"].get("fallback_from") == "landscape_area"

    def test_조경면적_제공시_landscape_드라이버(self):
        draft = generate_draft(
            {"gfa_sqm": REF_GFA_SQM, "landscape_area_sqm": REF_LANDSCAPE_AREA_SQM * 2},
            disciplines=["조경"],
        )
        land = _by_driver(draft, "조경", "landscape_area")
        assert land
        for it in land:
            assert it["qty"] == round_qty(it["qty_basis"]["sample_qty"] * 2.0)
            assert it["qty_basis"]["ref_value"] == REF_LANDSCAPE_AREA_SQM
        assert not any("조경면적 미제공" in w for w in draft["summary"]["warnings"])

    def test_세대규칙이_조경규칙보다_우선(self):
        """규칙 순서: 세대 > 조경 > 식 > gfa (조경엔 세대 항목 없음 — 데이터 확인)."""
        draft = generate_draft(
            {"gfa_sqm": REF_GFA_SQM, "landscape_area_sqm": 1000.0},
            disciplines=["조경"],
        )
        drivers = {it["qty_basis"]["driver"] for it in _items(draft, "조경")}
        assert drivers <= {"landscape_area"}  # 폴백·세대 없음 → 전부 조경면적 비례


# ──────────────────────────────────────────────
# 응답 계약 · 정직성 · 입력 검증
# ──────────────────────────────────────────────


class TestContract:
    def test_전공종_기본_생성(self):
        draft = generate_draft({"gfa_sqm": 50000.0})
        assert list(draft["disciplines"]) == ALL_DISCIPLINES
        assert draft["summary"]["total_items"] == sum(
            d["item_count"] for d in draft["disciplines"].values()
        )
        assert draft["summary"]["total_items"] == 3997  # 5공종 고유항목 합계

    def test_badges_정직성(self):
        draft = generate_draft({"gfa_sqm": 50000.0}, disciplines=["조경"])
        assert draft["badges"]["note"] == "실적 1건 기반 원단위 초안 — 전문 적산 검토 필수"
        assert draft["badges"]["confidence"] == "낮음(n=1)"

    def test_provenance_동봉(self):
        draft = generate_draft({"gfa_sqm": 50000.0}, disciplines=["토목"])
        prov = draft["provenance"]
        assert "의정부동 424" in prov["name"]
        assert prov["sample_count"] == 1

    def test_params_used_와_ref_basis(self):
        draft = generate_draft(
            {"gfa_sqm": 50000.0, "households": 300, "site_area_sqm": 9000.0},
            disciplines=["조경"],
        )
        used = draft["summary"]["params_used"]
        assert used["gfa_sqm"] == 50000.0
        assert used["households"] == 300.0
        assert used["site_area_sqm"] == 9000.0
        assert used["landscape_area_sqm"] is None
        assert used["ref"]["gfa_sqm"]["value"] == REF_GFA_SQM
        assert "추정" in used["ref"]["households"]["basis"]  # 추정 출처 정직 명시
        assert "추정" in used["ref"]["landscape_area_sqm"]["basis"]

    def test_항목_출력_키(self):
        draft = generate_draft({"gfa_sqm": 50000.0}, disciplines=["조경"])
        it = _items(draft, "조경")[0]
        for key in ("id", "discipline", "section_code", "section_name",
                    "name", "spec", "unit", "qty", "qty_basis"):
            assert key in it
        assert it["discipline"] == "조경"

    def test_전기_ref_mat_price_보존_타공종_부재(self):
        draft = generate_draft(
            {"gfa_sqm": 50000.0}, disciplines=["전기통신소방", "조경"],
        )
        assert any("ref_mat_price" in it for it in _items(draft, "전기통신소방"))
        assert all("ref_mat_price" not in it for it in _items(draft, "조경"))

    def test_영문_별칭_수용(self):
        draft = generate_draft({"gfa_sqm": 50000.0}, disciplines=["landscape"])
        assert list(draft["disciplines"]) == ["조경"]

    def test_gfa_필수(self):
        with pytest.raises(ValueError):
            generate_draft({})
        with pytest.raises(ValueError):
            generate_draft({"gfa_sqm": 0})
        with pytest.raises(ValueError):
            generate_draft({"gfa_sqm": -10})

    def test_미존재_공종_ValueError(self):
        with pytest.raises(ValueError):
            generate_draft({"gfa_sqm": 1000.0}, disciplines=["플랜트"])

    def test_결정론_동일입력_동일출력(self):
        p = {"gfa_sqm": 77777.0, "households": 500}
        assert generate_draft(p, ["기계소방"]) == generate_draft(p, ["기계소방"])

    def test_레지스트리_미가용시_JSON_폴백_동일결과(self, monkeypatch):
        """B1 레지스트리 임포트 불가 환경에서도 동일 SSOT JSON 폴백으로 동작."""
        import app.services.cost.boq_parametric_engine as eng
        via_registry = eng.generate_draft({"gfa_sqm": 50000.0}, ["조경"])
        monkeypatch.setattr(eng, "_registry", lambda: None)
        via_json = eng.generate_draft({"gfa_sqm": 50000.0}, ["조경"])
        key = lambda d: [(i["id"], i["qty"], i["qty_basis"]) for i in d["disciplines"]["조경"]["items"]]  # noqa: E731
        assert key(via_json) == key(via_registry)
        assert via_json["provenance"]["sample_count"] == 1


# ──────────────────────────────────────────────
# 엑셀 내보내기 — PK 시그니처·시트수·공란 단가
# ──────────────────────────────────────────────


class TestExcelExport:
    @pytest.fixture(scope="class")
    def draft(self):
        return generate_draft({"gfa_sqm": 50000.0}, disciplines=["조경", "토목"])

    def test_bytes_pk_시그니처(self, draft):
        data = build_xlsx(draft)
        assert isinstance(data, bytes)
        assert data[:2] == b"PK"  # xlsx = zip 컨테이너

    def test_시트수_공종수_일치(self, draft):
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(build_xlsx(draft)))
        assert len(wb.sheetnames) == len(draft["disciplines"]) == 2
        assert set(wb.sheetnames) == {"조경", "토목"}

    def test_헤더_2행_포맷(self, draft):
        from openpyxl import load_workbook
        ws = load_workbook(BytesIO(build_xlsx(draft)))["조경"]
        assert ws["A1"].value == "품명"
        assert ws["B1"].value == "규격"
        assert ws["C1"].value == "단위"
        assert ws["D1"].value == "수량"
        assert ws["E1"].value == "재료비"
        assert ws["M1"].value == "비고"
        assert ws["E2"].value == "단가"
        assert ws["F2"].value == "금액"

    def test_단가_금액_공란(self, draft):
        """공내역서 표준 — 데이터 행의 단가·금액(E~L) 전부 공란."""
        from openpyxl import load_workbook
        ws = load_workbook(BytesIO(build_xlsx(draft)))["조경"]
        data_rows = 0
        for row in ws.iter_rows(min_row=3):
            if row[2].value:  # 단위 칸 채워진 행 = 항목 행
                data_rows += 1
                assert all(row[c].value in (None, "") for c in range(4, 12))
        assert data_rows == draft["disciplines"]["조경"]["item_count"]

    def test_하단_출처_주의문구(self, draft):
        from openpyxl import load_workbook
        ws = load_workbook(BytesIO(build_xlsx(draft)))["토목"]
        texts = [
            str(row[0].value) for row in ws.iter_rows(min_row=3)
            if row[0].value and str(row[0].value).startswith("※")
        ]
        joined = "\n".join(texts)
        assert "전문 적산 검토 필수" in joined
        assert "공내역서" in joined
        assert "의정부동 424" in joined

    def test_빈_draft_안내시트(self):
        data = build_xlsx({"disciplines": {}, "summary": {}, "badges": {}, "provenance": {}})
        assert data[:2] == b"PK"
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(data))
        assert len(wb.sheetnames) == 1
