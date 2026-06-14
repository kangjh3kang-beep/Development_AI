"""N3 단가DB 결합(boq_price_join.join_prices) 단위 테스트 — 결정론·정직성.

검증 범위(§3 N3 명세):
- 우선순위: ①키워드→단가키 매핑(단위 정합 시 DB/fallback 단가) ②전기 ref_mat_price
  단독 결합(재료비만) ③둘 다 없으면 빈칸(가짜 단가 금지).
- 단위 정합 가드: 키워드가 매칭돼도 단위 불일치면 단가 미적용(정직) — unit_mismatch 집계.
- coverage 통계(priced_count/total/coverage_pct/by_source).
- additive·비파괴: 기존 draft/아이템 키 0개 변경, 입력 객체 불변.
- 금액 엑셀: 단가/금액 칸 채움 + 공종별 소계 + 총계 행(기존 빈칸 모드는 default 유지).

단가 기준값(fallback SSOT = standard_quantity_estimator.UNIT_PRICES_2026):
  concrete(㎥) mat 85,000 + labor 35,000 + exp 12,000 = 132,000.
"""
from __future__ import annotations

import os
import sys
from io import BytesIO
from typing import Any

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from app.services.cost.boq_price_join import join_prices  # noqa: E402


def _draft(items_by_disc: dict[str, list[dict[str, Any]]]) -> dict[str, Any]:
    """generate_draft 와 동형(disciplines/summary/provenance/badges)의 합성 초안."""
    total = sum(len(v) for v in items_by_disc.values())
    return {
        "disciplines": {
            d: {"items": its, "item_count": len(its), "sections": []}
            for d, its in items_by_disc.items()
        },
        "summary": {"total_items": total, "params_used": {"gfa_sqm": 52000.0}, "warnings": []},
        "provenance": {"name": "의정부동 424 주상복합", "sample_count": 1},
        "badges": {"note": "실적 1건 기반 원단위 초안 — 전문 적산 검토 필수",
                   "confidence": "낮음(n=1)"},
    }


def _item(name: str, unit: str, qty: float, *, spec: str = "", **extra: Any) -> dict[str, Any]:
    base = {
        "id": extra.get("id", "x-0001"), "discipline": extra.get("discipline", "건축"),
        "section_code": "0101", "section_name": "공통가설공사",
        "name": name, "spec": spec, "unit": unit, "qty": qty,
        "qty_sample": qty, "driver": "gfa", "basis": "표본 비례", "confidence": "낮음(n=1)",
    }
    if "ref_mat_price" in extra:
        base["ref_mat_price"] = extra["ref_mat_price"]
    return base


def _all_items(priced: dict[str, Any]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for block in (priced.get("disciplines") or {}).values():
        out.extend(block.get("items") or [])
    return out


def _by_name(priced: dict[str, Any], name: str) -> dict[str, Any]:
    for it in _all_items(priced):
        if it.get("name") == name:
            return it
    raise AssertionError(f"항목 '{name}' 없음")


# ──────────────────────────────────────────────
# ① 키워드 매핑 + 단위 정합 → DB/fallback 단가 결합
# ──────────────────────────────────────────────

class TestKeywordJoin:
    def test_concrete_m3_가산_단가금액(self):
        d = _draft({"건축": [_item("레미콘 타설", "m3", 10.0, spec="25-24-15")]})
        out = join_prices(d)
        it = _by_name(out, "레미콘 타설")
        assert it["price_source"] == "fallback"
        assert it["price_key"] == "concrete"
        assert it["mat_unit"] == 85_000.0
        assert it["labor_unit"] == 35_000.0
        assert it["exp_unit"] == 12_000.0
        # 금액 = qty × (mat+labor+exp) = 10 × 132,000
        assert it["amount"] == 1_320_000

    def test_단위_불일치는_단가_미적용_정직(self):
        # '식' 단위 콘크리트 항목 — ㎥ 단가를 곱하면 날조 → 미적용(price_source None)
        d = _draft({"건축": [_item("콘크리트 가설", "식", 1.0)]})
        out = join_prices(d)
        it = _by_name(out, "콘크리트 가설")
        assert it["price_source"] is None
        assert it.get("amount") is None
        assert out["summary"]["pricing"]["unit_mismatch_count"] >= 1
        # 투명성: 미적용 사유 표기
        assert "단위" in (it.get("price_note") or "")

    def test_주입_DB_단가_출처_passthrough(self):
        d = _draft({"건축": [_item("레미콘 타설", "m3", 2.0)]})
        injected = {"concrete": {"key": "concrete", "spec": "25-24-15", "unit": "m3",
                                 "mat_unit": 90_000.0, "labor_unit": 30_000.0, "exp_unit": 10_000.0,
                                 "price_source": "db", "price_basis_year": 2026, "region": "경기도"}}
        out = join_prices(d, prices=injected)
        it = _by_name(out, "레미콘 타설")
        assert it["price_source"] == "db"
        assert it["amount"] == 2 * 130_000


# ──────────────────────────────────────────────
# ② 전기 ref_mat_price 단독 결합(재료비만 — 노무·경비 빈칸 정직)
# ──────────────────────────────────────────────

class TestRefMatPriceJoin:
    def test_ref_mat_price_재료비만(self):
        d = _draft({"전기통신소방": [
            _item("0.6/1KV PVC절연 접지용전선", "m", 100.0, spec="F-GV 10㎟",
                  ref_mat_price=1938.0, discipline="전기통신소방")]})
        out = join_prices(d)
        it = _by_name(out, "0.6/1KV PVC절연 접지용전선")
        assert it["price_source"] == "도면참고단가"
        assert it["mat_unit"] == 1938.0
        assert it["labor_unit"] is None  # 노무비 빈칸 — 정직
        assert it["exp_unit"] is None
        assert it["amount"] == 193_800  # 재료비만 = 100 × 1938

    def test_키워드_우선순위_ref_mat_price_보다_DB(self):
        # ref_mat_price 가 있어도 키워드+단위정합이면 DB 단가(①)가 우선
        d = _draft({"전기통신소방": [
            _item("레미콘 매립", "m3", 1.0, ref_mat_price=999.0, discipline="전기통신소방")]})
        out = join_prices(d)
        it = _by_name(out, "레미콘 매립")
        assert it["price_key"] == "concrete"
        assert it["price_source"] == "fallback"  # 도면참고단가 아님
        assert it["mat_unit"] == 85_000.0


# ──────────────────────────────────────────────
# ③ 미매칭 빈칸 + ④ 통계 + ⑤ 비파괴
# ──────────────────────────────────────────────

class TestCoverageAndPurity:
    def test_미매칭_항목_빈칸(self):
        d = _draft({"건축": [_item("가설국기게양시설", "식", 1.0)]})
        out = join_prices(d)
        it = _by_name(out, "가설국기게양시설")
        assert it["price_source"] is None
        assert it.get("amount") is None
        # 키워드 후보가 없으므로 unit_mismatch 가 아니다
        assert out["summary"]["pricing"]["unit_mismatch_count"] == 0

    def test_coverage_통계(self):
        d = _draft({"건축": [
            _item("레미콘 타설", "m3", 10.0),               # priced
            _item("철근 가공조립", "ton", 2.0, spec="SD400"),  # priced
            _item("가설국기게양시설", "식", 1.0),            # blank
            _item("콘크리트 가설", "식", 1.0),               # unit mismatch → blank
        ]})
        out = join_prices(d)
        p = out["summary"]["pricing"]
        assert p["total_items"] == 4
        assert p["priced_count"] == 2
        assert p["coverage_pct"] == 50.0
        assert p["by_source"].get("fallback") == 2
        assert p["unit_mismatch_count"] == 1
        # 부분 합계(priced 항목만) — 정직 표기
        assert p["priced_amount_won"] == 10 * 132_000 + 2 * (950_000 + 280_000 + 70_000)

    def test_additive_비파괴_입력불변(self):
        src_item = _item("레미콘 타설", "m3", 10.0)
        d = _draft({"건축": [src_item]})
        out = join_prices(d)
        # 입력 객체는 단가 키가 추가되지 않는다(비파괴)
        assert "amount" not in src_item
        assert "price_source" not in src_item
        # 기존 아이템 키 보존
        it = _by_name(out, "레미콘 타설")
        for k in ("id", "qty", "qty_sample", "driver", "basis", "confidence"):
            assert k in it
        # 기존 배지/출처 보존
        assert out["badges"]["confidence"] == "낮음(n=1)"
        assert out["provenance"]["sample_count"] == 1


# ──────────────────────────────────────────────
# 금액 엑셀 모드(boq_excel_export.build_xlsx priced=True)
# ──────────────────────────────────────────────

class TestExcelAmountMode:
    def _priced(self) -> dict[str, Any]:
        d = _draft({"건축": [
            _item("레미콘 타설", "m3", 10.0, spec="25-24-15"),
            _item("가설국기게양시설", "식", 1.0),
        ]})
        return join_prices(d)

    def test_빈칸모드_default_하위호환(self):
        # priced 인자 없이 호출 시 기존(단가 공란) 동작 — PK xlsx 반환
        from app.services.cost.boq_excel_export import build_xlsx
        data = build_xlsx(self._priced())
        assert data[:4] == b"PK\x03\x04"

    def test_금액모드_단가금액칸_및_소계총계(self):
        from openpyxl import load_workbook

        from app.services.cost.boq_excel_export import build_xlsx
        data = build_xlsx(self._priced(), priced=True)
        assert data[:4] == b"PK\x03\x04"
        wb = load_workbook(BytesIO(data))
        ws = wb["건축"]
        rows = [[c.value for c in row] for row in ws.iter_rows()]
        flat = ["" if v is None else str(v) for r in rows for v in r]
        # 단가/금액 칸이 실제로 채워졌는지 — 레미콘 행 재료비 단가(col5)=85,000, 금액(col6)=850,000
        mat_unit_cells = [r[4] for r in rows if r[0] == "레미콘 타설"]
        amount_cells = [r[5] for r in rows if r[0] == "레미콘 타설"]
        assert mat_unit_cells and mat_unit_cells[0] == 85_000
        assert amount_cells and amount_cells[0] == 850_000
        # 공종별 소계 + 총계 행 존재
        assert any("소계" in s for s in flat), "공종 소계 행 없음"
        assert any("총계" in s or "합계" in s for s in flat), "총계 행 없음"
