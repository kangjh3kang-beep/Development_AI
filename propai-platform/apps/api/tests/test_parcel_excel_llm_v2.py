"""토지조서 엑셀 LLM 보조 반복검증 v2 — 다양식 픽스처 매트릭스(T6).

검증 범위:
  T1 use_llm 게이트 — False면 LLM 0호출(비용/신뢰 보장).
  T2 S1 구조인식 — 시트선택·전치판정·복합셀분해·컬럼역할을 '한 번의 질의'로 처리(중복 프롬프트
    금지), 실존 검증 실패(가짜 시트/컬럼명)는 거부.
  T3 S3 반복검증 — 결정론 게이트(지번형식·PNU19자리·미해소상태)+합계행 감지+선별 재질의(최대2회,
    원문 부분문자열만 채택).
  T4 S4 분류 — verified/corrected/needs_review + verification_report(additive, 기존 키 불변).

픽스처는 openpyxl로 테스트 내 동적 생성(바이너리 커밋 금지). LLM은 전부 mock(get_llm 패치,
실제 API 미호출). VWorld는 결정적 stub(네트워크 없는 환경에서도 안정적으로 검증).
"""
from __future__ import annotations

import asyncio
import io
import json
import re

import pytest

from app.services.land_intelligence import parcel_excel_service as pes


# ── 공용 헬퍼 ────────────────────────────────────────────────────────────
def _xlsx(rows: list[list], sheet_title: str = "토지조서",
          extra_sheets: dict[str, list[list]] | None = None) -> bytes:
    """행 리스트 → xlsx 바이트(테스트 내 동적 생성)."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    for r in rows:
        ws.append(r)
    for name, rs in (extra_sheets or {}).items():
        s = wb.create_sheet(name)
        for r in rs:
            s.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class _StubVWorld:
    """결정적 VWorld 대역 — PNU가 직접 주어진 필지는 조회 성공(happy path 안정),
    주소만으로 지오코딩해야 하는 필지는 실패(네트워크 없는 테스트 환경 — 정직 실패 검증용).
    """

    async def geocode_address(self, query):
        return None

    async def get_land_characteristics(self, pnu):
        return {"zone_type": "제2종일반주거지역", "land_category": "대",
                "official_price_per_sqm": 1_000_000}

    async def search_address(self, query, size=8):
        return []


@pytest.fixture(autouse=True)
def _patch_vworld(monkeypatch):
    import app.services.external_api.vworld_service as vmod
    monkeypatch.setattr(vmod, "VWorldService", _StubVWorld)


@pytest.fixture(autouse=True)
def _clear_llm_struct_cache():
    """구조질의 캐시(_STRUCT_CACHE)는 모듈 전역 — 테스트 간 시그니처 충돌(다른 테스트의 mock
    응답이 캐시로 새는 것) 방지를 위해 매 테스트 전후로 비운다."""
    pes._STRUCT_CACHE.clear()
    yield
    pes._STRUCT_CACHE.clear()


def _fake_llm(responder):
    """ainvoke가 human 프롬프트 텍스트를 responder(human)->json문자열 에 넘겨 응답을 만드는 가짜 LLM."""

    class _Resp:
        def __init__(self, content: str):
            self.content = content
            self.usage_metadata = {"input_tokens": 10, "output_tokens": 5}

    class _LLM:
        model = "fake-model"

        async def ainvoke(self, messages):
            human = messages[-1].content if messages else ""
            return _Resp(responder(human))

    def _factory(*_a, **_k):
        return _LLM()

    return _factory


def _patch_llm(monkeypatch, responder):
    monkeypatch.setattr("app.services.ai.llm_provider.get_llm", _fake_llm(responder))


def _no_llm_reply(_human: str) -> str:
    """구조질의엔 응답 없음(호출되면 실패해야 하는 use_llm=False 테스트용)."""
    raise AssertionError("use_llm=False인데 LLM이 호출됨")


# ── ① 표준양식 — 규칙기반만으로 완결, LLM 0호출 ──────────────────────────
def test_standard_form_rule_based_zero_llm_calls(monkeypatch):
    monkeypatch.setattr("app.services.ai.llm_provider.get_llm", _fake_llm(_no_llm_reply))
    raw = _xlsx([
        ["연번", "소재지(주소)", "지번", "법정동코드(bcode·10자리)", "PNU(필지고유번호·19자리)", "지목", "면적(㎡)", "소유구분"],
        [1, "서울특별시 동작구 상도동", "210-453", "1159010300", "1159010300102100453", "대", "200", "사유"],
        [2, "경기도 의정부시 의정부동", "224-1", "4115010100", "4115010100201100224", "대", "150", "사유"],
    ])
    out = asyncio.run(pes.ParcelExcelService().parse(raw, "표준.xlsx", use_llm=False))
    assert not out.get("error")
    assert len(out["parcels"]) == 2
    assert out["column_engine"] == "rule"
    vr = out["verification_report"]
    assert vr["llm_used"] is False
    assert vr["passes"] == 0
    assert vr["counts"] == {"verified": 2, "corrected": 0, "needs_review": 0, "excluded": 0}
    assert all(p["injectable"] for p in out["parcels"])


def test_standard_form_use_llm_true_still_zero_calls_when_confident(monkeypatch):
    """use_llm=True라도 규칙기반이 이미 신뢰도 높으면 LLM을 호출하지 않는다(비용 보호)."""
    monkeypatch.setattr("app.services.ai.llm_provider.get_llm", _fake_llm(_no_llm_reply))
    raw = _xlsx([
        ["소재지(주소)", "지번", "PNU(필지고유번호·19자리)"],
        ["서울특별시 동작구 상도동", "210-453", "1159010300102100453"],
    ])
    out = asyncio.run(pes.ParcelExcelService().parse(raw, "표준2.xlsx", use_llm=True))
    assert not out.get("error")
    assert out["verification_report"]["llm_used"] is False


# ── ② 비표준 헤더(동의어 밖) — LLM 컬럼 역할 분류 ────────────────────────
def test_nonstandard_headers_llm_column_mapping(monkeypatch):
    def responder(_human: str) -> str:
        return json.dumps({"columns": {"address": "A열", "jibun": "B열", "pnu": "C열"}}, ensure_ascii=False)

    _patch_llm(monkeypatch, responder)
    raw = _xlsx([
        ["A열", "B열", "C열"],
        ["서울특별시 동작구 상도동", "210-453", "1159010300102100453"],
    ])
    out = asyncio.run(pes.ParcelExcelService().parse(raw, "비표준.xlsx", use_llm=True))
    assert not out.get("error")
    assert out["column_engine"] == "rule+llm"
    assert len(out["parcels"]) == 1
    assert out["parcels"][0]["address"] == "서울특별시 동작구 상도동"
    assert out["parcels"][0]["jibun"] == "210-453"


# ── ③ 세로형(전치) — LLM is_transposed 판정 → 결정론 전치 후 재파싱 ──────
def test_transposed_form_deterministic_transpose(monkeypatch):
    def responder(_human: str) -> str:
        return json.dumps({"is_transposed": True}, ensure_ascii=False)

    _patch_llm(monkeypatch, responder)
    raw = _xlsx([
        ["항목", "필지1", "필지2"],
        ["소재지", "서울특별시 동작구 상도동", "경기도 의정부시 의정부동 224"],
        ["지번", "210-453", "224"],
        ["PNU", "1159010300102100453", "4115010100201100224"],
        ["면적", "200", "300"],
        ["지목", "대", "대"],
    ])
    out = asyncio.run(pes.ParcelExcelService().parse(raw, "전치.xlsx", use_llm=True))
    assert not out.get("error")
    assert out["column_engine"] == "rule+llm"
    assert len(out["parcels"]) == 2
    addrs = {p["address"] for p in out["parcels"]}
    assert addrs == {"서울특별시 동작구 상도동", "경기도 의정부시 의정부동 224"}
    jibuns = {p["jibun"] for p in out["parcels"]}
    assert jibuns == {"210-453", "224"}


# ── ④ 다중시트(2번째가 토지조서) — LLM sheet_name 재선택(실존 검증) ──────
def test_multi_sheet_llm_reselects_correct_sheet(monkeypatch):
    def responder(_human: str) -> str:
        return json.dumps({"sheet_name": "필지목록"}, ensure_ascii=False)

    _patch_llm(monkeypatch, responder)
    raw = _xlsx(
        rows=[["프로젝트 개요"], ["작성일자: 2026-01-01"], ["담당자: 홍길동"]],
        sheet_title="표지",
        extra_sheets={
            "필지목록": [
                ["소재지(주소)", "지번", "PNU(필지고유번호·19자리)"],
                ["서울특별시 동작구 상도동", "210-453", "1159010300102100453"],
                ["경기도 의정부시 의정부동 224", "224-1", "4115010100201100224"],
            ],
        },
    )
    out = asyncio.run(pes.ParcelExcelService().parse(raw, "다중시트.xlsx", use_llm=True))
    assert not out.get("error")
    assert out["column_engine"] == "rule+llm"
    assert len(out["parcels"]) == 2


# ── ⑤ 복합셀("의정부동 224-1 대 500㎡") — LLM regex 제안 → 매치율≥60% 채택 ─
def test_compound_cell_decomposition_applied_when_match_rate_high(monkeypatch):
    def responder(human: str) -> str:
        if "compound_cell" in human:
            return json.dumps({
                "compound_cell": {
                    "column": "복합정보",
                    "regex": r"(?P<addr>[가-힣]+동)\s+(?P<jibun>\d+(-\d+)?)\s+(?P<jimok>[가-힣])\s+(?P<area>\d+)㎡",
                },
            }, ensure_ascii=False)
        return "{}"

    _patch_llm(monkeypatch, responder)
    raw = _xlsx([
        ["복합정보"],
        ["의정부동 224-1 대 500㎡"],
        ["상도동 210-453 대 200㎡"],
    ])
    out = asyncio.run(pes.ParcelExcelService().parse(raw, "복합.xlsx", use_llm=True))
    assert not out.get("error")
    assert out["column_engine"] == "rule+llm"
    assert len(out["parcels"]) == 2
    addrs = {p["address"] for p in out["parcels"]}
    assert addrs == {"의정부동", "상도동"}
    jibuns = {p["jibun"] for p in out["parcels"]}
    assert jibuns == {"224-1", "210-453"}
    areas = {p["area_sqm"] for p in out["parcels"]}
    assert areas == {500.0, 200.0}


def test_compound_cell_discarded_when_match_rate_below_threshold(monkeypatch):
    """매치율<60%면 폐기 — 채택되지 않아 결국 필수컬럼(address/pnu/bcode) 미확보 에러."""
    def responder(human: str) -> str:
        if "compound_cell" in human:
            return json.dumps({
                "compound_cell": {"column": "복합정보", "regex": r"(?P<addr>동작구)"},
            }, ensure_ascii=False)
        return "{}"

    _patch_llm(monkeypatch, responder)
    raw = _xlsx([
        ["복합정보"],
        ["서울특별시 동작구 상도동 210-453"],
        ["경기도 의정부시 의정부동 224-1"],
        ["부산광역시 해운대구 우동 500"],
    ])
    out = asyncio.run(pes.ParcelExcelService().parse(raw, "복합폐기.xlsx", use_llm=True))
    assert out.get("error"), "매치율 1/3(33%)<60% 이므로 복합셀 분해가 폐기되고 필수컬럼 에러여야 함"


# ── ⑥ 합계행 오염 — 키워드/누적합 두 경로 모두 제외 + 대조 경고 ─────────
def test_summary_row_excluded_by_keyword_and_mismatch_warning():
    raw = _xlsx([
        ["소재지(주소)", "지번", "PNU(필지고유번호·19자리)", "면적(㎡)"],
        ["서울특별시 동작구 상도동", "210-453", "1159010300102100453", "200"],
        ["경기도 의정부시 의정부동 224", "224-1", "4115010100201100224", "150"],
        ["합계", "", "", "999"],
    ])
    out = asyncio.run(pes.ParcelExcelService().parse(raw, "합계.xlsx", use_llm=False))
    assert not out.get("error")
    assert len(out["parcels"]) == 2
    vr = out["verification_report"]
    assert vr["counts"]["excluded"] == 1
    assert any("합계" in w and "차이" in w for w in vr["warnings"])


def test_summary_row_excluded_by_cumulative_area_match():
    """키워드가 없어도(예: '집계') 면적이 상위행 누적합과 ±1% 이내면 집계행으로 감지·제외."""
    raw = _xlsx([
        ["소재지(주소)", "지번", "PNU(필지고유번호·19자리)", "면적(㎡)"],
        ["서울특별시 동작구 상도동", "210-453", "1159010300102100453", "200"],
        ["경기도 의정부시 의정부동 224", "224-1", "4115010100201100224", "150"],
        ["집계", "", "", "350"],
    ])
    out = asyncio.run(pes.ParcelExcelService().parse(raw, "집계.xlsx", use_llm=False))
    assert not out.get("error")
    assert len(out["parcels"]) == 2
    assert out["verification_report"]["counts"]["excluded"] == 1


# ── ⑦ 병합셀 — forward-fill(공유지분) 회귀 확인(_rebuild 리팩토링 후에도 유지) ─
def test_merged_cells_forward_fill():
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "토지조서"
    ws.append(["소재지(주소)", "지번", "PNU(필지고유번호·19자리)", "소유구분"])
    ws.append(["서울특별시 동작구 상도동", "210-453", "1159010300102100453", "김철수"])
    ws.append(["", "", "", "이영희"])
    ws.merge_cells("A2:A3")
    ws.merge_cells("B2:B3")
    ws.merge_cells("C2:C3")
    buf = io.BytesIO()
    wb.save(buf)
    raw = buf.getvalue()

    out = asyncio.run(pes.ParcelExcelService().parse(raw, "병합.xlsx", use_llm=False))
    assert not out.get("error")
    assert len(out["parcels"]) == 2
    assert out["parcels"][0]["jibun"] == out["parcels"][1]["jibun"] == "210-453"
    assert all(p.get("co_owner") for p in out["parcels"]), "같은 PNU 공유지분(병합 복원)으로 표시돼야 함"


# ── ⑦-b 병합 복원이 안 되는 엑셀 — 원문 복원 + 침묵 금지 + 위양성 금지 ─────
def _merged_two_owner_xlsx(merged: bool = True) -> bytes:
    """지번·소재지를 두 행에 세로 병합한(또는 안 한) 실제 토지조서 형태."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "토지조서"
    ws.append(["소재지(주소)", "지번", "소유구분"])
    ws.append(["서울특별시 동작구 상도동", "210-453", "김철수"])
    ws.append(["", "", "이영희"] if merged else ["서울특별시 동작구 상도동", "210-454", "이영희"])
    if merged:
        ws.merge_cells("A2:A3")
        ws.merge_cells("B2:B3")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _break_full_workbook_read(raw: bytes, part: str = "sheet1.xml") -> bytes:
    """★모킹이 아닌 '진짜 손상 입력' — 셀에 styles.xml 에 없는 서식번호(s="9999")를 심는다.

    엑셀을 표로만 읽는 pandas 경로(openpyxl read_only)는 서식을 찾아보지 않아 통과하지만,
    병합범위를 얻으려면 필요한 '전체 읽기'(read_only=False)는 그 번호를 서식표에서 찾다가
    IndexError 로 죽는다. 오피스가 아닌 도구로 내보낸 엑셀에서 실제로 나오는 형태다.
    """
    import zipfile

    zin = zipfile.ZipFile(io.BytesIO(raw))
    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zo:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.endswith(part):
                data = data.replace(b'<c r="A1"', b'<c s="9999" r="A1"', 1)
            zo.writestr(item.filename, data)
    return out.getvalue()


def _parse_x(raw: bytes, name: str = "t.xlsx") -> dict:
    return asyncio.run(pes.ParcelExcelService().parse(raw, name, use_llm=False))


def _warns(out: dict) -> list[str]:
    return out.get("verification_report", {}).get("warnings", [])


def test_merged_expand_failure_recovers_from_source_and_reports():
    """전체읽기가 죽어도 병합범위는 원문(zip XML)에 있다 → 지번을 복원하고 그 사실을 알린다."""
    ok = _parse_x(_merged_two_owner_xlsx(), "병합.xlsx")
    bad = _parse_x(_break_full_workbook_read(_merged_two_owner_xlsx()), "병합_손상.xlsx")

    # 전제(공허한 참 방지): 두 입력 모두 읽혀야 두 모집단 비교가 성립한다.
    assert not ok.get("error"), f"정상 입력이 읽혀야 함: {ok.get('error')}"
    assert not bad.get("error"), f"손상 입력도 표 자체는 읽혀야 함: {bad.get('error')}"

    # 전제: 성공 모집단에서 병합 복원이 실제로 일어나야 한다(복원 대상 0이면 잠금이 공허).
    ok_jibun = [p.get("jibun") for p in ok["parcels"]]
    assert ok_jibun == ["210-453", "210-453"], f"병합된 지번이 두 행 모두에 복원돼야 함: {ok_jibun}"

    # ★핵심: 손상 파일에서도 '원문 복원'으로 지번이 살아야 한다(행 탈락 0).
    #   사용자에게 "엑셀을 고쳐 오라"고 떠넘기지 않는다 — 복원 가능한 것은 복원한다.
    bad_jibun = [p.get("jibun") for p in bad["parcels"]]
    assert bad_jibun == ok_jibun, f"원문에서 복원했어야 함(떠넘기기 금지): {bad_jibun}"

    # ★두 모집단이 '다른 결과'를 낸다 — 경고 유무. 같으면 배선을 끊어도 통과한다.
    assert not _warns(ok), f"정상 파일엔 경고가 없어야 함: {_warns(ok)}"
    note = [w for w in _warns(bad) if "원문에서 직접 복원" in w]
    assert note, f"복원 사실을 알려야 한다(침묵 금지): {_warns(bad)}"
    assert "확인해 주세요" in note[0], f"확인을 권해야 한다: {note[0]}"
    # ★복원됐는데 '실패'라고 말하면 거짓말이다.
    assert not [w for w in _warns(bad) if "복원 실패" in w], f"복원했는데 실패라 함: {_warns(bad)}"


def test_no_merged_cells_never_warns_about_merges():
    """★위양성 대조군 — 병합이 0개인 파일에 '병합을 해제하라'고 하면 안 된다.

    같은 손상을 넣어도 병합이 없으면 잃은 값이 없다. 이 대조군이 없으면 '없는 병합을 풀라'는
    엉뚱한 지시가 멀쩡한 파일에까지 뜨는 것을 못 잡는다(가드의 위양성도 결함이다).
    """
    plain = _parse_x(_merged_two_owner_xlsx(merged=False), "병합없음.xlsx")
    plain_bad = _parse_x(_break_full_workbook_read(_merged_two_owner_xlsx(merged=False)),
                         "병합없음_손상.xlsx")

    # 전제: 두 입력 모두 데이터가 온전해야 '경고 0'이 의미를 갖는다(공허한 참 방지).
    for label, out in (("정상", plain), ("손상", plain_bad)):
        assert not out.get("error"), f"{label} 입력이 읽혀야 함"
        assert [p.get("jibun") for p in out["parcels"]] == ["210-453", "210-454"], (
            f"{label}: 병합이 없으므로 데이터가 100% 온전해야 함"
        )
    assert not [w for w in _warns(plain_bad) if "병합" in w], (
        f"병합이 0개인데 병합 경고가 떴다(위양성): {_warns(plain_bad)}"
    )


def test_merged_expand_reports_failure_when_source_recovery_also_fails(monkeypatch):
    """원문 복원마저 실패하면 그때는 정직하게 '복원 실패'를 말해야 한다.

    ★진짜 손상 입력으로는 이 갈래를 못 만든다(병합범위 XML 이 깨지면 pandas 도 같이 죽어
    파일 자체가 안 읽힌다 — 실측). 그래서 이 한 갈래만 폴백 함수를 막아 태운다.
    """
    def _boom(_raw):
        raise RuntimeError("zip 폴백 불가")

    monkeypatch.setattr(pes, "_merged_ranges_from_zip", _boom)
    bad = _parse_x(_break_full_workbook_read(_merged_two_owner_xlsx()), "복원불가.xlsx")
    assert not bad.get("error")
    fail = [w for w in _warns(bad) if "병합 셀 복원 실패" in w]
    assert fail, f"복원 실패를 알려야 한다: {_warns(bad)}"
    msg = fail[0]
    # 문구는 세 절을 다 말해야 한다 — 한 단어만 보면 여러 줄이 그 단어를 나눠 갖고 있어
    # 어느 한 줄을 지워도 통과한다(실제로 변이 생존으로 드러났다).
    assert "빈칸" in msg, f"②무엇을 잃는지: {msg}"
    assert "병합을 해제" in msg, f"③무엇을 해야 하는지: {msg}"
    assert "직접 적어" in msg, f"③실행지시: {msg}"
    assert "확인해 주세요" in msg, f"①먼저 확인을 권해야 한다: {msg}"
    # ★모르는 수를 지어내지 않는다 — 병합이 몇 곳인지 못 셌으면 개수를 말하지 않는다.
    assert "0곳" not in msg, f"셀 수 없는데 개수를 지어냄: {msg}"


def test_early_missing_column_return_still_carries_warnings():
    """★파서에서 가장 조용한 경로 — 필수 컬럼 부재 조기 반환이 사유를 버리면 안 된다.

    이 경로는 verification_report 자체가 없어서, 여기서 structure_notes 를 버리면 사용자는
    '필수 컬럼을 찾지 못했습니다'라는 **틀린 사유**만 받는다(진짜 원인은 따로 있는데).
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "토지조서"
    ws.append(["항목", "값", "비고"])          # 소재지·PNU·법정동코드 어느 것도 없다
    ws.append(["연면적", "1000", ""])
    ws.append(["", "", "메모"])
    ws.merge_cells("A2:A3")
    buf = io.BytesIO()
    wb.save(buf)
    out = _parse_x(_break_full_workbook_read(buf.getvalue()), "컬럼없음_손상.xlsx")

    assert out.get("error"), "필수 컬럼이 없으므로 error 여야 함(전제)"
    assert "warnings" in out, "조기 반환도 warnings 키를 실어야 한다(현재 침묵)"
    assert any("병합" in w or "복원" in w for w in out["warnings"]), (
        f"진짜 사유(병합/서식을 못 읽음)가 사용자에게 닿아야 한다: {out.get('warnings')}"
    )


def test_expand_merged_cells_returns_no_note_on_success():
    """함수 층 직접 확인 — 성공 경로는 사유 None(호출측이 경고를 싣지 않는다)."""
    import pandas as pd

    raw = _merged_two_owner_xlsx()
    df0 = pd.read_excel(io.BytesIO(raw), dtype=str, engine="openpyxl", header=None)
    df = df0.iloc[1:].reset_index(drop=True)
    df.columns = [str(v) for v in df0.iloc[0].tolist()]
    filled, note = pes._expand_merged_cells(raw, df, header_row=0)
    assert note is None, f"성공인데 사유가 생김: {note}"
    assert str(filled.iat[1, 1]) == "210-453", "성공 경로가 실제로 병합을 채워야 한다(공허 방지)"


def test_llm_sheet_reselect_reread_failure_is_reported(monkeypatch):
    """LLM 이 고른 시트를 다시 읽지 못하면 그 사실을 알려야 한다(조용히 옛 시트 결과 금지).

    ★이 분기가 발화하는 이유는 두 경로가 '읽는 양'이 다르기 때문이다 —
    미리보기(_sheet_previews_xlsx)는 15행에서 끊고(max_rows) pd.read_excel 은 전 행을 읽는다.
    그래서 16행 이후가 깨진 시트는 **미리보기는 되는데 본문은 못 읽는** 상태가 된다.
    (병합 건의 'read_only vs 전체읽기' 비대칭과 같은 종류의 결함이다 — 예전 skip 사유
    "두 경로가 같은 openpyxl 을 쓰므로 실패 입력을 만들 수 없다"는 **틀렸다**.)
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "표지"
    ws.append(["프로젝트 개요"])
    ws.append(["작성일자: 2026-01-01"])
    s2 = wb.create_sheet("필지목록")
    s2.append(["소재지(주소)", "지번", "면적(㎡)"])
    for i in range(1, 26):  # 26행 — 미리보기 캡(15행) 훨씬 뒤까지
        s2.append(["서울특별시 동작구 상도동", f"210-{400 + i}", 100 + i])
    buf = io.BytesIO()
    wb.save(buf)

    import zipfile

    zin = zipfile.ZipFile(io.BytesIO(buf.getvalue()))
    out_buf = io.BytesIO()
    injected = False
    with zipfile.ZipFile(out_buf, "w", zipfile.ZIP_DEFLATED) as zo:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.endswith("sheet2.xml"):
                new_data = data.replace(b"<v>125</v>", b"<v>NOT_A_NUMBER</v>", 1)
                injected = new_data != data
                data = new_data
            zo.writestr(item.filename, data)
    raw = out_buf.getvalue()
    assert injected, "손상 주입이 실제로 됐는지 먼저 확인(주입 실패를 생존으로 오독 방지)"

    # 전제(공허한 참 방지): 미리보기는 되고 본문 읽기는 실패해야 이 분기가 열린다.
    previews = pes._sheet_previews_xlsx(raw)
    assert "필지목록" in previews, "미리보기에 시트가 보여야 LLM 이 그 시트를 고를 수 있다"
    import pandas as pd
    reread_failed = False
    try:
        pd.read_excel(io.BytesIO(raw), dtype=str, engine="openpyxl", header=None,
                      sheet_name="필지목록")
    except Exception:  # noqa: BLE001 — 예외 '종류'가 아니라 '본문을 못 읽는다'는 사실이 전제다
        reread_failed = True
    assert reread_failed, "본문 읽기가 실패해야 이 분기가 열린다(전제)"

    _patch_llm(monkeypatch, lambda _h: json.dumps({"sheet_name": "필지목록"}, ensure_ascii=False))
    out = asyncio.run(pes.ParcelExcelService().parse(raw, "시트재선택실패.xlsx", use_llm=True))

    notes = out.get("verification_report", {}).get("warnings", []) or out.get("warnings", [])
    hit = [w for w in notes if "다시 읽지 못해" in w]
    assert hit, f"시트를 다시 읽지 못한 사실을 알려야 한다(침묵 금지): {notes}"
    assert "필지목록" in hit[0], f"어느 시트인지 말해야 한다: {hit[0]}"
    assert "확인이 필요합니다" in hit[0], f"확인을 권해야 한다: {hit[0]}"


def _merge_edge_case_xlsx() -> bytes:
    """경계 가드를 태우는 픽스처 — 제목행 병합·빈 좌상단·표 밖으로 넘친 병합.

    현실의 토지조서에 흔한 형태다(제목을 세로로 합치거나, 표 오른쪽 여백까지 병합).
    제목 2행을 둬서 머리글이 3행이 되게 한다 → 제목 병합의 행 번호가 표 기준으로 **음수**가 된다.
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "토지조서"
    ws.append(["토지조서", "", "관리번호 A-1"])   # 1행: 제목(C열에 값이 있어야 행 경계를 태운다)
    ws.append(["작성일: 2026-01-01"])           # 2행: 부제
    ws.append(["소재지(주소)", "지번", "비고", "소유구분", "메모", "여백"])  # 3행: 머리글
    ws.append(["서울특별시 동작구 상도동", "210-453", "", "김철수", "", "끝"])
    ws.append(["", "", "", "이영희", "", ""])
    # ★데이터 행을 3개로 둔다. 행이 2개뿐이면 음수 인덱스가 범위를 벗어나 예외가 나서,
    #   '조용히 뒤에서부터 감기는' 진짜 위험이 드러나지 않는다(변이 생존으로 확인했다).
    ws.append(["", "", "", "박민수", "", ""])
    ws.merge_cells("A4:A6")   # 정상(데이터 행 안)
    ws.merge_cells("B4:B6")   # 정상
    # ★제목 영역만 병합 → 표 기준 행번호가 전부 음수. 좌상단(C1)에 값을 둬야 '빈 좌상단'
    #   가드에서 먼저 빠져나가지 않고 **행 경계 가드가 실제로 실행**된다.
    ws.merge_cells("C1:C2")
    ws.merge_cells("E4:E6")   # ★좌상단이 빈칸 → 채울 값이 없다
    ws.merge_cells("F4:H6")   # ★표 오른쪽 밖(G·H)까지 넘침 → 열 범위초과
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_merge_fill_respects_table_boundaries():
    """★경계 가드 — 지우면 음수 인덱스가 뒤에서부터 감기거나 표 밖을 짚어 터진다.

    상·하한은 한 쌍이다(한쪽만 걸면 반대쪽이 무제한). 행·열 양방향과 '빈 좌상단'을 함께 본다.
    """
    import pandas as pd

    raw = _merge_edge_case_xlsx()
    df0 = pd.read_excel(io.BytesIO(raw), dtype=str, engine="openpyxl", header=None)
    hdr = pes._detect_header_row(df0)
    # 전제(공허한 참 방지): 머리글이 3행(0-based 2)으로 잡혀야 제목 병합이 '음수 행'이 된다.
    assert hdr == 2, f"제목 2행 뒤가 머리글이어야 이 픽스처가 경계를 태운다: hdr={hdr}"

    df = df0.iloc[hdr + 1:].reset_index(drop=True)
    df.columns = [str(v) for v in df0.iloc[hdr].tolist()]
    ncols_before, nrows_before = df.shape[1], df.shape[0]

    # 전제(가드 우회 방지): 제목 병합의 좌상단에 값이 있어야 행 경계 가드가 실행된다.
    assert str(df0.iat[0, 2]).strip() not in ("", "nan"), "제목 병합 좌상단이 비면 경계를 못 태운다"

    filled, note = pes._expand_merged_cells(raw, df, header_row=hdr)

    # ★가드를 지우면 표 밖을 짚어 예외가 나고, 예외는 사유 문구로 새어 나온다 —
    #   정상 파일에서 사유가 생기면 그 자체가 경계 위반의 신호다.
    assert note is None, f"정상 파일인데 사유가 생김(경계 위반 의심): {note}"

    # 전제: 정상 병합은 실제로 채워져야 한다 — 안 채워지면 아래 단언들이 무의미하다.
    assert str(filled.iat[1, 1]) == "210-453", "데이터 행 안의 병합은 채워져야 한다"

    # ① 제목 영역만의 병합(C1:C2): 표 안으로 한 칸도 넘어오지 않으므로 **아무것도 채우면 안 된다**.
    #    ★가드를 지우면 음수 인덱스가 뒤에서부터 감겨 제목값("관리번호 A-1")이 비고 칸에
    #    조용히 들어앉는다 — 예외도 안 나고 경고도 없는 날조다.
    for r in range(len(filled)):
        assert str(filled.iat[r, 2]).strip().lower() in ("", "nan"), (
            f"제목 영역 병합값이 표 안으로 샜다(음수 인덱스 되감기): {filled.iat[r, 2]!r}"
        )
    # ② 빈 좌상단(E4:E6): 채울 값이 없으면 아무것도 넣지 않는다("None"·"nan" 날조 금지).
    for r in range(len(filled)):
        assert str(filled.iat[r, 4]).strip().lower() in ("", "nan"), (
            f"빈 병합에 가짜 글자가 들어갔다: {filled.iat[r, 4]!r}"
        )
    # ③ 표 밖으로 넘친 병합(F4:H6): 열/행이 늘거나 옆 열을 덮으면 안 된다.
    assert filled.shape == (nrows_before, ncols_before), "병합이 표 밖으로 넘쳤다고 표가 커지면 안 된다"
    assert str(filled.iat[1, 3]) == "이영희", "옆 열(소유구분)이 병합값으로 덮이면 안 된다"


def test_partial_recovery_reports_how_many_were_lost():
    """원문 복원이 **일부만** 되면 몇 곳을 못 읽었는지 말해야 한다(모르면 말하지 않는다)."""
    out = _parse_x(_break_full_workbook_read(_merge_edge_case_xlsx()), "부분복원.xlsx")
    assert not out.get("error"), "표 자체는 읽혀야 함(전제)"

    # 전제: 복원 가능한 병합(A4:A6·B4:B6)은 실제로 복원돼야 '부분복원' 상황이 성립한다.
    assert [p.get("jibun") for p in out["parcels"]] == ["210-453"] * 3, (
        f"복원 가능한 지번은 원문에서 살아야 한다: {[p.get('jibun') for p in out['parcels']]}"
    )
    fail = [w for w in _warns(out) if "병합 셀 복원 실패" in w]
    assert fail, f"일부 복원 실패를 알려야 한다: {_warns(out)}"
    # ★개수를 실제로 센다 — 셀 수 있을 때는 말하고, 못 세면 말하지 않는다(위 대조 테스트).
    assert "1곳" in fail[0], f"못 읽은 병합 수를 말해야 한다: {fail[0]}"
    # ★어느 칸인지 모르면서 '지번 칸을 못 읽었다'고 단정하지 않는다 — 실제로 지번은 복원됐다.
    assert "지번·소재지 칸" not in fail[0], f"측정하지 않은 것을 단정함: {fail[0]}"


def _two_sheet_reordered_xlsx() -> bytes:
    """★반례 픽스처 — 파일 이름 순서와 시트 순서가 **어긋난** 워크북.

    표지가 `sheet1.xml`, 토지조서가 `sheet2.xml` 로 저장되지만, `workbook.xml` 의 `<sheet>`
    나열 순서만 뒤집어 **첫 시트를 토지조서**로 만든다. 엑셀 파일 형식에서 파일 이름과 시트
    순서는 무관하므로 이런 파일은 정상이다(비Office 도구가 내보내면 실제로 나온다).
    그리고 토지조서 쪽에만 서식 손상을 심어 원문 복원 경로를 태운다.
    """
    import re as _re
    import zipfile

    from openpyxl import Workbook

    wb = Workbook()
    cover = wb.active
    cover.title = "표지"
    cover.append(["프로젝트", "", "", ""])
    cover.append(["담당", "홍길동", "", ""])
    cover.append(["", "", "", ""])
    cover.merge_cells("D2:E3")          # 표지에만 있는 병합 — 여기가 새면 날조가 된다
    land = wb.create_sheet("토지조서")
    land.append(["소재지(주소)", "지번", "비고", "소유구분"])
    land.append(["서울특별시 동작구 상도동", "210-453", "", "김철수"])
    land.append(["", "", "", "이영희"])
    land.merge_cells("A2:A3")
    land.merge_cells("B2:B3")
    land.merge_cells("C2:C3")
    buf = io.BytesIO()
    wb.save(buf)

    zin = zipfile.ZipFile(io.BytesIO(buf.getvalue()))
    out = io.BytesIO()
    flipped = corrupted = False
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zo:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "xl/workbook.xml":
                tags = _re.findall(rb"<sheet [^>]*/>", data)
                assert len(tags) == 2, f"시트 태그 2개여야 함: {tags}"
                data = data.replace(tags[0] + tags[1], tags[1] + tags[0])
                flipped = True
            if item.filename.endswith("sheet2.xml"):
                new_data = data.replace(b'<c r="A1"', b'<c s="9999" r="A1"', 1)
                corrupted = new_data != data
                data = new_data
            zo.writestr(item.filename, data)
    assert flipped and corrupted, "픽스처 주입이 실제로 됐는지 확인(주입 실패 오독 방지)"
    return out.getvalue()


def test_zip_fallback_reads_the_same_sheet_pandas_read():
    """★시트를 **추측하지 않는다** — 표를 읽은 시트와 병합을 읽은 시트가 같아야 한다.

    파일 이름 순서로 첫 시트를 고르면, 표는 `토지조서`에서 읽고 병합은 `표지`에서 읽는 어긋남이
    생긴다. 그러면 세 가지가 한꺼번에 일어난다 — ①원문에 없던 값이 다른 칸에 복사되고(날조)
    ②진짜 병합이 적용 안 돼 행이 탈락하며 ③그런데도 "원문에서 직접 복원했습니다"라고
    **거짓 보고**한다(침묵보다 나쁘다).
    """
    import pandas as pd

    raw = _two_sheet_reordered_xlsx()

    # 전제(반례 성립 확인): 파일 이름 순서와 실제 첫 시트가 어긋나야 이 테스트가 의미를 갖는다.
    assert pd.ExcelFile(io.BytesIO(raw), engine="openpyxl").sheet_names[0] == "토지조서", (
        "첫 시트가 토지조서여야 반례가 성립한다"
    )
    import zipfile
    with zipfile.ZipFile(io.BytesIO(raw)) as z:
        part = pes._first_worksheet_part(z)
    assert part == "xl/worksheets/sheet2.xml", (
        f"첫 시트는 파일 이름이 sheet2.xml 이다(이름 순서로 고르면 sheet1.xml 을 집는다): {part}"
    )

    # ★병합범위는 토지조서 것이어야 한다 — 표지의 D2:E3 가 섞이면 날조가 시작된다.
    refs = pes._merged_ranges_from_zip(raw)
    assert sorted(refs) == ["A2:A3", "B2:B3", "C2:C3"], f"다른 시트의 병합을 집었다: {refs}"
    assert "D2:E3" not in refs, "표지(다른 시트)의 병합이 새어 들어왔다"

    out = _parse_x(raw, "시트순서뒤집힘.xlsx")
    assert not out.get("error")
    # ① 행 탈락 0 — 진짜 병합이 적용돼 둘째 행이 산다.
    assert [p.get("jibun") for p in out["parcels"]] == ["210-453", "210-453"], (
        f"병합 복원이 어긋나 행이 탈락했다: {[p.get('jibun') for p in out['parcels']]}"
    )
    # ② 날조 0 — 원문에 없던 값이 다른 칸에 복사되면 안 된다(비고 칸은 비어 있어야 한다).
    for p in out["parcels"]:
        assert not p.get("label"), f"원문에 없던 값이 만들어졌다: {p.get('label')!r}"
    # ③ 보고가 참이어야 한다 — 실제로 복원됐으니 '복원했습니다'가 맞다.
    assert [w for w in _warns(out) if "원문에서 직접 복원" in w], _warns(out)


def test_single_sheet_control_still_resolves_first_sheet():
    """대조군 — 단일 시트에서도 같은 해석 경로가 정답을 낸다(반례만 통과하는 잠금 방지)."""
    import zipfile

    raw = _merged_two_owner_xlsx()
    with zipfile.ZipFile(io.BytesIO(raw)) as z:
        assert pes._first_worksheet_part(z) == "xl/worksheets/sheet1.xml"
    assert sorted(pes._merged_ranges_from_zip(raw)) == ["A2:A3", "B2:B3"]


def test_unresolvable_first_sheet_fails_loudly_not_silently():
    """★첫 시트를 해석 못 하면 '병합 없음'이라 답하면 안 된다 — 그게 곧 침묵이다.

    빈 목록을 돌려주면 호출측이 "잃은 값이 없다"로 읽어 경고 없이 넘어간다. 그래서 예외를
    내고, 호출측은 정직한 실패 경고로 바꾼다.
    """
    import zipfile

    raw = _merged_two_owner_xlsx()
    broken = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(raw)) as zin, zipfile.ZipFile(broken, "w") as zo:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "xl/_rels/workbook.xml.rels":
                data = b'<?xml version="1.0"?><Relationships/>'  # 관계를 지운다
            zo.writestr(item.filename, data)

    with pytest.raises(ValueError):
        pes._merged_ranges_from_zip(broken.getvalue())


def _rewrite_part(raw: bytes, part: str, data: bytes) -> bytes:
    """xlsx(zip) 안의 한 파트만 갈아끼운다(테스트용)."""
    import zipfile

    zin = zipfile.ZipFile(io.BytesIO(raw))
    out = io.BytesIO()
    replaced = False
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zo:
        for item in zin.infolist():
            body = zin.read(item.filename)
            if item.filename == part:
                body, replaced = data, True
            zo.writestr(item.filename, body)
    assert replaced, f"교체 대상 파트가 없다: {part}"
    return out.getvalue()


def test_first_sheet_resolution_handles_both_target_forms():
    """관계의 Target 이 절대(`/xl/...`)든 상대(`./worksheets/...`)든 같은 시트를 짚어야 한다.

    ★엑셀 파일 형식은 둘 다 허용한다. 실측하니 openpyxl 이 저장한 파일은 **절대경로**였는데,
    다른 도구는 상대경로로 쓴다(이 PR 의 대상이 바로 '비Office 도구가 내보낸 엑셀'이다).
    경로 정규화를 빼면 `./` 가 낀 상대경로에서 시트를 못 찾아, 병합이 멀쩡히 있는 파일이
    '복원 불가' 경고를 받는다(정상 파일을 막는 위양성).
    """
    import zipfile

    raw = _merged_two_owner_xlsx()
    with zipfile.ZipFile(io.BytesIO(raw)) as z:
        rels = z.read("xl/_rels/workbook.xml.rels")
    # 전제(주입 성공 확인): 원본은 절대경로 형태다 — 여기가 바뀌면 아래 치환이 헛돈다.
    assert b'Target="/xl/worksheets/sheet1.xml"' in rels, f"원본 Target 형태가 바뀌었다: {rels!r}"

    for label, target in (
        ("절대", b'Target="/xl/worksheets/sheet1.xml"'),
        ("상대", b'Target="worksheets/sheet1.xml"'),
        ("상대(./)", b'Target="./worksheets/sheet1.xml"'),
    ):
        blob = _rewrite_part(
            raw, "xl/_rels/workbook.xml.rels",
            rels.replace(b'Target="/xl/worksheets/sheet1.xml"', target),
        )
        with zipfile.ZipFile(io.BytesIO(blob)) as z:
            assert pes._first_worksheet_part(z) == "xl/worksheets/sheet1.xml", f"{label} 형태 해석 실패"
        assert sorted(pes._merged_ranges_from_zip(blob)) == ["A2:A3", "B2:B3"], f"{label} 형태"


def test_first_sheet_resolution_fails_loudly_on_each_broken_link():
    """해석의 각 고리가 끊겼을 때 **조용히 '병합 없음'이라 답하지 않는다**.

    고리는 셋이다 — ①workbook 의 시트 목록 ②그 시트의 r:id ③rels 의 Target.
    빈 목록을 돌려주면 호출측이 "잃은 값이 없다"로 읽어 경고 없이 넘어가므로 예외를 낸다.

    ★정직하게 적어 두는 한계: 실측 결과 이 셋은 모두 **pandas 도 함께 죽인다**(openpyxl 이 같은
    관계를 타고 시트를 찾기 때문). 따라서 실제 업로드에서는 병합 복원 경로에 닿기 전에 정직한
    error 로 끝난다 — 이 가드들은 **심층 방어**이고, 여기서 잠그는 것은 "빈 목록이 아니라
    예외를 낸다"는 계약이다(그 계약이 깨지면 훗날 도달 가능해졌을 때 침묵이 된다).
    """
    import zipfile

    raw = _merged_two_owner_xlsx()
    with zipfile.ZipFile(io.BytesIO(raw)) as z:
        wb_xml = z.read("xl/workbook.xml")
        rels = z.read("xl/_rels/workbook.xml.rels")

    variants = {
        # ① 시트 목록이 없다
        "시트목록": _rewrite_part(raw, "xl/workbook.xml",
                              re.sub(rb"<sheets>.*?</sheets>", b"", wb_xml, flags=re.S)),
        # ② 첫 시트에 r:id 가 없다 — 가드가 없으면 None 을 그대로 써서 AttributeError 로 샌다
        "r:id": _rewrite_part(raw, "xl/workbook.xml",
                              re.sub(rb'\s\w+:id="[^"]+"', b"", wb_xml, count=1)),
        # ③ 관계는 있는데 Target 이 없다
        "Target": _rewrite_part(raw, "xl/_rels/workbook.xml.rels",
                                re.sub(rb'\sTarget="[^"]+"', b"", rels)),
    }
    # 전제(주입 성공 확인): 셋 다 원본과 실제로 달라야 한다.
    for label, blob in variants.items():
        assert blob != raw, f"{label} 변형이 주입되지 않았다"

    for label, blob in variants.items():
        # ★빈 목록(=잃은 값 없음)이 아니라 예외. 이것이 이 테스트가 잠그는 계약이다.
        with pytest.raises(ValueError):
            pes._merged_ranges_from_zip(blob)
        # ★끝단에서도 침묵이면 안 된다 — 여기서는 표를 못 읽어 정직한 error 로 끝난다.
        out = _parse_x(_break_full_workbook_read(blob), f"{label}끊김.xlsx")
        assert out.get("error"), f"{label}: 아무 말 없이 넘어갔다: {out}"
        assert out.get("parcels") == [], f"{label}: 못 읽었는데 필지가 나왔다: {out.get('parcels')}"





# ── ⑧ CSV cp949 인코딩 폴백 ───────────────────────────────────────────
def test_csv_cp949_encoding_fallback():
    csv_text = ("소재지(주소),지번,PNU(필지고유번호·19자리)\n"
                "서울특별시 동작구 상도동,210-453,1159010300102100453\n")
    raw = csv_text.encode("cp949")
    out = asyncio.run(pes.ParcelExcelService().parse(raw, "cp949.csv", use_llm=False))
    assert not out.get("error")
    assert len(out["parcels"]) == 1
    assert out["parcels"][0]["address"] == "서울특별시 동작구 상도동"


# ── ⑨ 예시행 잔존 — 예시값 미삭제 사고(하드코딩 면적) 재발 방지 확인 ──────
def test_template_example_rows_no_fake_area_injection():
    raw = pes.build_template_xlsx()
    out = asyncio.run(pes.ParcelExcelService().parse(raw, "template.xlsx", use_llm=False))
    assert not out.get("error")
    for p in out["parcels"]:
        assert p.get("area_sqm") not in (14959, 8500), "예시행의 옛 하드코딩 면적값이 재발하면 안 됨"


# ── ⑩ 빈/깨진 파일 — 정직 실패(크래시 없이 error) ────────────────────────
def test_empty_and_corrupt_file_honest_failure():
    out1 = asyncio.run(pes.ParcelExcelService().parse(b"", "empty.xlsx", use_llm=False))
    assert out1.get("error")
    assert out1["parcels"] == []

    out2 = asyncio.run(pes.ParcelExcelService().parse(b"not a real xlsx content", "broken.xlsx", use_llm=False))
    assert out2.get("error")
    assert out2["parcels"] == []


# ── 반복검증(S3): 중복 필지·PNU 형식·합계 불일치 ─────────────────────────
def test_duplicate_pnu_marks_ambiguous_and_needs_review():
    raw = _xlsx([
        ["소재지(주소)", "지번", "PNU(필지고유번호·19자리)"],
        ["서울특별시 동작구 상도동", "210-453", "1159010300102100453"],
        ["경기도 성남시 분당구 정자동", "999-1", "1159010300102100453"],
    ])
    out = asyncio.run(pes.ParcelExcelService().parse(raw, "중복.xlsx", use_llm=False))
    assert not out.get("error")
    assert out["duplicate_pnu_warning"]
    assert {p["status"] for p in out["parcels"]} == {"ambiguous"}
    vr = out["verification_report"]
    assert vr["counts"]["needs_review"] == 2
    # ★H3: injectable=False는 표에서 완전히 제외된 행(합계/집계)에만 쓴다 — ambiguous(needs_review)
    #   도 일단 주입해 주입 후 2차 enrich의 재지오코딩·재검증으로 자기치유되게 한다(과거엔
    #   이 자기치유 경로가 자동반영 제외로 조용히 끊겼었음). 분류·사유는 verification_status/
    #   verification_reasons로 계속 노출된다.
    assert all(p["injectable"] for p in out["parcels"])
    assert all(p["verification_status"] == "needs_review" for p in out["parcels"])


def test_jibun_format_gate_allows_san_and_rejects_garbage():
    assert pes._JIBUN_RE.match("210-453")
    assert pes._JIBUN_RE.match("산12-3")
    assert pes._JIBUN_RE.match("224")
    assert not pes._JIBUN_RE.match("확인필요")
    assert not pes._JIBUN_RE.match("N/A")


def test_jibun_format_issue_triggers_reverify_and_correction(monkeypatch):
    """지번 형식 불량 행 — 같은 행의 다른 원본 셀(비고)에 있는 값을 재질의로 부분문자열 채택."""
    def responder(human: str) -> str:
        if "검증 실패 사유" in human:
            return json.dumps({"jibun": "210-453"}, ensure_ascii=False)
        return "{}"

    _patch_llm(monkeypatch, responder)
    raw = _xlsx([
        ["소재지(주소)", "지번", "비고"],
        ["서울특별시 동작구 상도동", "확인필요", "실제 지번 210-453 로 추정"],
    ])
    out = asyncio.run(pes.ParcelExcelService().parse(raw, "재질의.xlsx", use_llm=True))
    vr = out["verification_report"]
    assert vr["passes"] >= 1
    assert vr["llm_used"] is True
    corr = [c for c in vr["corrections"] if c["field"] == "jibun"]
    assert corr and corr[0]["after"] == "210-453"
    assert out["parcels"][0]["jibun"] == "210-453"


def test_reverify_stops_after_pass1_when_zero_corrections(monkeypatch):
    """H4-①: pass1이 한 건도 교정하지 못하면 pass2를 생략한다(동일 셀 재질의 과금 폭주 방지).
    (구 테스트명 test_reverify_capped_at_two_passes_when_unresolvable — 2회 고정 재시도였던
    구동작을 대체: 이제 무교정이면 1회로 멈춘다.)"""
    call_count = {"n": 0}

    def responder(_human: str) -> str:
        call_count["n"] += 1
        return "{}"  # 원문에 유효 후보가 없다고 가정(교정 불가) — pass1에서 즉시 멈춰야 함

    _patch_llm(monkeypatch, responder)
    raw = _xlsx([
        ["소재지(주소)", "지번"],
        ["서울특별시 동작구 상도동", "확인불가"],
    ])
    out = asyncio.run(pes.ParcelExcelService().parse(raw, "미해결.xlsx", use_llm=True))
    vr = out["verification_report"]
    assert vr["passes"] == 1
    assert call_count["n"] == 1, "pass1 무교정 → LLM 재질의 호출은 pass1의 1건뿐이어야 함(pass2 생략)"
    assert vr["counts"]["needs_review"] == 1


def test_reverify_continues_to_pass2_when_pass1_has_correction(monkeypatch):
    """H4-①의 반대 경계: pass1에서 '어떤 행이든' 교정이 있었으면 아직 미해결 행에 대해
    pass2까지 시도한다(무교정일 때만 조기종료 — 교정 성공 시 조기종료로 다른 행 기회를
    빼앗으면 안 됨)."""
    class _Geo:
        async def geocode_address(self, query):
            if "210-453" in query:
                return {"lat": 37.5, "lon": 127.0, "pnu": "1159010300102100453"}
            return None
        async def get_land_characteristics(self, pnu):
            return {"zone_type": "제2종일반주거지역", "land_category": "대", "official_price_per_sqm": 1_000_000}
        async def search_address(self, query, size=8):
            return []

    import app.services.external_api.vworld_service as vmod
    monkeypatch.setattr(vmod, "VWorldService", _Geo)

    def responder(human: str) -> str:
        if "행1" in human:
            return json.dumps({"jibun": "210-453"}, ensure_ascii=False)
        return "{}"  # 행2는 끝까지 교정 불가

    _patch_llm(monkeypatch, responder)
    raw = _xlsx([
        ["소재지(주소)", "지번", "비고"],
        ["서울특별시 동작구 상도동", "확인필요", "행1 실제 지번 210-453 로 추정"],
        ["경기도 의정부시 의정부동", "확인불가", "행2 단서 없음"],
    ])
    out = asyncio.run(pes.ParcelExcelService().parse(raw, "부분교정.xlsx", use_llm=True))
    vr = out["verification_report"]
    assert vr["passes"] == 2, "행1이 교정·재지오코딩으로 해소됐으므로 행2를 위해 pass2까지 시도해야 함"
    assert out["parcels"][0]["status"] == "ok"
    assert out["parcels"][0]["verification_status"] == "corrected"
    assert out["parcels"][1]["verification_status"] == "needs_review"
    assert vr["counts"] == {"verified": 0, "corrected": 1, "needs_review": 1, "excluded": 0}


# ── use_llm=False 게이트 — 구조가 아무리 나빠도(전치+비표준) LLM 0호출 ────
def test_use_llm_false_zero_llm_calls_even_on_bad_form(monkeypatch):
    monkeypatch.setattr("app.services.ai.llm_provider.get_llm", _fake_llm(_no_llm_reply))
    raw = _xlsx([
        ["항목", "필지1"],
        ["소재지", "서울특별시 동작구 상도동"],
        ["지번", "210-453"],
        ["PNU", "1159010300102100453"],
    ])
    out = asyncio.run(pes.ParcelExcelService().parse(raw, "bad.xlsx", use_llm=False))
    assert out["verification_report"]["llm_used"] is False
    assert out["verification_report"]["passes"] == 0


# ── 환각 차단 ─────────────────────────────────────────────────────────
def test_parse_rejects_hallucinated_sheet_and_column_names(monkeypatch):
    """LLM이 실존하지 않는 시트명·컬럼명을 답해도 채택되지 않는다(정직 실패로 귀결)."""
    calls: list[str] = []

    def responder(human: str) -> str:
        calls.append(human)
        return json.dumps({"sheet_name": "존재안함시트", "columns": {"address": "존재안함컬럼"}},
                           ensure_ascii=False)

    _patch_llm(monkeypatch, responder)
    raw = _xlsx([
        ["X1", "X2", "X3"],
        ["서울특별시 동작구 상도동", "210-453", "200"],
    ])
    out = asyncio.run(pes.ParcelExcelService().parse(raw, "환각.xlsx", use_llm=True))
    assert calls, "LLM은 호출됐어야 함(구조질의 시도)"
    assert out.get("error"), "가짜 시트/컬럼명은 거부되어 필수컬럼 에러로 귀결해야 함"


def test_reverify_hallucination_guard_rejects_non_substring(monkeypatch):
    """LLM이 원본 셀에 없는 값을 답하면(새로 생성) 채택하지 않는다."""
    def responder(_human: str) -> str:
        return json.dumps({"jibun": "완전조작된값"}, ensure_ascii=False)

    _patch_llm(monkeypatch, responder)
    cand, hit = asyncio.run(
        pes._llm_reverify_row({"지번": "확인필요", "비고": "실제지번 아님"}, ["jibun_format"])
    )
    assert hit is True
    assert cand == {}, "원본 셀에 실존하지 않는 값은 채택되지 않아야 함(환각 차단)"


def test_reverify_hallucination_guard_accepts_valid_substring(monkeypatch):
    def responder(_human: str) -> str:
        return json.dumps({"jibun": "224-1"}, ensure_ascii=False)

    _patch_llm(monkeypatch, responder)
    cand, hit = asyncio.run(
        pes._llm_reverify_row({"지번": "확인필요", "비고": "실제 지번 224-1"}, ["jibun_format"])
    )
    assert hit is True
    assert cand == {"jibun": "224-1"}


# ── M1: 재질의 환각가드 강화(셀단위 검사 + 역할별 형식 게이트) ────────────
def test_reverify_area_rejects_pnu_fragment_via_haystack(monkeypatch):
    """M1 재현: 전 셀을 이어붙인 haystack 검사면 PNU 숫자파편("10300")이 area로 오채택될 수
    있다 — 셀 단위 검사(그 파편이 속한 셀 전체가 PNU이지 면적이 아님)로 거부해야 한다."""
    def responder(_human: str) -> str:
        return json.dumps({"area": "10300"}, ensure_ascii=False)

    _patch_llm(monkeypatch, responder)
    cand, hit = asyncio.run(
        pes._llm_reverify_row({"PNU": "9910300000", "비고": "확인 필요"}, ["area_format"])
    )
    assert hit is True
    assert cand == {}, "PNU 숫자파편이 area로 채택되면 안 됨(출처 셀 전체가 PNU이지 면적이 아님)"


def test_reverify_area_accepts_pure_numeric_or_unit_cell(monkeypatch):
    """M1: 정상 면적 후보 셀(순수 숫자, 또는 ㎡ 단위 표기 셀)의 값은 area로 채택된다."""
    def responder(_human: str) -> str:
        return json.dumps({"area": "200"}, ensure_ascii=False)

    _patch_llm(monkeypatch, responder)
    cand, hit = asyncio.run(
        pes._llm_reverify_row({"면적": "200㎡", "비고": "확인 필요"}, ["area_format"])
    )
    assert hit is True
    assert cand == {"area": "200"}


# ── M2: 교정 후 재지오코딩 조건(need_geocode 상태 누락 수정) ─────────────
def test_reverify_correction_retriggers_geocode_from_need_geocode_status(monkeypatch):
    """M2 재현: 기존엔 status in (failed, ambiguous)만 재지오코딩 트리거로 체크해 need_geocode
    상태(주소는 있으나 PNU 미확보)인 행의 지번 교정이 재지오코딩으로 이어지지 못하고 사장됐다.
    _UNRESOLVED_STATUSES(need_geocode 포함)로 통일해 교정 후 자기치유(PNU 확보)까지 이어진다."""
    class _Geo:
        async def geocode_address(self, query):
            if "210-453" in query:
                return {"lat": 37.5, "lon": 127.0, "pnu": "1159010300102100453"}
            return None
        async def get_land_characteristics(self, pnu):
            return {"zone_type": "제2종일반주거지역", "land_category": "대", "official_price_per_sqm": 1_000_000}
        async def search_address(self, query, size=8):
            return []

    import app.services.external_api.vworld_service as vmod
    monkeypatch.setattr(vmod, "VWorldService", _Geo)

    def responder(human: str) -> str:
        if "검증 실패 사유" in human:
            return json.dumps({"jibun": "210-453"}, ensure_ascii=False)
        return "{}"

    _patch_llm(monkeypatch, responder)
    raw = _xlsx([
        ["소재지(주소)", "지번", "비고"],
        ["서울특별시 동작구 상도동", "확인필요", "실제 지번 210-453 로 추정"],
    ])
    out = asyncio.run(pes.ParcelExcelService().parse(raw, "재지오코딩.xlsx", use_llm=True))
    assert out["parcels"][0]["jibun"] == "210-453"
    assert out["parcels"][0]["status"] == "ok", "교정 후 재지오코딩으로 PNU 확보(자기치유)돼야 함"
    assert out["parcels"][0]["pnu"] == "1159010300102100453"


# ── M4: 누적합 합계행 오제외 경계(bcode 있는 정상 행 보호) ────────────────
def test_cumulative_area_summary_gate_respects_bcode_presence():
    """M4 재현: 누적합 분기는 키워드 분기와 달리 not bcode 조건이 없어, bcode가 있는 정상
    행(면적이 우연히 이전 누적합과 일치)이 집계행으로 오제외될 수 있었다."""
    raw = _xlsx([
        ["소재지(주소)", "지번", "법정동코드(bcode·10자리)", "면적(㎡)"],
        ["서울특별시 동작구 상도동", "210-453", "1159010300", "100"],
        ["경기도 의정부시 의정부동", "224-1", "4115010100", "50"],
        # 지번은 비어있지만 bcode가 있는 정상 필지 — 면적이 우연히 위 두 행의 합(150)과 일치.
        ["부산광역시 해운대구 우동", "", "2635010100", "150"],
    ])
    out = asyncio.run(pes.ParcelExcelService().parse(raw, "경계.xlsx", use_llm=False))
    assert not out.get("error")
    assert len(out["parcels"]) == 3, "bcode 있는 정상 행이 누적합 오탐으로 제외되면 안 됨"
    assert out["verification_report"]["counts"]["excluded"] == 0


def test_summary_row_exclusion_is_not_silent():
    """M4: 제외된 행의 원문 요약이 warnings에 남아야 한다(무음 제외 금지)."""
    raw = _xlsx([
        ["소재지(주소)", "지번", "PNU(필지고유번호·19자리)", "면적(㎡)"],
        ["서울특별시 동작구 상도동", "210-453", "1159010300102100453", "200"],
        ["합계", "", "", "200"],
    ])
    out = asyncio.run(pes.ParcelExcelService().parse(raw, "무음제외.xlsx", use_llm=False))
    assert not out.get("error")
    warnings = out["verification_report"]["warnings"]
    assert any("집계/합계 추정 행 제외" in w and "합계" in w for w in warnings), (
        "제외된 행의 원문이 warnings에 표면화돼야 함(무음 금지)"
    )


# ── L5: co_owner(공유지분 연속행)는 'corrected'가 아니라 'verified' ───────
def test_co_owner_rows_classified_verified_not_corrected():
    """L5: 병합복원 등으로 co_owner=True 표시된 공유지분 연속행은 실제 값 보정이 아니므로
    'corrected'가 아니라 'verified'로 분류하고, 사유에 '공유지분 연속행'을 명시한다."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "토지조서"
    ws.append(["소재지(주소)", "지번", "PNU(필지고유번호·19자리)", "소유구분"])
    ws.append(["서울특별시 동작구 상도동", "210-453", "1159010300102100453", "김철수"])
    ws.append(["", "", "", "이영희"])
    ws.merge_cells("A2:A3")
    ws.merge_cells("B2:B3")
    ws.merge_cells("C2:C3")
    buf = io.BytesIO()
    wb.save(buf)
    raw = buf.getvalue()

    out = asyncio.run(pes.ParcelExcelService().parse(raw, "공유지분.xlsx", use_llm=False))
    assert not out.get("error")
    assert len(out["parcels"]) == 2
    for p in out["parcels"]:
        assert p["verification_status"] == "verified"
        assert "공유지분 연속행" in p["verification_reasons"]
        assert p["injectable"] is True


# ── L6: _reverify_loop gather 예외 로그(무음 금지) ────────────────────────
def test_reverify_loop_logs_gather_exceptions_not_silent(monkeypatch):
    """L6: asyncio.gather(return_exceptions=True)로 삼켜지던 개별 행 처리 예외가 로그로
    남아야 한다(무음 금지). 예외가 나도 parse()는 크래시 없이 정상 반환해야 한다."""
    logged: list[tuple] = []

    class _FakeLogger:
        def warning(self, *a, **k):
            logged.append((a, k))

        def info(self, *a, **k):
            pass

    monkeypatch.setattr(pes, "logger", _FakeLogger())

    async def _boom(*_a, **_k):
        raise RuntimeError("boom")

    monkeypatch.setattr(pes, "_llm_reverify_row", _boom)
    raw = _xlsx([
        ["소재지(주소)", "지번"],
        ["서울특별시 동작구 상도동", "확인불가"],
    ])
    out = asyncio.run(pes.ParcelExcelService().parse(raw, "예외.xlsx", use_llm=True))
    assert not out.get("error")
    assert logged, "gather에서 삼켜지던 개별 행 처리 예외가 로그로 남아야 함(무음 금지)"


# ── C1: _STRUCT_CACHE에 행범위가 남아 재업로드 절단 ───────────────────────
def test_struct_cache_excludes_data_range_no_truncation_on_reupload(monkeypatch):
    """C1 재현: 캐시 키(시트목록+현재시트+헤더)에는 행수가 없는데 캐시값에 data_start/end_row가
    남아 있으면, 같은 양식을 행 늘려 재업로드할 때 이전 행범위로 절단된다(1행 업로드 후 4행
    재업로드 → 1필지로 절단·LLM 0호출). 캐시엔 구조속성만 남아야 재업로드 시 전체 행이 산다."""
    call_count = {"n": 0}

    def responder(_human: str) -> str:
        call_count["n"] += 1
        # 1행짜리 업로드 시점 기준 데이터범위(0-based, 헤더행 다음 1개행=인덱스1)를 그대로 답한다.
        return json.dumps({
            "columns": {"address": "A열", "jibun": "B열", "pnu": "C열"},
            "data_start_row": 1, "data_end_row": 1,
        }, ensure_ascii=False)

    _patch_llm(monkeypatch, responder)
    svc = pes.ParcelExcelService()

    raw1 = _xlsx([
        ["A열", "B열", "C열"],
        ["서울특별시 동작구 상도동", "210-453", "1159010300102100453"],
    ])
    out1 = asyncio.run(svc.parse(raw1, "비표준1.xlsx", use_llm=True))
    assert not out1.get("error")
    assert len(out1["parcels"]) == 1
    assert call_count["n"] == 1

    raw4 = _xlsx([
        ["A열", "B열", "C열"],
        ["서울특별시 동작구 상도동", "210-453", "1159010300102100453"],
        ["경기도 의정부시 의정부동 224", "224-1", "4115010100201100224"],
        ["부산광역시 해운대구 우동 500", "500", "2635010100105000000"],
        ["대구광역시 수성구 범어동", "1-1", "2726010600100010000"],
    ])
    out2 = asyncio.run(svc.parse(raw4, "비표준4.xlsx", use_llm=True))
    assert not out2.get("error")
    assert len(out2["parcels"]) == 4, "캐시 히트로 이전 1행 범위에 절단되면 안 됨"
    assert call_count["n"] == 1, "구조질의 캐시 히트 — LLM 재호출 없이도 전체 행이 반영돼야 함"
    assert out2["verification_report"]["llm_used"] is True, "캐시 히트라도 LLM 유래 구조 적용은 llm_used=True(M3)"


# ── H1: data_start/end_row off-by-one(마지막 행 상시 절단) ───────────────
def test_data_range_slice_is_0_based_not_off_by_one(monkeypatch):
    """H1 재현: 프롬프트는 0-based 미리보기 행을 지시하는데 적용부가 base=hdr+2(1-based)로
    매핑해 마지막 데이터 행이 상시 절단됐다(3행+end=3행(0-based 마지막)→2필지로 절단).
    base=hdr+1(0-based)로 통일하면 3행 모두 온전히 남아야 한다."""
    def responder(_human: str) -> str:
        return json.dumps({
            "columns": {"address": "A열", "jibun": "B열", "pnu": "C열"},
            # 헤더가 0행, 데이터는 1~3행(0-based, 미리보기 grid 좌표) — 3행 전부 포함하려는 의도.
            "data_start_row": 1, "data_end_row": 3,
        }, ensure_ascii=False)

    _patch_llm(monkeypatch, responder)
    raw = _xlsx([
        ["A열", "B열", "C열"],
        ["서울특별시 동작구 상도동", "210-453", "1159010300102100453"],
        ["경기도 의정부시 의정부동 224", "224-1", "4115010100201100224"],
        ["부산광역시 해운대구 우동", "500", "2635010100105000000"],
    ])
    out = asyncio.run(pes.ParcelExcelService().parse(raw, "off-by-one.xlsx", use_llm=True))
    assert not out.get("error")
    assert len(out["parcels"]) == 3, "0-based 데이터범위 3행이 모두 온전히 남아야 함(off-by-one 회귀)"


def test_data_range_slice_truncation_is_reported_in_warnings(monkeypatch):
    """H1: 슬라이스로 실제 행이 줄면 warnings에 '구조인식으로 N행 제외'가 표면화돼야 한다."""
    def responder(_human: str) -> str:
        return json.dumps({
            "columns": {"address": "A열", "jibun": "B열", "pnu": "C열"},
            "data_start_row": 1, "data_end_row": 2,  # 3번째 데이터 행(인덱스3)은 범위 밖 — 실제 제외.
        }, ensure_ascii=False)

    _patch_llm(monkeypatch, responder)
    raw = _xlsx([
        ["A열", "B열", "C열"],
        ["서울특별시 동작구 상도동", "210-453", "1159010300102100453"],
        ["경기도 의정부시 의정부동 224", "224-1", "4115010100201100224"],
        ["부산광역시 해운대구 우동", "500", "2635010100105000000"],
    ])
    out = asyncio.run(pes.ParcelExcelService().parse(raw, "슬라이스경고.xlsx", use_llm=True))
    assert not out.get("error")
    assert len(out["parcels"]) == 2
    warnings = out["verification_report"]["warnings"]
    assert any("구조인식으로" in w and "제외" in w for w in warnings), "무음 절단 금지 — 제외 사유가 표면화돼야 함"


# ── H2: LLM 컬럼매핑 경로에서도 병합셀 forward-fill이 유지돼야 함(main 회귀) ─
def test_llm_column_mapping_preserves_merged_cell_expansion(monkeypatch):
    """H2 재현: struct가 truthy(컬럼역할만 응답)면 시트/전치 불변이어도 무조건 rebuild해
    병합셀 forward-fill이 적용된 df를 버리고 원본(미병합) df0에서 다시 만들었다 — 공유지분
    병합이 1필지로 소실(main은 2필지). 시트/전치 변경이 없으면 rebuild를 생략해야 한다.

    ★pd.read_excel(header=None)은 전 셀이 빈 '완전 공백 행'은 통째로 드롭한다(병합 자식
    셀은 None) — 그래서 기존 test_merged_cells_forward_fill처럼 병합 안 되는 컬럼(소유구분)에
    행마다 다른 값을 둬서 pandas가 그 행을 드롭하지 못하게 한다.
    """
    def responder(_human: str) -> str:
        return json.dumps({"columns": {"address": "A열", "jibun": "B열", "pnu": "C열"}},
                           ensure_ascii=False)

    _patch_llm(monkeypatch, responder)
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "토지조서"
    ws.append(["A열", "B열", "C열", "소유구분"])
    ws.append(["서울특별시 동작구 상도동", "210-453", "1159010300102100453", "김철수"])
    ws.append(["", "", "", "이영희"])
    ws.merge_cells("A2:A3")
    ws.merge_cells("B2:B3")
    ws.merge_cells("C2:C3")
    buf = io.BytesIO()
    wb.save(buf)
    raw = buf.getvalue()

    out = asyncio.run(pes.ParcelExcelService().parse(raw, "비표준병합.xlsx", use_llm=True))
    assert not out.get("error")
    assert out["column_engine"] == "rule+llm"
    assert len(out["parcels"]) == 2, "LLM 컬럼매핑 경유에도(시트/전치 불변) 병합 복원 2행이 유지돼야 함"
    assert all(p["jibun"] == "210-453" for p in out["parcels"])
