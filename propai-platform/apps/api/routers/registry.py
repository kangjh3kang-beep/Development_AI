"""부동산 등기부(소유관계) 라우터 — 단건/다필지 일괄 조회·다운로드 + 토지조서."""

import io
import logging
import uuid
from typing import Any

logger = logging.getLogger(__name__)

from fastapi import APIRouter, Depends, File, HTTPException, Request, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

from app.core.charge_guard import charge_once
from app.services.common.job_store import JobStore
from app.services.registry.registry_service import RegistryService
from apps.api.auth.jwt_handler import CurrentUser, get_current_user
from apps.api.rate_limit import ai_limiter, limiter

router = APIRouter(prefix="/registry", tags=["부동산 등기부"])

# ── 비동기 등기분석 작업 저장소(모바일 안정: 긴 동기요청 대신 제출+폴링) ──
# ★공용 잡 스토어(Redis 우선·인메모리 폴백) — design_audit과 동일 계약. 인메모리 백킹은 기존
#   _JOBS 전역 dict를 재사용해 폴백 경로 동작을 보존한다(무악화). Redis 설정 시 블루그린 컷오버·
#   다중 워커에서도 잡 공유(프로세스 경계 404 봉합).
_JOBS: dict[str, dict[str, Any]] = {}
_JOB_TTL = 3600  # 잡 보관 TTL(초)
_REGISTRY_STORE = JobStore("job:registry:", memory_backing=_JOBS, default_ttl_s=_JOB_TTL)


async def _set_registry_job(job_id: str, **fields: Any) -> None:
    """job_id 항목 병합 갱신(user_id 등 소유 필드 보존) — 공용 스토어 경유(get→merge→put).

    ★status 전이 시 소유 필드 소실 버그 봉합: 과거 _run_registry_job이 잡 엔트리를 통째로
    교체(replace)해 user_id를 떨어뜨려, 완료된 잡을 소유자가 폴링하면 소유권 불일치로 404가
    나던 결함이 있었다. 병합으로 소유 필드를 보존한다.
    """
    cur = dict(await _REGISTRY_STORE.get(job_id) or {})
    cur.update(fields)
    await _REGISTRY_STORE.put(job_id, cur, _JOB_TTL)


# 실제 발급/열람이 일어났다고 볼 수 있는 출처.
_ISSUED_ORIGINS = ("hyphen", "tilko", "custom")


def issued_count(result: Any) -> int:
    """**실제로 발급·열람된 건수**를 센다. 과금은 딱 이 수만큼만 한다.

    ★2026-08-12 — 블랙리스트를 **화이트리스트로 뒤집었다**. 종전 `_issue_failed` 는
    `("unavailable","error","failed")` 만 실패로 봤다. 그런데 등기 조회의 실제 실패 상태는
    `not_configured`·`provider_error`·`no_match`·`bad_request`·`forbidden` 이라 **하나도
    걸리지 않았고**, 실패한 조회마다 1,200원이 청구됐다.

    ★추정이 아니라 실측이다: 진단으로 `/registry/bulk` 를 4회 호출해 **전부 실패**했는데
    원장에 `service_fee -1200` 이 **4건** 남았다(23:16~23:18). 합계 4,800원.

    ★블랙리스트가 위험한 이유가 여기 있다 — **새 실패 상태를 추가하는 사람이 돈 가드를
    같이 고쳐야 한다는 것을 모른다**(이 PR 이 `provider_error` 를 추가하며 실제로 놓쳤다).
    화이트리스트는 "성공을 증명하지 못하면 과금하지 않는다" 로 기본값이 안전하다.

    ★★2026-08-15 — **같은 결함 클래스가 두 번째로 났다: 캐시 적중에 발급료를 재청구했다.**
    `RegistryAnalysisService.analyze()` 는 캐시 적중 시 `{**cached, "cached": True}` 를
    돌려주는데 `status` 는 `"ok"` 이고 `origin` 도 원본 그대로다. `analysis_charged` 는
    `cached` 를 보고 건너뛰었지만 **이 함수는 보지 않았다** — 실측:
    `issued_count({"status":"ok","origin":"hyphen","cached":True})` 가 **1** 을 냈다.

    결과: `/registry/survey/strategy` 를 같은 20필지로 두 번 부르면 **외부 발급이 0건인데**
    두 번째 호출에 24,000원이 청구된다. 캐시는 DB 영속·세션 공유라 다른 세션에서도 터진다.
    호출부가 넷(단건·bulk 2곳·전략)이라 **판정 함수 한 곳**에서 막는다 — 라우트마다
    `cached` 를 따로 보게 하면 새 호출부가 또 빠뜨린다(2026-08-12 블랙리스트와 같은 구조).

    - `results` 를 가진 일괄 응답은 **성공 건수만** 센다(요청 필지 수가 아니라).
    - PDF 업로드 파싱은 외부 발급이 아니므로 과금 대상이 아니다.
    """
    if not isinstance(result, dict):
        return 0
    # ★캐시 적중 = 외부 발급이 일어나지 않았다 → 0건. `analysis_charged` 와 **같은 자리·같은
    #   방식**으로 본다(두 판정이 갈라지면 한쪽만 고치는 일이 반복된다).
    if result.get("cached"):
        return 0
    if isinstance(result.get("results"), list):
        return sum(issued_count(r) for r in result["results"])
    if result.get("error"):
        return 0
    status = str(result.get("status", "")).lower()
    # ★`ok is True and not status` 는 **방어적 절이다** — 2026-08-12 기준 tilko·hyphen·
    #   pdf 파서·registry_service 의 모든 return 을 전수로 확인한 결과, `ok:True` 를 내면서
    #   `status` 가 없는 shape 은 **하나도 없다**. 지금은 도달 불가라 변이가 생존한다
    #   (그 사실을 적어 둔다 — 변이 점수 부풀리기 방지). 신규 프로바이더 대비로 남긴다.
    ok = status == "ok" or (result.get("ok") is True and not status)
    if not ok:
        return 0
    # ★이중 가드다 — 단건 라우트가 이미 `if not pdf_input:` 으로 과금을 건너뛰고,
    #   `bulk` 은 PDF 를 넘기지 않는다. 한쪽만 죽어도 동작이 옳아 변이가 생존한다.
    #   그래도 남기는 이유: 판정 함수 자체가 "업로드 파싱은 발급이 아니다" 를 말해야
    #   새 호출부가 생겼을 때 라우트 가드를 잊어도 안전하다.
    if result.get("origin") == "pdf_upload":
        return 0
    # 발급 근거: 알려진 프로바이더가 처리했거나, 실제 문서(PDF)를 받았다.
    has_doc = bool(result.get("has_pdf") or result.get("pdf_data") or result.get("pdf_base64"))
    return 1 if (result.get("origin") in _ISSUED_ORIGINS or has_doc) else 0


async def _charge_registry_issue(user_id: Any, result: Any, times: int | None = None) -> None:
    """등기부등본 발급·열람 사용료 누적(best-effort). **발급된 건수만** 과금한다.

    `times` 는 하위호환용 상한이다 — 실제 과금은 `issued_count` 가 센 수를 넘지 않는다.
    (종전에는 `times=len(items)` 로 **요청 수만큼** 과금해, 10필지 중 1건만 발급돼도
    10건이 청구될 수 있었다.)
    """
    n = issued_count(result)
    # ★이 상한은 지금 **한 번도 구속하지 않는다**(변이로 무력화해도 결과가 같다).
    #   `issued_count` 가 이미 **실제 발급 수**를 세므로, 모든 호출부에서 `times >= n` 이다
    #   (단건 `times=1`·일괄 `times=len(items)`). 하위호환·신규 호출부 대비로만 남긴다.
    if times is not None:
        n = min(n, max(0, times))
    # ★이중 가드 — 아래 `range(n)` 이 n=0 이면 어차피 한 번도 돌지 않는다.
    #   조기 반환은 DB 세션을 열지 않기 위한 것이다(실패 조회가 대부분인 상황에서 유의미).
    if n <= 0:
        return
    try:
        from app.core.database import async_session_factory
        from app.services.billing import billing_service

        async with async_session_factory() as _db:
            for _ in range(n):
                await billing_service.charge_service(_db, user_id, "registry_issue")
    except Exception:  # noqa: BLE001
        pass


def analysis_charged(result: Any) -> bool:
    """권리분석을 **실제로 수행했는가**. 과금은 이 판정이 참일 때만 한다.

    ★`issued_count` 와 **대칭**이다. 발급 쪽만 화이트리스트로 뒤집고 분석 쪽을 두면
    같은 결함이 같은 파일 안에 남는다(리뷰가 잡았다 — 30줄 아래에 그대로 있었다).

    분석기가 등기부를 못 받으면 `status="not_available"`·`"empty"` 에 **`ai=None`** 을
    돌려준다(`registry_analysis_service.py:391,403`). 종전 라우트는 결과를 **보지도 않고**
    1,200원을 청구해, **AI 분석이 0인 응답에 돈을 받고 있었다**.

    ★캐시 적중(`cached=True`)도 과금하지 않는다 — 신규 분석이 일어나지 않았다.
    """
    if not isinstance(result, dict):
        return False
    if result.get("cached"):
        return False
    return str(result.get("status", "")).lower() == "ok" and bool(result.get("ai"))


async def _charge_registry_analysis(user_id: Any, result: Any) -> Any:
    """등기 권리분석 사용료(1,200원) — **분석이 실제로 나온 경우만**(best-effort)."""
    if not analysis_charged(result):
        return None
    try:
        from app.core.database import async_session_factory
        from app.services.billing import billing_service

        async with async_session_factory() as _db:
            return await billing_service.charge_service(_db, user_id, "registry_analysis")
    except Exception:  # noqa: BLE001
        return None


async def _run_registry_job(job_id: str, params: dict[str, Any]) -> None:
    try:
        from app.services.registry.registry_analysis_service import RegistryAnalysisService

        res = await RegistryAnalysisService().analyze(**params)
        await _set_registry_job(job_id, status="done", result=res)
        # ★과금을 **제출 시점에서 완료 시점으로** 옮겼다. 종전에는 작업을 시작하기도 전에
        #   청구하고, 그 작업이 실패해도 환불 경로가 없었다(리뷰 지적).
        #   소유자는 잡 엔트리에 기록돼 있다(IDOR 스코프용) — 그 값을 그대로 쓴다.
        owner = (await _REGISTRY_STORE.get(job_id) or {}).get("user_id")
        if owner:
            await _charge_registry_analysis(owner, res)
    except Exception as e:  # noqa: BLE001
        await _set_registry_job(job_id, status="error", error=str(e)[:200])


# ── 유료·다필지 경로의 **지출/부하 상한** (형제 공용) ────────────────────
# ★★이 상한은 원래 `/survey/strategy` **하나에만** 걸려 있었다. 그게 결함이다 —
#   `/registry/bulk` 는 **더 오래됐고 더 많이 쓰이며 똑같이 유료**인데(아래 `times=len(items)`)
#   상한이 없었다. CLAUDE.md D20: *"처방을 적용한 범위 = 결함이 사는 범위인지 확인하라."*
#   신규 엔드포인트에만 방어를 걸고 **형제를 스윕하지 않은** 전형적 형태였다.
# ★상수를 한 곳에 두는 이유: 값이 갈리면 "어느 게 진짜 상한인가"를 아무도 모르게 된다.
MAX_BULK_ITEMS = 100
MIN_BULK_ITEMS = 1

# 토지조서 엑셀(무과금·메모리 축)은 업로드 관례와 맞춘다 —
# `parcel_excel_service._MAX_ROWS = 500`. 과금 상한과 값이 다른 것은 **의도**다(축이 다르다).
MAX_LAND_SCHEDULE_ROWS = 500


class RegistryBulkRequest(BaseModel):
    """다필지 등기부 일괄 조회. **건당 과금**이라 길이가 곧 청구액이다.

    ★`items`·`addresses` 둘 다 상한을 건다 — `addresses` 는 "단축 입력"이라 같은 경로로
      들어오므로 한쪽만 막으면 우회된다(양방향 경계, D19).
    ★하한도 건다 — 빈 요청이 조용히 200 을 내면 사용자가 "조회됐다"로 오독한다.
    """

    items: list[dict[str, Any]] = Field(
        default_factory=list,
        max_length=MAX_BULK_ITEMS,
        description="[{pnu?, address?}]",
    )
    addresses: list[str] | None = Field(default=None, max_length=MAX_BULK_ITEMS)  # 단축 입력


@router.get("/status", summary="등기부 API 연동 상태")
async def registry_status() -> dict[str, Any]:
    # 키 존재만이 아니라 실제 호출 권한까지 확인한 상태를 돌려준다(공용 판정).
    return await RegistryService().live_status()


@router.get("/tilko/status", summary="틸코(Tilko) 등기 연동 상태 점검")
async def tilko_status() -> dict[str, Any]:
    """TILKO_API_KEY·IROS 자격 설정 여부 + 공개키 도달성 점검(키 입력 후 검증용)."""
    from app.services.registry import tilko_client as tk

    out = {"key_set": tk.tilko_ready(), "iros_set": tk.iros_ready(), "public_key_ok": False}
    if tk.tilko_ready():
        pub = await tk.get_public_key()
        out["public_key_ok"] = bool(pub)
        out["message"] = (
            "틸코 API키 정상(공개키 수신). IROS 자격까지 설정되면 등기 조회 가능."
            if pub else "공개키 조회 실패 — TILKO_API_KEY 확인 필요."
        )
    else:
        out["message"] = "TILKO_API_KEY 미설정(관리자 키화면 입력 필요)."
    return out


@router.post("/tilko/search", summary="틸코 등기물건 주소검색(주소→부동산 고유번호)")
async def tilko_search(
    req: dict[str, Any],
    current_user: CurrentUser = Depends(get_current_user),
) -> dict[str, Any]:
    """주소 → 부동산 고유번호 목록(RISUConfirmSimpleC). IROS 로그인·전자결제 불필요(Tilko API키만)."""
    from app.services.registry import tilko_client as tk
    return await tk.search_unique_no(str(req.get("address") or ""), page=str(req.get("page") or "1"))


@router.post("/tilko/realty", summary="틸코 등기부등본 조회/발급(IROS ID로그인)")
async def tilko_realty(
    request: Request,
    req: dict[str, Any],
    current_user: CurrentUser = Depends(get_current_user),
) -> dict[str, Any]:
    """틸코로 등기부등본 조회/발급. unique_no(고유번호) 또는 address(자동 주소검색)로 부동산 지정.

    ⚠ 발급 수수료가 IROS 전자지급수단에서 차감됨(실호출).
    """
    from app.services.registry import tilko_client as tk

    # unique_no(부동산 고유번호 14자리) = Pin 필드. property_params.Pin/UniqueNo도 허용(하위호환).
    uno = str(req.get("unique_no") or req.get("pin")
              or (req.get("property_params") or {}).get("Pin")
              or (req.get("property_params") or {}).get("UniqueNo") or "").replace("-", "").strip()
    # 고유번호 미지정 + 주소 제공 시 → 주소검색으로 자동 해석.
    # 한 주소에 여러 물건(토지·건물·집합건물 각 호)이 나오므로 공용 선택기로 고른다 —
    # 첫 건을 맹목적으로 집으면 요청과 다른 물건을 1,200원 과금하며 발급하게 된다.
    select_note: str | None = None
    if not uno and req.get("address"):
        from app.services.registry.realty_kind import select_registry_item

        s = await tk.search_unique_no(str(req["address"]))
        items = s.get("items") or []
        picked, select_note = select_registry_item(
            items,
            realty_type=req.get("realty_type"),
            dong=req.get("dong"),
            ho=req.get("ho"),
        )
        if picked and picked.get("unique_no"):
            uno = picked["unique_no"]
        else:
            return {"ok": False, "status": s.get("status", "need_unique_no"),
                    "message": s.get("message") or "주소로 부동산 고유번호를 찾지 못했습니다.",
                    "search": s}
    # ★이 라우트는 **파생형 락이 잡아냈다.** 나는 과금 경로를 손으로 넷(get-one·bulk·analyze·
    #   survey/strategy)만 세고 여기를 빠뜨렸다 — 목록형으로 열거했으면 영영 안 걸렸을 자리다.
    #   틸코는 IROS 전자지급수단에서 **실제로 수수료가 차감**되므로 재전송 노출이 형제들과 같다.
    async with charge_once(
        request, endpoint="registry.tilko_realty", payload=req,
        tenant_id=current_user.tenant_id, user_id=current_user.user_id,
    ) as guard:
        result = await tk.fetch_realty_registry(
            unique_no=uno,
            cmort_flag=str(req.get("cmort_flag", "N")),
            trade_seq_flag=str(req.get("trade_seq_flag", "N")),
            abs_cls=str(req.get("abs_cls", "11")),
            rgs_mttr_smry=str(req.get("rgs_mttr_smry", "")),
        )
        # 등기부등본 발급·열람 1건 1,200원(발급 성공 시, best-effort).
        if guard.billable:
            await _charge_registry_issue(current_user.user_id, result, times=1)
        if select_note and isinstance(result, dict):
            # 요청과 다른 물건을 골랐을 수 있다는 사실을 과금 결과와 함께 반드시 전달한다.
            result["select_note"] = select_note
        return result


@router.post("/get-one", summary="단건 등기부 조회 / 비상 PDF 업로드 파싱")
@limiter.limit(ai_limiter)
async def registry_get_one(
    request: Request,
    req: dict[str, Any],
    current_user: CurrentUser = Depends(get_current_user),
) -> dict[str, Any]:
    """단건 등기부 조회(주소·PNU·고유번호) 또는 **비상 등기부 PDF 업로드 파싱**.

    ★이 라우트는 **없었다**(2026-08-12 라이브 진단으로 발각). 프론트
    `RegistryUploadModal` 은 `POST /registry/get-one` 을 부르는데(2026-07-24 `588ea8ed`)
    백엔드에 그 경로가 **한 번도 존재한 적이 없어** 약 3주간 상시 404 였다.
    서비스 함수 `RegistryService.get_one(pdf_input=...)` 은 PDF 파서 분기까지 구현돼
    있었으므로 **문만 없던 셈**이다.

    ★왜 이게 특히 나빴나: 조회 실패 응답이 사용자에게 "'비상 등기부 PDF 직접 업로드'
    기능을 이용하세요" 라고 안내한다. 즉 **주 경로가 막혔을 때의 탈출구가 404** 였다.

    과금: PDF 업로드 파싱은 **발급이 아니므로 과금하지 않는다**(외부 발급 없음).
    주소·고유번호로 실제 발급이 일어난 경우에만 건당 1,200원(발급 성공 시).
    """
    pdf_input = req.get("pdf_input")
    # ★`req` 는 pydantic 모델이 아니라 dict 다 — `req.model_dump()` 를 부르면 즉시 500 이다.
    #   `compute_request_hash` 는 dict 를 그대로 받는다.
    async with charge_once(
        request, endpoint="registry.get_one", payload=req,
        tenant_id=current_user.tenant_id, user_id=current_user.user_id,
    ) as guard:
        result = await RegistryService().get_one(
            pnu=req.get("pnu"),
            address=req.get("address"),
            unique_no=req.get("unique_no") or req.get("pin"),
            pdf_input=pdf_input,
            realty_type=req.get("realty_type"),
            dong=req.get("dong"),
            ho=req.get("ho"),
        )
        if not pdf_input and guard.billable:
            await _charge_registry_issue(current_user.user_id, result, times=1)
        return result


@router.post("/bulk", summary="다필지 등기부 일괄 조회/다운로드")
@limiter.limit(ai_limiter)
async def registry_bulk(
    request: Request,
    req: RegistryBulkRequest,
    current_user: CurrentUser = Depends(get_current_user),
) -> dict[str, Any]:
    """여러 필지의 등기부를 일괄 조회/발급한다(공급자 키 설정 시). 미설정 시 안내 반환."""
    items = list(req.items or [])
    if not items and req.addresses:
        items = [{"address": a} for a in req.addresses if a and a.strip()]
    # ★이 경로가 이중청구 노출이 가장 크다 — 100필지 × 1,200원이 한 번에 나간다.
    #   프론트 타임아웃(120초)이 서버 실행보다 먼저 끊기면 사용자는 "실패"를 보고 다시 누른다.
    #   그때 두 요청이 서버에서 겹친다 → 선점이 없으면 둘 다 과금된다.
    async with charge_once(
        request, endpoint="registry.bulk", payload={"items": items},
        tenant_id=current_user.tenant_id, user_id=current_user.user_id,
    ) as guard:
        result = await RegistryService().bulk(items)
        # 발급·열람 1건당 1,200원 × 필지수(발급 성공 시, best-effort).
        if guard.billable:
            await _charge_registry_issue(current_user.user_id, result, times=max(1, len(items)))
        return result


class RegistryAnalyzeRequest(BaseModel):
    address: str | None = None
    pnu: str | None = None
    registry_text: str | None = None  # 등기부등본 내용 직접 입력(연동 미설정 시)
    realty_type: str | None = None    # 0토지+건물 1집합건물 2토지 3건물(기본=env)
    dong: str | None = None           # 집합건물 동
    ho: str | None = None             # 집합건물 호
    # 부지분석에서 이미 확보한 토지정보(전달 시 백엔드 재조회 생략 → 지연 단축)
    land_hint: dict[str, Any] | None = None
    # ★이미 발급받은 등기부가 있으면 재사용하는 것이 기본이다(발급은 민원캐시를 차감한다).
    #   True 면 캐시를 건너뛰고 **새로 발급**한다 — 돈이 드는 행위라 호출측이 명시할 때만.
    force_reissue: bool = False


@router.post("/analyze", summary="부동산 등기정보 권리분석(법무사·변호사 AI)")
@limiter.limit(ai_limiter)
async def registry_analyze(
    request: Request,
    req: RegistryAnalyzeRequest,
    current_user: CurrentUser = Depends(get_current_user),
) -> dict[str, Any]:
    """등기부(연동 조회 또는 직접 입력)를 법무사·변호사 관점에서 분석해 소유정보·소유기간·
    매입금액·보유지분·가등기·압류·근저당·매도청구 가능여부 등 권리관계를 제공한다.
    집합건물은 realty_type=1 + dong/ho로 특정 호 등기를 조회한다."""
    from app.services.registry.registry_analysis_service import RegistryAnalysisService

    async with charge_once(
        request, endpoint="registry.analyze", payload=req,
        tenant_id=current_user.tenant_id, user_id=current_user.user_id,
    ) as guard:
        result = await RegistryAnalysisService().analyze(
            address=req.address, pnu=req.pnu, registry_text=req.registry_text,
            realty_type=req.realty_type, dong=req.dong, ho=req.ho,
            land_hint=req.land_hint, force_reissue=req.force_reissue,
        )
        # 서비스 사용료: 등기부등본 권리분석 1건 1,200원(LLM 과금 별개, best-effort).
        # ★분석이 실제로 나온 경우만 청구한다 — 종전에는 결과를 **보지도 않고** 청구해
        #   `ai: null` 인 응답에도 돈을 받았다(`analysis_charged` 주석 참조).
        charge = await _charge_registry_analysis(current_user.user_id, result) if guard.billable else None
        if charge is not None and isinstance(result, dict):
            result["service_charge"] = charge
        return result


@router.post("/analyze/jobs", summary="등기 권리분석 비동기 작업 제출(모바일 안정)")
async def registry_analyze_submit(
    req: RegistryAnalyzeRequest,
    current_user: CurrentUser = Depends(get_current_user),
) -> dict[str, Any]:
    """긴 동기요청(CODEF ~50s) 대신 작업을 제출하고 즉시 job_id 반환.
    캐시 적중 시 즉시 결과 반환(작업 생략). 진행은 GET /analyze/jobs/{id}로 폴링."""
    from app.services.registry.registry_analysis_service import peek_analyze_cache

    # ★재발급을 명시 요청했으면 캐시를 보지 않는다 — 보면 옛 결과가 즉시 반환돼
    #   "새로 발급"이 조용히 무시된다(요청과 결과가 어긋나는 침묵 실패).
    cached = None if req.force_reissue else await peek_analyze_cache(
        address=req.address, pnu=req.pnu, realty_type=req.realty_type,
        dong=req.dong, ho=req.ho, registry_text=req.registry_text,
    )
    if cached is not None:
        # 캐시 적중 = 신규 분석 없음 → 과금 안 함(동일 입력 재조회 무료).
        return {"job_id": None, "status": "done", "result": cached}

    # ★과금은 여기서 하지 않는다 — 작업이 **끝나고** 분석이 실제로 나왔을 때
    #   `_run_registry_job` 이 청구한다. 종전에는 시작 전에 청구하고 실패해도 환불이 없었다.
    job_id = uuid.uuid4().hex
    # ★소유권 기록(IDOR 봉합 — R1 범위외 발견): 등기 권리분석 결과는 개인정보 급이라
    #   제출자만 조회 가능해야 한다. GET 이 이 user_id 로 스코프한다(불일치=404).
    #   프루닝은 스토어가 put 시 lazy 수행(별도 _prune 호출 불필요).
    await _REGISTRY_STORE.put(
        job_id, {"status": "pending", "user_id": str(current_user.user_id)}, _JOB_TTL
    )
    params = dict(
        address=req.address, pnu=req.pnu, registry_text=req.registry_text,
        realty_type=req.realty_type, dong=req.dong, ho=req.ho, land_hint=req.land_hint,
        force_reissue=req.force_reissue,
    )
    # ★태스크 강참조 보관(GC 유실 방지 — design_audit 과 동일 공용 헬퍼).
    from app.services.common.bg_tasks import create_tracked_task

    create_tracked_task(_run_registry_job(job_id, params))
    return {"job_id": job_id, "status": "pending"}


@router.get("/analyze/jobs/{job_id}", summary="등기 권리분석 작업 상태/결과 조회")
async def registry_analyze_status(
    job_id: str,
    current_user: CurrentUser = Depends(get_current_user),
) -> dict[str, Any]:
    """작업 상태(pending/done/error)와 완료 시 결과를 반환."""
    j = await _REGISTRY_STORE.get(job_id)
    # ★소유권 스코프(IDOR 봉합): 타인 job_id 를 추측·탈취해도 404(존재 여부 비노출).
    #   구(user_id 미기록) 잡은 프루닝 TTL 내 잔존 가능 — 소유 확인 불가라 동일하게 404(fail-closed).
    if not j or j.get("user_id") != str(current_user.user_id):
        raise HTTPException(404, "작업을 찾을 수 없습니다(만료되었거나 잘못된 ID).")
    return {"status": j["status"], "result": j.get("result"), "error": j.get("error")}


@router.post("/cleanup", summary="등기부 PDF TTL 자동삭제(경과분 정리)")
async def registry_cleanup(days: int = 30) -> dict[str, Any]:
    """비공개 버킷의 등기부 PDF 중 days 경과분을 삭제한다(워커 cron/수동 호출용)."""
    from apps.api.services.storage_service import cleanup_registry_pdfs

    try:
        deleted = await cleanup_registry_pdfs(days=days)
        return {"status": "ok", "deleted": deleted, "days": days}
    except Exception as e:  # noqa: BLE001
        return {"status": "error", "message": str(e)[:200]}


# ── 토지조서 엑셀 ──

class LandRow(BaseModel):
    jibun: str = ""
    owner: str = ""
    share: str = ""
    area_sqm: float | None = None  # 면적(집합건물 세대행은 대지지분 면적=실토지 기여분)
    exclusive_area_sqm: float | None = None  # 세대 전유면적(집합건물 세대행)
    unit_label: str = ""  # 동·호(집합건물 세대행)
    owner_type: str = ""
    expected_price: float | None = None
    purchase_price: float | None = None
    contracted: bool = False
    land_use_consent: bool = False
    district_consent: bool = False
    note: str = ""


class LandScheduleExcelRequest(BaseModel):
    """토지조서 엑셀 생성. **무과금**이지만 상한을 건다 — 메모리에 워크북을 통째로 짓는다.

    ★상한값은 업로드 쪽 관례(`parcel_excel_service._MAX_ROWS = 500`)에 맞춘다. 다운로드가
      업로드보다 좁으면 **자기가 받은 조서를 되돌려받지 못하는** 비대칭이 생긴다.
      과금 경로(`MAX_BULK_ITEMS=100`)와 값이 다른 것은 **의도적**이다 — 축이 다르다
      (여기는 지갑이 아니라 메모리).
    """

    project_name: str = "토지조서"
    rows: list[LandRow] = Field(default_factory=list, max_length=MAX_LAND_SCHEDULE_ROWS)


@router.post("/land-schedule/excel", summary="토지조서 엑셀 다운로드")
async def land_schedule_excel(req: LandScheduleExcelRequest):
    """토지조서(편입토지 명세 + 집계)를 엑셀(.xlsx)로 생성한다."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "토지조서"

    title = f"토지조서 — {req.project_name}"
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=14)
    ws["A1"].font = Font(size=14, bold=True)

    # 대지지분(평)·세대면적: 집합건물(공동주택·다세대·집합상가) 세대행에서 채워진다.
    PY = 3.305785  # 1평 = 3.305785㎡
    headers = ["번호", "지번/동·호", "소유자", "대지권비율/지분", "대지지분(㎡)", "대지지분(평)",
               "세대전유면적(㎡)", "소유구분",
               "매입예정가(원)", "매입가(원)", "계약확정", "토지사용동의", "지구단위동의"]
    ws.append(headers)
    hdr_fill = PatternFill("solid", fgColor="0E7490")
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=c)
        cell.fill = hdr_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    tot_area = priv_area = pub_area = excl_area = 0.0
    sum_expected = sum_purchase = 0.0
    contracted_n = use_consent_n = dist_consent_n = 0
    for i, r in enumerate(req.rows, start=1):
        area = r.area_sqm or 0
        tot_area += area
        excl = r.exclusive_area_sqm or 0
        excl_area += excl
        if r.owner_type == "국공유지":
            pub_area += area
        elif r.owner_type == "사유지":
            priv_area += area
        sum_expected += r.expected_price or 0
        sum_purchase += r.purchase_price or 0
        contracted_n += 1 if r.contracted else 0
        use_consent_n += 1 if r.land_use_consent else 0
        dist_consent_n += 1 if r.district_consent else 0
        ws.append([
            i, r.jibun, r.owner, r.share,
            round(area, 2), round(area / PY, 3),  # 대지지분 ㎡ / 평
            round(excl, 2) if excl else "", r.owner_type,
            int(r.expected_price) if r.expected_price else "",
            int(r.purchase_price) if r.purchase_price else "",
            "○" if r.contracted else "", "○" if r.land_use_consent else "",
            "○" if r.district_consent else "",
        ])

    n = len(req.rows)
    pct = lambda a, b: f"{round(a / b * 100, 1)}%" if b else "-"  # noqa: E731
    ws.append([])
    summary = [
        ["총 필지/세대수", f"{n}건"],
        ["대지면적 합계(Σ대지지분=실토지면적)", f"{round(tot_area):,}㎡ ({round(tot_area / PY):,}평)"],
        ["  - 사유지", f"{round(priv_area):,}㎡"],
        ["  - 국공유지", f"{round(pub_area):,}㎡"],
        ["세대 전유면적 합계(집합건물)", f"{round(excl_area):,}㎡ ({round(excl_area / PY):,}평)" if excl_area else "-"],
        ["확보비율(계약확정)", f"{pct(contracted_n, n)} ({contracted_n}/{n})"],
        ["토지사용 동의율", f"{pct(use_consent_n, n)} ({use_consent_n}/{n})"],
        ["지구단위 동의율", f"{pct(dist_consent_n, n)} ({dist_consent_n}/{n})"],
        ["매입예정가 합계", f"{int(sum_expected):,}원"],
        ["매입가 합계(확정)", f"{int(sum_purchase):,}원"],
        ["미확보 잔여 보상비(예정-매입)", f"{int(sum_expected - sum_purchase):,}원"],
    ]
    for label, val in summary:
        ws.append([label, val])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    widths = [6, 28, 14, 14, 12, 11, 14, 10, 16, 16, 9, 12, 12]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="land_schedule.xlsx"'},
    )


@router.post("/land-schedule/import", summary="토지조서 엑셀 업로드(대량 지번 일괄 입력)")
async def land_schedule_import(file: UploadFile = File(...)) -> dict[str, Any]:
    """토지조서 엑셀(.xlsx)을 업로드해 행으로 파싱한다. 헤더에 '지번' 포함 행을 기준으로
    소유자/지분/면적/소유구분/매입예정가/매입가/계약/동의 컬럼을 유연 매핑한다."""
    from openpyxl import load_workbook

    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="빈 파일입니다.")

    # ★공용 콘텐츠 검증(WP-H 세션2 전역 스윕·fail-closed) — openpyxl 파싱 전에 압축폭탄(xlsx=zip
    # 계열)·실행/스크립트 위장·MIME 위장·경로순회를 차단한다. xlsx 전용이라 실측 계열을 zip 으로
    # 화이트리스트한다(xlsx 는 항상 PK/zip). 검증 실패는 http_status(4xx).
    from app.services.security.content_inspection import http_status_for, inspect_upload

    _verdict = inspect_upload(raw, file.filename or "", file.content_type, expected_kinds={"zip"})
    if not _verdict.allowed:
        raise HTTPException(
            status_code=http_status_for(_verdict.code),
            detail=f"업로드가 거부되었습니다: {_verdict.reason}",
        )
    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001
        return {"status": "error", "message": f"엑셀 읽기 실패: {str(e)[:120]}", "rows": []}
    ws = wb.active

    def _num(v: Any) -> float | None:
        if v is None:
            return None
        try:
            return float(str(v).replace(",", "").replace("원", "").strip())
        except (TypeError, ValueError):
            return None

    def _bool(v: Any) -> bool:
        s = str(v or "").strip().lower()
        return s in ("○", "o", "y", "yes", "true", "1", "예", "완료", "v")

    headers: list[str] = []
    out: list[dict[str, Any]] = []
    # 모든 행을 보관(LLM 폴백용 그리드).
    all_rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        cells = [("" if c is None else str(c)).strip() for c in row]
        all_rows.append(cells)
        # 헤더 탐지: 공백 제거 후 '지번' 매칭(예 '지 번'도 인식).
        if not headers:
            if any("지번" in c.replace(" ", "") for c in cells):
                headers = cells
            continue
        if not any(cells):  # 빈 행 → 데이터 끝(집계 푸터 앞에서 중단)
            break
        rd = {headers[i]: (cells[i] if i < len(cells) else "") for i in range(len(headers))}

        def pick(*keys: str, exclude: tuple[str, ...] = ()) -> str:
            for k, v in rd.items():
                if any(key in k for key in keys) and not any(e in k for e in exclude):
                    return v
            return ""

        jibun = pick("지번", "주소")
        # 집계 푸터 잔재(필지수·면적·비율·금액 등) 방어적 스킵.
        # '평'은 정상 지번(평창동·평택 등)을 오스킵하므로 제외 — ㎡/원/%로 면적·금액 푸터를 잡는다.
        if not jibun or any(t in jibun for t in ("필지", "㎡", "%", "원")):
            continue
        ot = pick("소유구분")
        owner_type = "국공유지" if ("국" in ot or "공" in ot) else ("사유지" if ot else "")
        # 면적: 신규 양식의 '대지지분(㎡)' 우선, 없으면 '면적'(단 '전유'·'평' 컬럼 제외).
        area_val = pick("대지지분", exclude=("평",)) or pick("면적", exclude=("전유", "평"))
        out.append({
            "jibun": jibun,
            "owner": pick("소유자"),
            "share": pick("지분", "대지권비율"),
            "area_sqm": _num(area_val),
            "exclusive_area_sqm": _num(pick("전유")),  # 세대 전유면적(집합건물 세대행)
            "owner_type": owner_type,
            "expected_price": _num(pick("매입예정")),
            "purchase_price": _num(pick("매입가")),
            "contracted": _bool(pick("계약")),
            "land_use_consent": _bool(pick("토지사용")),
            "district_consent": _bool(pick("지구단위")),
        })
    if out:
        return {"status": "ok", "count": len(out), "rows": out, "engine": "rule"}

    # ── LLM 폴백 ── 규칙기반이 0행(병합셀·다층헤더·집계 혼재 등 복잡 레이아웃)일 때
    # LLM이 시트 전체를 읽어 필지/소유자 행을 구조화 추출(원본 최대 복원, 무목업).
    llm_rows = await _llm_extract_land_schedule(all_rows)
    if llm_rows:
        return {"status": "ok", "count": len(llm_rows), "rows": llm_rows, "engine": "llm"}
    return {
        "status": "ok", "count": 0, "rows": [],
        "message": "지번을 인식하지 못했습니다. '지번/소재지'·소유자·면적이 포함된 토지조서인지 확인하세요.",
    }


async def _llm_extract_land_schedule(all_rows: list[list[str]]) -> list[dict[str, Any]]:
    """복잡 레이아웃 토지조서를 LLM으로 구조화 추출(병합셀 지번 상속·집계행 제외)."""
    # 그리드 텍스트화: 앞 90행, 셀 길이 제한, 빈 trailing 컬럼 제거.
    lines: list[str] = []
    for i, cells in enumerate(all_rows[:90]):
        trimmed = [c[:40] for c in cells]
        while trimmed and not trimmed[-1]:
            trimmed.pop()
        if trimmed:
            lines.append(f"R{i + 1}: " + " | ".join(trimmed))
    grid = "\n".join(lines)
    if not grid.strip():
        return []
    try:
        from langchain_core.messages import HumanMessage, SystemMessage

        from app.services.ai.llm_provider import get_llm

        llm = get_llm(timeout=60, max_tokens=4000)
        sys = (
            "너는 한국 부동산 토지조서(편입토지조서) 엑셀을 정확히 구조화하는 전문가다. "
            "병합셀·다층헤더·집계행을 이해하고 각 소유자/필지 행을 추출한다. 근거 없는 값은 비운다."
        )
        human = (
            "다음은 토지조서 엑셀 셀 내용이다(R행번호: 열1 | 열2 ...). 각 데이터 행을 JSON 배열로만 "
            "출력하라(설명·코드펜스 금지).\n"
            "스키마: [{\"jibun\":\"소재지+지번 예 '사당동 219-16'\",\"owner\":\"소유자명\","
            "\"share\":\"지분 예 '1/2', 없으면 빈문자\",\"area_sqm\":편입면적㎡_숫자_또는_null,"
            "\"owner_type\":\"사유지|국공유지|빈문자\"}]\n"
            "규칙: ①병합셀로 지번이 빈 행은 바로 위 유효 지번을 상속 ②합계/소계/구성비/집계 행 제외 "
            "③헤더 행 제외 ④면적은 편입면적 우선(없으면 지적면적), 숫자만 ⑤JSON 배열만 출력.\n\n"
            f"[엑셀]\n{grid}"
        )
        resp = await llm.ainvoke(
            [SystemMessage(content=sys), HumanMessage(content=human)]
        )
        # 계측: BaseInterpreter 밖 직접 호출도 동일하게 토큰·과금 기록(best-effort)
        from app.services.ai.base_interpreter import record_llm_response_billing
        await record_llm_response_billing(llm, resp, service="registry")
        text = resp.content if isinstance(resp.content, str) else str(resp.content)
        import json
        import re

        m = re.search(r"\[.*\]", text, re.S)
        if not m:
            return []
        data = json.loads(m.group(0))
        rows: list[dict[str, Any]] = []
        for r in data:
            if not isinstance(r, dict):
                continue
            jb = str(r.get("jibun", "") or "").strip()
            if not jb:
                continue
            ot = str(r.get("owner_type", "") or "").strip()
            owner_type = (
                "국공유지" if ("국" in ot or "공" in ot)
                else ("사유지" if ot else "")
            )
            area = r.get("area_sqm")
            try:
                area = float(area) if area not in (None, "") else None
            except (TypeError, ValueError):
                area = None
            rows.append({
                "jibun": jb,
                "owner": str(r.get("owner", "") or "").strip(),
                "share": str(r.get("share", "") or "").strip(),
                "area_sqm": area,
                "owner_type": owner_type,
                "expected_price": None,
                "purchase_price": None,
                "contracted": False,
                "land_use_consent": False,
                "district_consent": False,
            })
        logger.info("토지조서 LLM 파싱 성공: %d행", len(rows))
        return rows
    except Exception as e:  # noqa: BLE001
        logger.warning("토지조서 LLM 파싱 실패: %s", str(e)[:160])
        return []


# ── 토지필지 종합분석 P0 — 견적·선별(무과금) ───────────────────────────
# ★★이 엔드포인트는 **과금하지 않는다.** 존재 이유가 "비용을 쓰기 전에 알려주는 것"이라
#   여기서 과금하면 목적과 정면으로 모순된다(견적을 보려고 돈을 내는 꼴).
#   대신 **인증은 요구**한다 — 요율·정책은 계정 컨텍스트에 딸린 정보다.
# ★계산은 순수함수(`parcel_survey_quote_service`)에 있고 이 라우터는 배선만 한다.

class ParcelSurveyQuoteRequest(BaseModel):
    """필지 목록. 각 행은 `pnu`/`address` 와 선택적 `has_building`·`geometry` 를 담는다.

    ★무과금이지만 상한을 건다 — `free_preview` 가 `build_parcel_graph` 를 태우고 그건
      O(V²) shapely 거리계산이다. **지갑이 아니라 CPU 를 막는 상한**이며, 형제 3경로가
      같은 값을 쓰게 해 "어느 게 진짜 상한인가"가 갈리지 않게 한다(D20 형제 스윕).
    """

    parcels: list[dict[str, Any]] = Field(default_factory=list, max_length=MAX_BULK_ITEMS)


@router.post(
    "/survey/quote",
    summary="토지필지 종합분석 견적·선별(무과금) — 발급 전 비용·판정가능성 안내",
)
async def parcel_survey_quote(
    req: ParcelSurveyQuoteRequest,
    current_user: CurrentUser = Depends(get_current_user),
) -> dict[str, Any]:
    """등기 발급 **전에** ①비용 ②등기 없이 아는 것 ③폴리곤 확보 여부를 돌려준다.

    ★견적은 **상한**이다 — 실제 청구는 발급에 성공한 건수(`issued_count`)만 집계한다.
      응답의 `billing_basis`·`note` 가 그 사실을 실어 나른다.
    """
    from app.services.land_intelligence.parcel_survey_quote_service import (
        free_preview,
        quote,
    )

    parcels = req.parcels or []
    result_quote = quote(parcels)

    # ── 성장루프 결속 ────────────────────────────────────────────────
    # ★P0 에서 **학습 가능한 신호**만 보낸다. 분류 결과는 아직 없다(P2) — 여기서 보낼 수 있는 건
    #   ①견적 규모 ②건축물 미상 비율 ③**폴리곤 미확보율**이다.
    #   ③이 특히 값지다: 이 비율이 높으면 제척 판정 자체가 불가하다는 뜻이라, 플랫폼이
    #   지적도 확보를 개선해야 한다는 신호가 된다(사용자 불만이 오기 전에).
    # ★식별자(pnu·주소)는 **보내지 않는다** — 집계에 불필요하고, 보내면 마스킹에 의존하게 된다.
    # ★best-effort: 적재 실패가 견적을 막으면 이 단계의 목적이 무너진다.
    try:
        from app.services.growth import capture_service

        # ★★도메인 메타는 반드시 `payload` **아래로** 넣는다.
        #   `capture_service._EVENT_COLS` 는 화이트리스트라 **평면 키를 조용히 버린다**
        #   (그 파일 주석: "그 외 키는 payload 로 흡수하지 않고 버림").
        #   실측 — 평면으로 보내면 적재 결과가 `{'event_type': 'parcel_survey_quote'}` 뿐이고
        #   도메인 필드 4개가 **전량 소실**된다. 즉 "성장루프에 실었다"가 거짓이 된다.
        #   ★이 규약은 형제 emitter 가 이미 주석으로 적어 둔 것이다
        #     (`design_ingest/orchestrator.py`·`ingest_service.py`) — 그걸 안 보고 재발시켰다.
        capture_service.record_event(
            "parcel_survey_quote",
            {
                "service": "parcel_survey",
                "payload": {
                    "parcel_count": result_quote["parcel_count"],
                    "building_unknown": result_quote["building_status"]["unknown"],
                    "geometry_missing": len(result_quote["geometry"]["missing"]),
                    "cost_max": result_quote["estimated_cost"]["max"],
                },
            },
        )
    except Exception:  # noqa: BLE001 — 성장루프 실패가 본기능을 막지 않는다
        logger.debug("성장루프 적재 스킵(견적 무손상)", exc_info=True)

    return {
        "quote": result_quote,
        # ★무료 미리보기는 **부가정보**다. 실패해도 견적(이 단계의 본질)은 나와야 하므로
        #   서비스 내부에서 이미 예외를 흡수하고 사유를 담아 돌려준다.
        "preview": free_preview(parcels),
    }


# ── 토지필지 종합분석 P2 — 매입전략 분류 ────────────────────────────────
# ★★이 엔드포인트의 존재 이유 하나는 **"소비처 0" 결함의 봉합**이다. P1
#   (`survey_selected_parcels`)은 머지된 뒤로 자기 테스트 밖에 호출부가 하나도 없었다 —
#   그런 코드는 배선을 끊어도 아무도 모른다. 그래서 여기서 **반드시 P1 을 태우고**,
#   그 카드를 P2 분류기(`build_strategy`)에 먹인다(우회 금지).
# ★계산은 전부 순수함수에 있고 이 라우터는 배선만 한다(/survey/quote 와 동일 패턴).

# ★★유료 엔드포인트의 **지출 상한**. 이 엔드포인트는 필지당 실제 발급(1,200원)+분석(2,000원)을
#   일으키므로, 입력 길이가 곧 청구액이다. 상한이 없으면 요청 하나로 임의 금액이 빠진다
#   (800필지 = 256만원). `MAX_PARCELS_FOR_GRAPH`(200)는 **그래프 연산 비용**을 막을 뿐
#   지갑을 막지 않는다 — 축이 다르므로 별도 상한을 둔다.
# ★경계는 양방향으로 건다(CLAUDE.md D19) — 상한만 걸면 반대쪽이 무제한이 된다.
#   ★자기적발: 최초 커밋은 이 줄에 "(min_length=1)" 이라고 **적어 놓고 실제로는 안 걸었다**.
#     실측 `parcels=[] → HTTP 200`. 주석이 없는 면역을 주장한 형태(C11)라 아래에 실제로 건다.
# ★형제와 **같은 상수**를 쓴다 — 값을 따로 적어 두면 한쪽만 바뀌었을 때
#   "어느 게 진짜 상한인가"를 아무도 모르게 된다(이 저장소가 반복해서 데인 형태).
MAX_STRATEGY_PARCELS = MAX_BULK_ITEMS
MIN_STRATEGY_PARCELS = MIN_BULK_ITEMS


class ParcelPurchaseStrategyRequest(BaseModel):
    """필지 목록 + 사업방식·기준일·주택건설대지면적.

    `scheme` 은 **기본값을 몰래 넣지 않는다** — 보유기간 10년 요건은 주택법 계열에만 있어
    방식이 없으면 판정 자체가 성립하지 않는다(미지정이면 미지정으로 판정 불가가 나온다).

    ★`parcels` 에 **상한·하한이 둘 다** 있다 — 유료 경로라 길이가 곧 청구액이다.
      초과하면 422 로 거부한다(조용히 잘라내면 사용자가 뺀 필지를 모른 채 결과를 신뢰한다).
      빈 목록도 422 다(빈 요청으로 판정표를 받는 것은 의미가 없고, 조용한 200 은 오독을 만든다).
    """

    parcels: list[dict[str, Any]] = Field(
        default_factory=list,
        min_length=MIN_STRATEGY_PARCELS,
        max_length=MAX_STRATEGY_PARCELS,
    )
    scheme: str | None = None
    district_plan_decision_date: str | None = None
    housing_site_area_sqm: float | None = None
    # ★파생형 락이 잡아낸 **내 필드**다 — 상한을 형제에만 걸고 자기 요청모델의 다른 리스트는
    #   빠뜨렸다. 제척 후보는 필지 부분집합이므로 필지 상한을 넘을 수 없다.
    exclusion_candidates: list[str] = Field(default_factory=list, max_length=MAX_BULK_ITEMS)


@router.post(
    "/survey/strategy",
    summary="토지필지 종합분석 P2 — 매입전략 분류(협의매수/매도청구/수용/제척검토/판정보류)",
)
@limiter.limit(ai_limiter)
async def parcel_purchase_strategy(
    request: Request,
    req: ParcelPurchaseStrategyRequest,
    current_user: CurrentUser = Depends(get_current_user),
) -> dict[str, Any]:
    """선택 필지를 **P1 로 발급·권리분석**한 뒤 매입전략으로 분류한다.

    ★발급은 건당 유료다(`registry_issue`+`registry_analysis`). P0(`/survey/quote`)에서 비용을
      확인한 뒤 진입하는 단계이며, 실제 청구는 **발급·분석에 성공한 건수만** 집계한다
      (`issued_count`/`analysis_charged` — 실패한 조회에는 청구하지 않는다).
    """
    from app.services.land_intelligence.parcel_purchase_strategy_service import (
        ACTION_UNDECIDED,
        build_strategy,
    )
    from app.services.land_intelligence.parcel_rights_survey_service import (
        survey_selected_parcels,
    )

    parcels = req.parcels or []

    # ★멱등 가드는 **발급이 일어나기 전**에 잡아야 한다 — 아래 P1 호출이 곧 외부 유료 발급이다.
    #   필지당 발급 1,200 + 분석 2,000 = 3,200원이라 중복 실행의 손해가 가장 크다.
    async with charge_once(
        request, endpoint="registry.survey_strategy", payload=req,
        tenant_id=current_user.tenant_id, user_id=current_user.user_id,
    ) as guard:
        # ① P1 을 **실제로** 태운다(우회하면 이 엔드포인트의 존재 이유가 사라진다).
        survey = await survey_selected_parcels(
            parcels,
            district_plan_decision_date=req.district_plan_decision_date,
            scheme=req.scheme,
        )

        # ② 과금 — P1 이 카드에 원본 분석 결과(`analysis`)를 그대로 실어 보내므로, 기존 라우터
        #    헬퍼가 그대로 소비한다(P1 서비스 자체는 과금하지 않는다 — 배선은 호출부의 몫).
        #    ★성공한 건만 청구한다: 두 헬퍼 모두 결과를 보고 판정한다(화이트리스트).
        #    ★같은 키로 이미 청구된 재요청이면 과금만 건너뛴다(결과는 새로 계산해 돌려준다).
        for card in (survey.get("cards") or []) if guard.billable else []:
            analysis = card.get("analysis")
            # ★이중 가드다(변이 감사 2026-08-15 — 이 조건을 무력화해도 결과가 같아 생존한다):
            #   `analysis` 가 None 이면 `issued_count`/`analysis_charged` 가 각각 0·False 를 내므로
            #   한 푼도 나가지 않는다. 그래도 남기는 이유는 **DB 세션을 열지 않기 위해서**다
            #   (조회 실패가 대부분인 상황에서 유의미하다 — `_charge_registry_issue` 와 같은 이유).
            if not analysis:
                continue
            await _charge_registry_issue(current_user.user_id, analysis, times=1)
            await _charge_registry_analysis(current_user.user_id, analysis)

    # ★★③ 과금이 끝나면 **원본 분석 블롭을 응답에서 걷어낸다.**
    #
    #   `_build_card` 는 P1 이 낸 `analysis`(RegistryAnalysisService 원본)를 카드에 통째로 싣는데,
    #   그 안에는 `pdf_url` 이 들어 있다 — `upload_registry_pdf(ttl_days=30)` 가 만든
    #   **30일짜리 서명 URL**이고 **인증이 걸려 있지 않다**(`services/storage_service.py`).
    #   즉 이 JSON 을 손에 넣은 사람은 누구나 30일간 등기부등본 전문을 내려받는다
    #   — 소유자 실명·지분·거래가액·근저당 채권최고액/근저당권자·압류권리자까지.
    #
    #   ★이 블롭은 **과금 외에는 아무도 쓰지 않는다**(위 루프가 유일한 소비처). 응답에 남길
    #     이유가 없고, 100필지면 페이로드가 수 MB 로 부푼다.
    #   ★카드가 이미 필요한 것만 추린 `registry` 블록을 따로 갖고 있으므로 표면 손실은 없다.
    #   ★순서 주의 — **반드시 과금 뒤**다. 앞에서 지우면 `issued_count` 가 셀 대상이 사라져
    #     발급 성공분이 청구되지 않는다(매출 누수). 두 방향 다 결함이라 순서가 계약이다.
    for card in survey.get("cards") or []:
        card.pop("analysis", None)

    # ③ 제척 위상판정용 인접 그래프 — 실패해도 분류(본기능)는 살아 있어야 한다.
    # ★이 초기화가 사라지면 `build_parcel_graph` 가 던졌을 때 `graph` 가 미정의라 500 이 난다
    #   (예외를 흡수한 의미가 사라진다). `test_그래프_계산이_던져도_분류는_살고...` 가 잠근다.
    graph: dict[str, Any] | None = None
    try:
        from app.services.zoning.parcel_graph import build_parcel_graph

        graph = build_parcel_graph(parcels)
    except Exception:  # noqa: BLE001 — 그래프 실패는 severability 가 '판정 불가'로 흡수한다
        # ★로그 문구는 동작이 아니다(변이 감사 생존 — 문자열 변경은 잡히지 않는다).
        #   **동작**의 잠금은 위 테스트가 담당한다: severable=None · 제척검토 미권고 · 200 응답.
        logger.warning("인접 그래프 계산 실패 — 제척 위상판정 미제공", exc_info=True)

    strategy = build_strategy(
        survey,
        parcels,
        scheme=req.scheme,
        housing_site_area_sqm=req.housing_site_area_sqm,
        graph=graph,
        exclusion_candidates=req.exclusion_candidates or None,
    )

    # ── 성장루프 결속(best-effort) ───────────────────────────────────────
    # ★식별자(pnu·주소)는 **보내지 않는다** — 집계에 불필요하고, 보내면 마스킹에 의존하게 된다.
    #   보낼 가치가 있는 것은 **플랫폼이 고쳐야 할 신호**다: 판정보류가 왜 나는지
    #   (확보율 미산정 · geometry 미확보)를 세면 데이터 확보 우선순위가 나온다.
    try:
        from app.services.growth import capture_service

        capture_service.record_event(
            "parcel_purchase_strategy",
            {
                "service": "parcel_survey",
                # ★★도메인 메타는 `payload` 아래로 — 평면 키는 화이트리스트에서 **버려진다**
                #   (`capture_service._EVENT_COLS`). 실측: 평면으로 보내면 적재 결과가
                #   `{'event_type': 'parcel_purchase_strategy'}` 뿐이고 아래 6개가 전량 소실됐다.
                "payload": {
                    "parcel_count": survey.get("parcel_count"),
                    "row_count": strategy["summary"]["row_count"],
                    # ★액션 라벨을 문자열로 다시 적으면 계약 상수와 갈라진다(상수 이름이 바뀌어도
                    #   이 줄은 조용히 0 을 세고, 성장루프는 "판정보류 없음"으로 오독한다).
                    "undecided_rows": strategy["summary"]["by_action"].get(ACTION_UNDECIDED, 0),
                    "secured_ratio_available": strategy["summary"]["secured_ratio_available"],
                    "geometry_unknown": strategy["summary"]["geometry_unknown_count"],
                    "scheme_provided": bool(req.scheme),
                    # ★`scheme_provided` 는 "문자열이 있었는가"만 본다 — **해석됐는가**는
                    #   별개다(미등록 사업방식은 문자열이 있어도 `governing_act=None` 이라
                    #   판정보류가 된다). 그 둘을 못 가르면 성장루프가 판정보류의 원인을
                    #   "사용자 미입력"으로 오귀속한다 → 해석 성공 여부를 따로 싣는다.
                    "scheme_resolved": strategy.get("legal", {}).get("governing_act") is not None,
                },
            },
        )
    except Exception:  # noqa: BLE001 — 성장루프 실패가 본기능을 막지 않는다
        logger.debug("성장루프 적재 스킵(분류 무손상)", exc_info=True)

    return {"survey": survey, "strategy": strategy}


# ── 권리분석 보고서 다운로드(PDF/PPTX/DOCX) ─────────────────────────────────
class RightsReportItem(BaseModel):
    """일괄 분석 결과 한 건. `result` 는 `/registry/analyze` 응답을 그대로 담는다."""

    jibun: str = Field("", description="지번(사람이 읽는 식별자)")
    result: dict[str, Any] | None = Field(None, description="analyze() 응답 원형")


# 한 번에 묶을 수 있는 필지 수 — **제품 상한**. 넘으면 **자르지 않고 거부**한다.
# 조용히 자르면 "전 필지 보고서"라는 이름으로 일부만 담긴 문서가 나간다.
_RIGHTS_REPORT_MAX = 300

# 파싱 단계 **안전 상한**(제품 상한과 다른 일을 한다). 제품 상한은 사람에게 "나눠서 받으세요"
# 라고 말하기 위한 것이고, 이쪽은 그 메시지에 닿기도 전에 **거대한 본문으로 서버를 태우는 것**을
# 막는다(리스트 필드는 모델 단계에서 상한을 갖는다 — 이 저장소의 계약이며 파생형 가드가 강제한다).
# 두 값이 같으면 한 층이 장식이 되므로 일부러 벌려 둔다: 301~1000 → 사람이 읽는 400,
# 1000 초과 → 파싱 단계에서 422.
_RIGHTS_REPORT_HARD_MAX = 1000


class RightsReportRequest(BaseModel):
    items: list[RightsReportItem] = Field(default_factory=list, max_length=_RIGHTS_REPORT_HARD_MAX)
    project_address: str | None = None
    format: str = Field("pdf", description="pdf | pptx | docx")


@router.post("/rights-report", summary="등기 권리분석 보고서 다운로드(PDF/PPTX/DOCX)")
async def registry_rights_report(
    req: RightsReportRequest,
    current_user: CurrentUser = Depends(get_current_user),
):
    """다필지 일괄 등기분석 결과를 **정본 보고서 엔진**으로 렌더해 내려보낸다.

    ★새로 조회하지 않는다. 호출측이 **이미 받은** 분석 결과를 형식화할 뿐이라 발급 과금이
    발생하지 않고, 새로 열리는 데이터 접근 권한도 없다(서식화 전용).

    ★미분석 필지를 빼지 않는다 — 어댑터가 §미분석 섹션으로 드러낸다. 빼면 보고서가
    "N필지 전부 안전"이라고 말하게 되고, 그건 없는 안전을 만드는 것이다.
    """
    import datetime as _dt

    if not req.items:
        raise HTTPException(status_code=400, detail="보고서로 만들 분석 결과가 없습니다.")
    if len(req.items) > _RIGHTS_REPORT_MAX:
        raise HTTPException(
            status_code=400,
            detail=(
                f"한 번에 {_RIGHTS_REPORT_MAX}필지까지 가능합니다(요청 {len(req.items)}필지). "
                "나눠서 내려받으세요 — 조용히 잘라 일부만 담긴 보고서를 만들지 않습니다."
            ),
        )

    from app.services.report.render import build_report_model_from_registry_rights, render_report

    model = build_report_model_from_registry_rights(
        [it.model_dump() for it in req.items],
        project_address=req.project_address,
        generated_at=_dt.datetime.now().strftime("%Y-%m-%d %H:%M"),
    )
    data, media_type, ext = render_report(model, req.format)
    logger.info(
        "권리분석 보고서 생성 user=%s 필지=%d fmt=%s bytes=%d",
        getattr(current_user, "id", None), len(req.items), ext, len(data),
    )
    return StreamingResponse(
        io.BytesIO(data),
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="propai_rights_report.{ext}"'},
    )


@router.post("/rights/ask", summary="권리분석 결과에 대한 추가 질의(LLM · 무과금)")
@limiter.limit(ai_limiter)
async def registry_rights_ask(
    request: Request,
    req: dict[str, Any],
    current_user: CurrentUser = Depends(get_current_user),
) -> dict[str, Any]:
    """이미 산출된 권리분석 JSON 에 대해 사용자가 원하는 추가 분석을 LLM 으로 받는다.

    ★★**등기부를 새로 발급하지 않는다.** 등기부는 1,200원/필지 유료이고, 저장소가
      *«실패를 캐시하기 싫으면 **파생물(해석)만** 재계산하라 — 원본을 다시 사지 마라»*
      를 규율로 남겼다. 추가질의는 **파생물 재계산**이므로 호출부가 **이미 가진**
      분석 JSON 만 받는다(`tests/test_registry_rights_interpreter_ssot.py` 가
      해석기의 임포트 그래프로 그 불가능성을 잠근다).

    Body:
        analysis: 권리분석 결과 JSON(`/registry/get-one` 등이 이미 돌려준 것)
        question: 사용자 질문(500자 상한, 초과분은 절단)

    Returns:
        {'ok', 'answer', 'basis', 'caveat'}  — 분석이 미완이면 answer='' + caveat 에 사유.
    """
    from app.services.ai.registry_rights_interpreter import RegistryRightsInterpreter

    analysis = req.get("analysis")
    if not isinstance(analysis, dict):
        return {"ok": False, "answer": "", "basis": "",
                "caveat": "권리분석 결과(analysis)가 필요합니다."}

    res = await RegistryRightsInterpreter().answer(analysis, str(req.get("question") or ""))
    return {"ok": bool(res.get("answer")), **res}
